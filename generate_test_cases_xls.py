import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

CLR = {
    "header_bg": "1F3864", "header_fg": "FFFFFF",
    "title_bg":  "2E75B6", "title_fg":  "FFFFFF",
    "smoke_bg":  "E2EFDA", "regr_bg":   "FFF2CC",
    "neg_bg":    "FCE4D6", "row_alt":   "F2F2F2",
    "row_white": "FFFFFF", "border_clr":"BFBFBF",
}

def mk_fill(h): return PatternFill("solid", fgColor=h)
def mk_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def mk_border():
    s = Side(style="thin", color=CLR["border_clr"])
    return Border(left=s, right=s, top=s, bottom=s)
def mk_align(wrap=True, h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

TYPE_FILL   = {"SMOKE": mk_fill(CLR["smoke_bg"]), "REGRESSION": mk_fill(CLR["regr_bg"]), "NEGATIVE": mk_fill(CLR["neg_bg"])}
STATUS_FILL = {"Active": mk_fill("C6EFCE"), "Skipped": mk_fill("FFEB9C"), "Pending": mk_fill("BDD7EE"), "Deprecated": mk_fill("F4CCCC")}

COLS       = ["TC ID", "Scenario", "Test Steps", "Expected Result", "Type", "Status"]
COL_WIDTHS = [15, 45, 65, 45, 14, 12]

# ── ALL TEST DATA ────────────────────────────────────────────────────────────
SECTIONS = [
  {
    "sheet": "UI Tests",
    "groups": [
      {
        "title": "1 · Login",
        "suite": "inventory-login-suite.xml  |  InventoryLoginTests.java",
        "rows": [
          ("UI-LOGIN-001","Successful login with valid credentials",
           "1. Clear browser cookies and session/local storage\n2. Launch the HLM Platform login page\n3. Verify login page heading 'Sign in to your account' is visible\n4. Enter valid admin email and password, click Sign In\n5. Verify Dashboard h1 heading is visible after redirect\n6. Verify 'Inventory & Assets' nav link is present in sidebar\n7. Verify welcome message contains 'Welcome back'",
           "User is redirected to the Dashboard; h1 header, nav link and welcome message are all visible","SMOKE","Active"),

          ("UI-LOGIN-002","Login fails with invalid credentials",
           "1. Clear session state and launch login page\n2. Enter invalid email and invalid password\n3. Click Sign In\n4. Verify red error banner (p.text-red-400) is visible\n5. Verify Sign In button is still present (no redirect occurred)",
           "Red error banner is displayed; user remains on the login page","NEGATIVE","Active"),

          ("UI-LOGIN-003","Login with empty email stays on login page",
           "1. Clear session state and launch login page\n2. Leave email field empty\n3. Enter valid password\n4. Click Sign In\n5. Verify Sign In button still visible (HTML5 blocked empty email)\n6. Verify login page heading is still present",
           "HTML5 validation blocks form submission; user remains on login page","NEGATIVE","Active"),

          ("UI-LOGIN-004","Login with empty password stays on login page",
           "1. Clear session state and launch login page\n2. Enter valid email address\n3. Leave password field empty\n4. Click Sign In\n5. Verify Sign In button still visible (no redirect)",
           "HTML5 validation blocks form submission; user remains on login page","NEGATIVE","Active"),

          ("UI-LOGIN-005","Login with both fields empty stays on login page",
           "1. Clear session state and launch login page\n2. Click Sign In without filling any credentials\n3. Verify Sign In button still visible\n4. Verify login page heading is still present",
           "HTML5 validation blocks empty form; user stays on login page","NEGATIVE","Active"),
        ],
      },
      {
        "title": "2 · Dashboard KPI Cards",
        "suite": "dashboard-kpi-suite.xml  |  DashboardKpiTests.java",
        "rows": [
          ("UI-KPI-001","Four KPI cards visible after login",
           "1. Re-navigate to Dashboard for a clean state\n2. Verify 'Welcome back' h2 heading is visible\n3. Verify Total Devices KPI card value element is visible\n4. Verify Active Deployments KPI card value element is visible\n5. Verify Pending Approvals KPI card value element is visible\n6. Verify Health Score KPI card value element is visible\n7. Verify no red error banner is present\n8. Verify no '—' loading placeholder remains",
           "All four KPI cards are visible, populated, and error-free","SMOKE","Active"),

          ("UI-KPI-002","Total Devices card has blue icon and numeric count",
           "1. Verify 'Total Devices' card label (div.text-sm) is visible\n2. Verify Total Devices numeric value (div.text-3xl) is visible\n3. Verify blue SVG Package icon (text-blue-* class) is present\n4. Verify no '—' loading placeholder remains",
           "Total Devices card shows label, blue icon, and a resolved numeric count","SMOKE","Active"),

          ("UI-KPI-003","Active Deployments card has green icon and in-progress count",
           "1. Verify 'Active Deployments' card label is visible\n2. Verify Active Deployments numeric value is visible\n3. Verify green SVG Download icon (text-green-* class) is present\n4. Verify no loading placeholder remains",
           "Active Deployments card shows label, green icon, and resolved count","SMOKE","Active"),

          ("UI-KPI-004","Pending Approvals card has orange icon and combined count",
           "1. Verify 'Pending Approvals' card label is visible\n2. Verify Pending Approvals numeric value (combined firmware + compliance) is visible\n3. Verify orange SVG Shield icon (text-orange-* class) is present\n4. Verify no loading placeholder remains",
           "Pending Approvals card shows label, orange icon, and aggregated count","SMOKE","Active"),

          ("UI-KPI-005","Health Score card has green icon and percentage format",
           "1. Verify 'Health Score' card label is visible\n2. Verify Health Score value contains '%' suffix (e.g., '77%')\n3. Verify green SVG Check-circle icon (text-green-* class) is present\n4. Verify no loading placeholder remains",
           "Health Score card shows label, green icon, and a value in N% format","SMOKE","Active"),

          ("UI-KPI-006","KPI cards show em-dash loading placeholder on page load",
           "1. Register route handler to delay all AppSync GraphQL responses by 2 s\n2. Re-navigate to Dashboard URL to trigger fresh page load\n3. Verify em-dash '—' (U+2014) is visible in at least one KPI card\n4. Unregister route handler (cleanup via finally block)",
           "At least one KPI card shows '—' placeholder while APIs are in-flight","REGRESSION","Active"),

          ("UI-KPI-007","KPI cards show zero fallback when no data exists",
           "1. Navigate to Dashboard, wait for live data to confirm loaded state\n2. Register route to return empty arrays for all fetch/xhr requests\n3. Click Refresh Dashboard to re-trigger all KPI APIs\n4. Verify no error banner is shown\n5. Verify Total Devices card shows '0'\n6. Verify Active Deployments card shows '0'\n7. Verify Pending Approvals card shows '0'\n8. Verify Health Score card shows '0%'\n9. Unregister route handler (cleanup)",
           "All KPI cards display '0' or '0%' with no error banner when APIs return empty data","REGRESSION","Active"),

          ("UI-KPI-008","KPI cards render in 4-column layout at desktop viewport",
           "1. Verify KPI grid container (div.grid) is visible at current desktop viewport\n2. Verify all four KPI card values are visible without horizontal scrolling\n3. Evaluate via JS whether grid-cols-4 Tailwind class is active on grid container\n4. Assert KPI grid container is present confirming 4-column layout",
           "All four cards visible in a single 4-column row at 1280px+ width","REGRESSION","Active"),

          ("UI-KPI-009","KPI cards render in 2x2 grid at tablet viewport",
           "1. Resize browser viewport to 768×1024 (tablet)\n2. Re-navigate to Dashboard to trigger reflow\n3. Verify Total Devices card is visible at tablet width\n4. Verify Active Deployments, Pending Approvals, Health Score cards are visible\n5. Restore viewport to 1620×1080 (desktop)",
           "Cards reflow into a 2×2 grid; all four remain visible at tablet width","REGRESSION","Active"),

          ("UI-KPI-010","KPI cards stack in single column at mobile viewport",
           "1. Resize browser viewport to 375×667 (mobile)\n2. Re-navigate to Dashboard to trigger reflow\n3. Verify Total Devices card is visible at mobile width\n4. Verify all four KPI cards are accessible without horizontal scrolling\n5. Restore viewport to 1620×1080 (desktop)",
           "All four cards stack vertically in a single column at 375px width","REGRESSION","Active"),

          ("UI-KPI-011","Refresh Dashboard button re-fetches all KPI data",
           "1. Verify Dashboard is loaded with live data (Total Devices card visible)\n2. Click the Refresh Dashboard button\n3. Verify all four KPI cards are populated after refresh\n4. Verify no error banner is present post-refresh",
           "All KPI cards update with fresh data after clicking Refresh Dashboard","SMOKE","Active"),

          ("UI-KPI-012","API failure shows red error banner and KPI fallback values",
           "1. Register route to return HTTP 500 from GraphQL endpoint\n2. Navigate to Dashboard with API failing\n3. Verify red error banner container (div.bg-red-50) is visible\n4. Verify error banner message is visible\n5. Unregister route handler (cleanup)",
           "Red error banner appears with message; KPI cards show fallback values","REGRESSION","Active"),

          ("UI-KPI-013","KPI cards render correctly in dark mode",
           "1. Activate dark mode on the Dashboard\n2. Verify all four KPI card value elements are visible\n3. Verify all four KPI card label elements are visible\n4. Verify all four KPI card icon elements are visible\n5. Verify no error banner is present in dark mode",
           "All KPI card icons, labels, and values render correctly with dark mode active","REGRESSION","Active"),

          ("UI-KPI-014","Health Score value conforms to rounded percentage format",
           "1. Verify Health Score KPI card value element is visible\n2. Assert value text matches pattern ^\\d{1,3}%$ (e.g., '0%', '77%', '100%')\n3. Verify no loading placeholder '—' remains",
           "Health Score displays as a rounded integer percentage with no decimals","REGRESSION","Active"),

          ("UI-KPI-015","KPI card icon colors match specification",
           "1. Verify Total Devices icon has a blue Tailwind class (text-blue-*)\n2. Verify Active Deployments icon has a green Tailwind class (text-green-*)\n3. Verify Pending Approvals icon has an orange Tailwind class (text-orange-*)\n4. Verify Health Score icon has a green Tailwind class (text-green-*)",
           "Each KPI card icon renders with the correct specification colour class","REGRESSION","Active"),

          ("UI-KPI-016","KPI values are never negative or invalid artifacts",
           "1. Read text content of Total Devices value element\n2. Assert value is a non-negative integer string\n3. Read Active Deployments, Pending Approvals values and assert ≥ 0\n4. Read Health Score value and assert it matches N% with 0 ≤ N ≤ 100\n5. Verify no value is 'NaN', 'undefined', 'null', or empty",
           "All KPI values are valid non-negative numbers; no JavaScript artifacts present","REGRESSION","Active"),

          ("UI-KPI-017","Unauthenticated access to dashboard redirects to login",
           "1. Clear all cookies and browser storage to remove auth session\n2. Navigate directly to the Dashboard URL\n3. Verify Sign In button is visible (redirected to login page)\n4. Verify login page heading is visible\n5. Verify Dashboard header is NOT present",
           "Unauthenticated navigation to dashboard redirects to login page","SMOKE","Active"),

          ("UI-KPI-018","Pending Approvals displays sum of firmware and compliance counts",
           "1. Intercept listFirmware and listCompliance GraphQL requests and capture responses\n2. Count records with status='Pending' from firmware API response\n3. Count records with status='Pending' from compliance API response\n4. Read Pending Approvals KPI card displayed value\n5. Assert displayed value = pending firmware count + pending compliance count",
           "Pending Approvals KPI equals the arithmetic sum of pending firmware + compliance records","REGRESSION","Active"),
        ],
      },
      {
        "title": "3 · Dashboard API Data Layer",
        "suite": "dashboard-api-suite.xml  |  DashboardApiTests.java",
        "rows": [
          ("UI-DAPI-001","Parallel API calls on dashboard load",
           "1. Navigate to Dashboard and trigger fresh page load\n2. Verify Total Devices KPI value is visible (devices API responded)\n3. Verify Active Deployments KPI value is visible (service orders API responded)\n4. Verify Pending Approvals KPI value is visible (firmware + compliance APIs responded)\n5. Verify no loading placeholder '—' remains",
           "All six data-domain APIs respond and their values are reflected across KPI cards","REGRESSION","Active"),

          ("UI-DAPI-002","All device records are fetched",
           "1. Verify Total Devices KPI card value element is present in the DOM\n2. Verify no loading placeholder '—' remains — data has fully loaded",
           "Total Devices KPI card displays a resolved numeric count","SMOKE","Active"),

          ("UI-DAPI-003","Offline devices are fetched separately",
           "1. Verify 'View Inventory' card is present in the Quick Actions section\n2. Verify orange offline-devices badge (span.bg-orange-500) is visible inside the 'View Inventory' card",
           "Offline device count badge is visible on the Quick Actions Inventory card","REGRESSION","Active"),

          ("UI-DAPI-004","In-progress service orders are fetched",
           "1. Verify Active Deployments KPI card value element is present\n2. Verify no loading placeholder '—' remains across all KPI cards",
           "Active Deployments KPI card shows a resolved count after APIs respond","SMOKE","Active"),

          ("UI-DAPI-005","Scheduled service orders are fetched",
           "1. Verify 'Scheduled Service' card is present in Quick Actions section\n2. Verify orange scheduled-orders badge (span.bg-orange-500) is visible inside the 'Scheduled Service' card",
           "Scheduled orders count badge is visible on the Quick Actions Scheduled Service card","REGRESSION","Active"),

          ("UI-DAPI-006","Pending firmware records are fetched",
           "1. Verify Pending Approvals KPI card value element is visible\n2. Verify 'Check Compliance' quick action link is visible",
           "Pending Approvals KPI includes pending firmware count; compliance link is present","SMOKE","Active"),

          ("UI-DAPI-007","Pending compliance records are fetched",
           "1. Verify Pending Approvals KPI card value element is visible (combined firmware + compliance)\n2. Verify no loading placeholder '—' remains in Pending Approvals card",
           "Pending Approvals KPI reflects both pending firmware and pending compliance counts","REGRESSION","Active"),

          ("UI-DAPI-008","KPI cards show loading placeholder on page load",
           "1. Register route to delay all GraphQL responses by 2 s\n2. Navigate directly to the Dashboard URL\n3. Verify '—' (em-dash, U+2014) is visible in at least one KPI card during fetch\n4. Unregister route handler (cleanup via finally block)",
           "Em-dash placeholder is visible in KPI cards while API responses are pending","REGRESSION","Active"),

          ("UI-DAPI-009","KPI cards transition from placeholder to live values",
           "1. Verify Total Devices KPI value is present (not placeholder)\n2. Verify Active Deployments KPI value is present\n3. Verify Pending Approvals KPI value is present\n4. Verify no loading placeholder '—' remains across any KPI card",
           "All KPI cards display live data values; no placeholder remains after load","SMOKE","Active"),

          ("UI-DAPI-010","Error banner displayed on single API failure",
           "1. Register route to return HTTP 500 from GraphQL endpoint\n2. Navigate to Dashboard with API returning 500\n3. Verify red error banner container (div.bg-red-50) is visible\n4. Verify error message text is visible within the banner\n5. Unregister route handler (cleanup)",
           "Red error banner with message appears when any API returns a server error","REGRESSION","Active"),

          ("UI-DAPI-011","Error banner and placeholders when all APIs fail",
           "1. Register route to abort all GraphQL requests (network-level failure)\n2. Navigate to Dashboard with all APIs unavailable\n3. Verify red error banner is visible\n4. Verify KPI cards show '0' fallback (app resolved error state)\n5. Verify Dashboard header is still rendered (no page crash)\n6. Unregister route handler (cleanup)",
           "Error banner shown; KPI cards show fallback '0'; dashboard does not crash","REGRESSION","Active"),

          ("UI-DAPI-012","Error banner displays meaningful message",
           "1. Register route to return HTTP 500 with specific error message payload\n2. Navigate to Dashboard with error route active\n3. Verify red error banner container is visible\n4. Verify error banner message element is visible and non-empty\n5. Unregister route handler (cleanup)",
           "Error banner contains a visible, non-empty, API-specific error message","REGRESSION","Active"),

          ("UI-DAPI-013","Audit logs fetched for alerts panel",
           "1. Register route to intercept listAuditLogs GraphQL query and inject a mock audit log entry\n2. Re-navigate to Dashboard so mock response is applied\n3. Verify Recent Alerts panel container is visible\n4. Verify at least one audit log list item is rendered in the panel\n5. Unregister route handler (cleanup)",
           "Recent Alerts panel is visible and shows at least one audit log entry","SMOKE","Active"),

          ("UI-DAPI-014","Audit logs 24-hour time window",
           "1. Register route to inject a mock audit log entry within last 24 hours\n2. Re-navigate to Dashboard so mock response is applied\n3. Verify Recent Alerts panel container is visible\n4. Verify alert items from within 24-hour window are displayed\n5. Verify 'No recent activity' message is NOT shown\n6. Unregister route handler (cleanup)",
           "Audit log entries from last 24 hours appear in the Alerts panel","REGRESSION","Active"),

          ("UI-DAPI-015","Alerts loading indicator displayed while fetching",
           "1. Register route to delay listAuditLogs GraphQL response by 2 s\n2. Re-navigate to Dashboard\n3. Verify loading indicator (spinner or skeleton) is visible in Alerts panel during fetch\n4. Unregister route handler (cleanup)",
           "A loading indicator is shown in the Alerts panel while audit log fetch is in-flight","REGRESSION","Active"),

          ("UI-DAPI-016","Alerts error message on API failure",
           "1. Register route to abort listAuditLogs GraphQL request\n2. Navigate to Dashboard with alerts API failing\n3. Verify Alerts panel container is visible\n4. Verify error message is displayed inside the Alerts panel\n5. Unregister route handler (cleanup)",
           "Alerts panel shows a descriptive error message when its API call fails","REGRESSION","Active"),

          ("UI-DAPI-017","No dashboard auto-refresh or polling",
           "1. Navigate to Dashboard and wait for full load\n2. Capture initial KPI values\n3. Wait 35 seconds without interaction\n4. Verify no additional API calls have been triggered\n5. Verify KPI values remain unchanged (no auto-refresh occurred)",
           "Dashboard data does not refresh automatically; values remain static until user action","REGRESSION","Active"),

          ("UI-DAPI-018","Device list is capped at 100 records",
           "1. Navigate to Dashboard and verify Total Devices KPI is visible\n2. Intercept the listDevices GraphQL request and inspect the limit variable in the request body\n3. Assert that limit ≤ 100 in the API request\n4. Verify Total Devices KPI value is ≤ 100",
           "The listDevices API call is sent with a limit of 100 or fewer records","REGRESSION","Active"),

          ("UI-DAPI-019","Dashboard loads with zero data",
           "1. Register route to return empty items arrays for all dashboard APIs\n2. Navigate to Dashboard\n3. Verify no red error banner appears\n4. Verify all four KPI cards show '0' or '0%'\n5. Verify dashboard header is still rendered\n6. Unregister route handler (cleanup)",
           "Dashboard renders cleanly with all-zero KPI values when APIs return empty datasets","REGRESSION","Active"),

          ("UI-DAPI-020","KPI and alerts fetch independently",
           "1. Register route to abort ONLY the listAuditLogs GraphQL request\n2. Navigate to Dashboard with alerts API failing\n3. Verify KPI cards (Total Devices, Active Deployments, Pending Approvals, Health Score) are populated\n4. Verify error message appears in Alerts panel only\n5. Unregister route handler (cleanup)",
           "KPI cards load successfully even when the Alerts panel API fails","REGRESSION","Active"),

          ("UI-DAPI-021","Refresh Dashboard button triggers complete data refetch",
           "1. Verify Dashboard is loaded with initial data\n2. Click the Refresh Dashboard button (circular arrow icon)\n3. Verify loading placeholders briefly appear in KPI cards\n4. Verify all KPI cards repopulate with fresh data\n5. Verify Alerts panel refreshes\n6. Verify no error banner is present post-refresh",
           "Clicking Refresh triggers re-fetch of all APIs; all panels update","SMOKE","Active"),

          ("UI-DAPI-022","Refresh button visible with circular arrow icon",
           "1. Verify Dashboard is loaded\n2. Verify the Refresh Dashboard button element is visible on the page\n3. Verify the button contains a circular arrow SVG icon",
           "Refresh Dashboard button is visible with a circular arrow icon","REGRESSION","Active"),

          ("UI-DAPI-023","KPI cards show placeholders during refresh",
           "1. Navigate to Dashboard and wait for full load\n2. Register route to delay GraphQL responses by 2 s\n3. Click the Refresh Dashboard button\n4. Immediately verify KPI cards show '—' loading placeholders\n5. Unregister route handler (cleanup)",
           "KPI cards revert to '—' loading state while a manual refresh is in progress","REGRESSION","Active"),

          ("UI-DAPI-024","Alerts panel shows loading state during refresh",
           "1. Navigate to Dashboard and wait for full load\n2. Register route to delay listAuditLogs response by 2 s\n3. Click the Refresh Dashboard button\n4. Verify loading indicator appears in Alerts panel during fetch\n5. Unregister route handler (cleanup)",
           "Alerts panel shows a loading indicator while a manual refresh is fetching audit logs","REGRESSION","Active"),

          ("UI-DAPI-025","Error banners cleared when refresh starts",
           "1. Register route to return HTTP 500 from GraphQL endpoint\n2. Navigate to Dashboard so error banner appears\n3. Unregister error route and register success route\n4. Click the Refresh Dashboard button\n5. Verify the error banner is cleared immediately when refresh starts\n6. Verify KPI cards repopulate with data after successful refresh",
           "Any existing error banners are dismissed when the Refresh Dashboard button is clicked","REGRESSION","Active"),

          ("UI-DAPI-026","System Status panel displays four services",
           "1. Verify Dashboard is loaded\n2. Verify System Status panel container is visible\n3. Verify exactly four service rows are rendered inside the panel\n4. Verify each row has a service name label and a status indicator",
           "System Status panel renders with exactly four service entries","SMOKE","Active"),

          ("UI-DAPI-027","System Status health percentages update from live KPI data",
           "1. Wait for Dashboard to fully load\n2. Read Total Devices and Health Score KPI values\n3. Read health percentage values displayed in the System Status panel\n4. Assert System Status health percentages correlate with live KPI data",
           "System Status health percentages are derived from and aligned with live KPI values","REGRESSION","Active"),

          ("UI-DAPI-028","System Status operational and degraded labels",
           "1. Wait for Dashboard to fully load\n2. Verify at least one service row with 100% health shows 'Operational' label\n3. Verify any service row below health threshold shows 'Degraded' label",
           "Service rows display 'Operational' or 'Degraded' labels based on health threshold","REGRESSION","Active"),

          ("UI-DAPI-029","Quick action badges hidden when count is zero",
           "1. Register route to return empty arrays for firmware and compliance APIs\n2. Click Refresh Dashboard\n3. Verify Quick Action Firmware badge is NOT visible\n4. Verify Quick Action Compliance badge is NOT visible\n5. Unregister route handler (cleanup)",
           "Quick Action count badges are hidden when the corresponding pending count is zero","REGRESSION","Active"),

          ("UI-DAPI-030","Quick action badges visible when count is greater than zero",
           "1. Verify Dashboard is loaded with live data\n2. Verify Pending Approvals KPI value is greater than zero\n3. Verify Quick Action badge for Firmware or Compliance is visible with a count",
           "Quick Action badges display numeric counts when pending items exist","SMOKE","Active"),

          ("UI-DAPI-031","No recent activity message when no audit logs exist",
           "1. Register route to return empty items for listAuditLogs query\n2. Re-navigate to Dashboard so empty audit logs response is applied\n3. Verify Alerts panel container is visible\n4. Verify 'No recent activity' (or equivalent empty-state) message is displayed\n5. Verify no audit log list items are rendered\n6. Unregister route handler (cleanup)",
           "Alerts panel shows 'No recent activity' empty-state when audit log API returns no items","REGRESSION","Active"),
        ],
      },
      {
        "title": "4 · Geo Location Map",
        "suite": "geo-location-suite.xml  |  GeoLocationTests.java",
        "rows": [
          ("UI-GEO-001","Geo Location tab renders interactive MapLibre map",
           "1. Navigate to Dashboard → Inventory module via sidebar\n2. Wait for Inventory page heading to appear\n3. Click the Geo Location tab\n4. Wait for MapLibre map canvas and device pins to render\n5. Verify div.maplibregl-map container is present\n6. Verify HTML5 canvas element is rendered inside map container\n7. Verify [aria-label='Map'] region is present\n8. Verify legacy static SVG map element is NOT present\n9. Verify at least one .maplibregl-marker pin is visible\n10. Verify MapLibre attribution link is visible",
           "Interactive MapLibre map renders; canvas and markers present; no legacy SVG map","SMOKE","Active"),

          ("UI-GEO-002","Map legend displays correct status entries",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'Online' text entry is visible in the legend panel\n3. Verify 'Offline' text entry is visible in the legend panel\n4. Verify 'Maintenance' text entry is visible in the legend panel",
           "Map legend shows Online, Offline, and Maintenance entries with colour indicators","SMOKE","Active"),

          ("UI-GEO-003","All twelve pins visible with correct color distribution",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'All' filter button is active by default\n3. Verify at least one map marker is visible\n4. Evaluate via JS: count background-color of all pin dot divs\n5. Assert total pin count = 12\n6. Assert 6 pins have green rgb(34,197,94), 3 red rgb(239,68,68), 3 orange rgb(249,115,22)",
           "12 pins are rendered; JS color evaluation confirms 6 green, 3 red, 3 orange distribution","SMOKE","Active"),

          ("UI-GEO-004","Online filter shows six green pins",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click 'Online(6)' filter button\n3. Verify Online filter button is active/highlighted\n4. Wait for filtered pins to appear in DOM\n5. Evaluate via JS: count all map marker elements\n6. Assert total visible pins = 6\n7. Evaluate via JS: count pins with green background rgb(34,197,94)\n8. Assert green pin count = 6",
           "Exactly 6 green pins visible after Online filter; all have correct green colour","REGRESSION","Active"),

          ("UI-GEO-005","Offline filter shows three red pins",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click 'Offline(3)' filter button\n3. Wait for filtered pins to appear in DOM\n4. Evaluate via JS: count all map marker elements\n5. Assert total visible pins = 3\n6. Evaluate via JS: count pins with red background rgb(239,68,68)\n7. Assert red pin count = 3",
           "Exactly 3 red pins visible after Offline filter; all have correct red colour","REGRESSION","Active"),

          ("UI-GEO-006","Maintenance filter shows three orange pins",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click 'Maintenance(3)' filter button\n3. Wait for filtered pins to appear in DOM\n4. Evaluate via JS: count all map marker elements\n5. Assert total visible pins = 3\n6. Evaluate via JS: count pins with orange background rgb(249,115,22)\n7. Assert orange pin count = 3",
           "Exactly 3 orange pins visible after Maintenance filter; all have correct orange colour","REGRESSION","Active"),

          ("UI-GEO-007","Stat pills display correct device counts",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'Total' stat pill is visible\n3. Verify 'Online' stat pill is visible\n4. Verify 'Offline' stat pill is visible\n5. Verify 'Maintenance' stat pill is visible\n6. Verify Online filter button label contains '6'\n7. Verify Offline filter button label contains '3'\n8. Verify Maintenance filter button label contains '3'",
           "All four stat pills visible; counts match Online(6), Offline(3), Maintenance(3)","SMOKE","Active"),

          ("UI-GEO-008","All four filter buttons are present",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'All' filter button is visible and active by default\n3. Verify 'Online(N)' filter button is visible with count suffix\n4. Verify 'Offline(N)' filter button is visible with count suffix\n5. Verify 'Maintenance(N)' filter button is visible with count suffix",
           "All four filter buttons (All, Online(6), Offline(3), Maintenance(3)) are present","SMOKE","Active"),

          ("UI-GEO-009","Filter sequence updates map markers correctly",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Assert initial pin count = 12 (All filter default)\n3. Click Online filter → wait → assert pin count = 6\n4. Click Offline filter → wait → assert pin count = 3\n5. Click Maintenance filter → wait → assert pin count = 3\n6. Click All filter → wait → assert pin count = 12",
           "Pin counts change correctly at each filter step: 12→6→3→3→12","REGRESSION","Active"),

          ("UI-GEO-010","Maintenance pin opens detail card with correct fields",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click Maintenance filter to show only orange pins\n3. Click on the first Maintenance pin (force:true via browser.evaluate)\n4. Wait for device detail card to appear\n5. Verify detail card container is visible\n6. Verify device name, status, address, coordinates, and last-seen fields are present",
           "Clicking a Maintenance pin opens a detail card with all required device fields","REGRESSION","Active"),

          ("UI-GEO-011","Device detail card closes on X button click",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click on any map marker to open the device detail card\n3. Verify detail card is visible\n4. Click the X (close) button on the detail card\n5. Verify detail card container is no longer visible",
           "Device detail card is dismissed when the X button is clicked","SMOKE","Active"),

          ("UI-GEO-012","Zoom in and zoom out controls are functional",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Evaluate via JS: get current map zoom level\n3. Click the Zoom In (+) control button\n4. Wait briefly for zoom animation\n5. Evaluate via JS: assert new zoom level > initial zoom level\n6. Click the Zoom Out (-) control button\n7. Evaluate via JS: assert zoom level decreased",
           "Zoom In increases map zoom level; Zoom Out decreases it","REGRESSION","Active"),

          ("UI-GEO-013","Compass button is present and interactive",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify compass/bearing reset button element is visible\n3. Click the compass button\n4. Verify no error is thrown and map remains interactive",
           "Compass button is visible and responds to user clicks without error","REGRESSION","Active"),

          ("UI-GEO-014","Hardware Inventory tab unaffected by PS20",
           "1. Navigate to Inventory module (via BeforeMethod)\n2. Click the Hardware Inventory tab\n3. Verify Hardware Inventory tab content is visible\n4. Verify no MapLibre elements or Geo Location UI are leaked into Hardware Inventory view\n5. Verify existing Hardware Inventory table/content renders correctly",
           "Hardware Inventory tab content is unchanged and unaffected by PS20 Geo Location changes","REGRESSION","Active"),

          ("UI-GEO-015","Firmware Status tab unaffected by PS20",
           "1. Navigate to Inventory module (via BeforeMethod)\n2. Click the Firmware Status tab\n3. Verify Firmware Status tab content is visible\n4. Verify no Geo Location UI elements are present in Firmware Status view\n5. Verify existing Firmware Status content renders correctly",
           "Firmware Status tab content is unchanged and unaffected by PS20 Geo Location changes","REGRESSION","Active"),

          ("UI-GEO-016","Stat pill counts are static across filter changes",
           "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Read Online stat pill count value (should be 6)\n3. Click Online filter button\n4. Assert Online stat pill count is still 6 (unchanged)\n5. Click Offline filter button\n6. Assert Offline stat pill count is still 3 (unchanged)\n7. Click Maintenance filter button\n8. Assert Maintenance stat pill count is still 3 (unchanged)\n9. Click All filter and verify counts unchanged",
           "Stat pill counts remain fixed at 6/3/3 regardless of which filter is active","REGRESSION","Active"),
        ],
      },
    ],
  },
  {
    "sheet": "API Tests",
    "groups": [
      {
        "title": "5 · Devices",
        "suite": "specs/api/device.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-DEV-001","listDevices returns paginated response with items",
           "1. Send authenticated GraphQL query: listDevices (no filter)\n2. Assert HTTP 200 status code\n3. Assert no errors in response\n4. Assert listDevices result object is defined\n5. Parse items array and assert length > 0",
           "Response is HTTP 200 with a populated items array","SMOKE","Active"),

          ("API-DEV-002","listDevices with limit returns at most N items",
           "1. Send authenticated GraphQL query: listDevices with limit=3\n2. Assert HTTP 200 status code\n3. Parse items array\n4. Assert items.length ≤ 3",
           "Response contains at most 3 items when limit=3 is specified","REGRESSION","Active"),

          ("API-DEV-003","listDevices filters by status=Online",
           "1. Send authenticated GraphQL query: listDevices with status='Online'\n2. Assert HTTP 200 status code\n3. Parse items array\n4. Assert every item has status === 'Online'",
           "All returned items have status='Online' when filter is applied","REGRESSION","Active"),

          ("API-DEV-004","listDevices filters by status=Offline",
           "1. Send authenticated GraphQL query: listDevices with status='Offline'\n2. Assert HTTP 200 status code\n3. Parse items array\n4. Assert every item has status === 'Offline'",
           "All returned items have status='Offline' when filter is applied","REGRESSION","Active"),

          ("API-DEV-005","listDevices filters by status=Maintenance",
           "1. Send authenticated GraphQL query: listDevices with status='Maintenance'\n2. Assert HTTP 200 status code\n3. Parse items array\n4. Assert every item has status === 'Maintenance'",
           "All returned items have status='Maintenance' when filter is applied","REGRESSION","Active"),

          ("API-DEV-006","listDevices response items have required fields",
           "1. Send authenticated GraphQL query: listDevices with limit=1\n2. Assert HTTP 200 status code\n3. Parse first item from response\n4. Assert item.id, item.deviceName, item.serialNumber, item.model, item.status, item.location are all present",
           "Each device item contains all required fields: id, deviceName, serialNumber, model, status, location","SMOKE","Active"),

          ("API-DEV-007","listDevices totalCount is a non-negative integer",
           "1. Send authenticated GraphQL query: listDevices\n2. Assert HTTP 200 status code\n3. Read totalCount from response\n4. Assert totalCount is a number and totalCount ≥ 0",
           "totalCount field is present and is a non-negative integer","REGRESSION","Active"),

          ("API-DEV-008","getDevice returns device by valid ID",
           "1. Seed firstDeviceId in beforeAll by calling listDevices\n2. Send authenticated GraphQL query: getDevice(id: firstDeviceId)\n3. Assert HTTP 200 status code\n4. Assert no errors in response\n5. Assert returned device.id === firstDeviceId",
           "getDevice returns the complete device record for the queried ID","SMOKE","Active"),

          ("API-DEV-009","getDevicesByCustomer returns devices for a customer",
           "1. Seed firstCustomerId in beforeAll\n2. Send authenticated GraphQL query: getDevicesByCustomer(customerId: firstCustomerId)\n3. Assert HTTP 200 status code\n4. Parse items array and assert length > 0\n5. Assert all returned items belong to the queried customer",
           "All devices belonging to the specified customer are returned","REGRESSION","Active"),

          ("API-DEV-010","getDevicesByLocation returns devices for a location",
           "1. Seed firstLocation in beforeAll\n2. Send authenticated GraphQL query: getDevicesByLocation(location: firstLocation)\n3. Assert HTTP 200 status code\n4. Parse items array and assert length > 0",
           "All devices at the specified location are returned","REGRESSION","Active"),

          ("API-DEV-011","getDevice with non-existent ID returns null or error",
           "1. Send authenticated GraphQL query: getDevice(id: 'non-existent-id-999')\n2. Assert HTTP 200 status code\n3. Assert response.data.getDevice is null OR response.errors is defined",
           "API returns null data or a descriptive error for an unknown device ID","NEGATIVE","Active"),

          ("API-DEV-012","listDevices with invalid status returns empty items or error",
           "1. Send authenticated GraphQL query: listDevices(status: 'INVALID_STATUS')\n2. Assert HTTP 200 status code\n3. Assert items array is empty OR response contains a validation error",
           "Invalid status filter returns empty items or a GraphQL validation error","NEGATIVE","Active"),

          ("API-DEV-013","getDevicesByCustomer with non-existent customer returns empty items",
           "1. Send authenticated GraphQL query: getDevicesByCustomer(customerId: 'non-existent-cust-999')\n2. Assert HTTP 200 status code\n3. Parse items array and assert length === 0",
           "Empty items array returned for an unknown customer ID","NEGATIVE","Active"),

          ("API-DEV-014","getDevicesByLocation with unknown location returns empty items",
           "1. Send authenticated GraphQL query: getDevicesByLocation(location: 'UNKNOWN_LOCATION_XYZ')\n2. Assert HTTP 200 status code\n3. Parse items array and assert length === 0",
           "Empty items array returned for an unknown location","NEGATIVE","Active"),

          ("API-DEV-015","Unauthenticated listDevices request returns 401 or auth error",
           "1. Build GraphQL request for listDevices WITHOUT authentication token\n2. Send unauthenticated request to the GraphQL endpoint\n3. Assert HTTP status is 401 OR response contains an authentication/authorization error",
           "Unauthenticated request is rejected with 401 or an auth error","NEGATIVE","Active"),
        ],
      },
      {
        "title": "6 · Service Orders",
        "suite": "specs/api/service-orders.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-SO-001","listServiceOrdersByStatus returns Pending orders",
           "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: 'Pending')\n2. Assert HTTP 200 status code\n3. Assert no errors\n4. Parse items array and assert length > 0\n5. Assert all items have status === 'Pending'",
           "Only Pending service orders are returned","SMOKE","Active"),

          ("API-SO-002","listServiceOrdersByStatus returns InProgress orders",
           "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: 'In Progress')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'In Progress'",
           "Only In Progress service orders are returned","REGRESSION","Active"),

          ("API-SO-003","listServiceOrdersByStatus returns Completed orders",
           "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: 'Completed')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Completed'",
           "Only Completed service orders are returned","REGRESSION","Active"),

          ("API-SO-004","listServiceOrdersByStatus returns Scheduled orders",
           "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: 'Scheduled')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Scheduled'",
           "Only Scheduled service orders are returned","REGRESSION","Active"),

          ("API-SO-005","listServiceOrdersByStatus response items have required fields",
           "1. Send authenticated GraphQL query: listServiceOrdersByStatus with any valid status\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.title, item.status, item.technicianId, item.scheduledDate are all present",
           "Each service order item contains all required fields","SMOKE","Active"),

          ("API-SO-006","listServiceOrdersByDate returns results within last 30 days",
           "1. Compute startDate = now - 30 days, endDate = now (ISO format)\n2. Send authenticated GraphQL query: listServiceOrdersByDate(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert all scheduledDate values fall within the 30-day range",
           "All returned orders have scheduledDate within the last 30 days","REGRESSION","Active"),

          ("API-SO-007","listServiceOrdersByDate with future range returns empty items",
           "1. Compute startDate = now + 1 year, endDate = now + 2 years\n2. Send authenticated GraphQL query: listServiceOrdersByDate(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert length === 0",
           "Empty items array returned for a fully future date range","NEGATIVE","Active"),

          ("API-SO-008","getServiceOrder returns order by valid ID",
           "1. Seed a valid service order ID from listServiceOrdersByStatus\n2. Send authenticated GraphQL query: getServiceOrder(id: seededId)\n3. Assert HTTP 200\n4. Assert returned order.id === seededId",
           "Complete service order record returned for the queried ID","SMOKE","Active"),

          ("API-SO-009","getServiceOrdersByTechnician returns orders for a technician",
           "1. Seed a valid technicianId from listServiceOrdersByStatus\n2. Send GraphQL query: getServiceOrdersByTechnician(technicianId: seededId)\n3. Assert HTTP 200\n4. Parse items and assert length > 0",
           "All orders for the specified technician are returned","REGRESSION","Active"),

          ("API-SO-010","getServiceOrder with non-existent ID returns null or error",
           "1. Send authenticated GraphQL query: getServiceOrder(id: 'non-existent-so-999')\n2. Assert HTTP 200\n3. Assert response.data.getServiceOrder is null OR response.errors is defined",
           "null data or descriptive error returned for unknown service order ID","NEGATIVE","Active"),

          ("API-SO-011","listServiceOrdersByStatus with invalid status returns empty or error",
           "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: 'FAKE_STATUS')\n2. Assert HTTP 200\n3. Assert items array is empty OR response.errors is defined",
           "Invalid status returns empty items or a validation error","NEGATIVE","Active"),

          ("API-SO-012","getServiceOrdersByTechnician with unknown ID returns empty items",
           "1. Send GraphQL query: getServiceOrdersByTechnician(technicianId: 'unknown-tech-999')\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
           "Empty items array returned for unknown technician ID","NEGATIVE","Active"),

          ("API-SO-013","listServiceOrdersByDate with endDate before startDate returns empty or error",
           "1. Set startDate = now, endDate = now - 1 day\n2. Send authenticated GraphQL query: listServiceOrdersByDate(startDate, endDate)\n3. Assert HTTP 200\n4. Assert items is empty OR response.errors is defined",
           "Reversed date range returns empty results or a validation error","NEGATIVE","Active"),
        ],
      },
      {
        "title": "7 · Firmware",
        "suite": "specs/api/firmware.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-FW-001","listFirmware returns all firmware with expected fields",
           "1. Send authenticated GraphQL query: listFirmware (no filter)\n2. Assert HTTP 200\n3. Assert no errors in response\n4. Parse items and assert length > 0",
           "Response is HTTP 200 with a populated firmware items array","SMOKE","Active"),

          ("API-FW-002","listFirmware response items have required fields",
           "1. Send authenticated GraphQL query: listFirmware with limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.name, item.version, item.status, item.deviceModel are all present and non-null",
           "Each firmware item contains all required fields","SMOKE","Active"),

          ("API-FW-003","listFirmware filters by status=Pending",
           "1. Send authenticated GraphQL query: listFirmware(status: 'Pending')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Pending'",
           "Only Pending firmware records are returned","REGRESSION","Active"),

          ("API-FW-004","listFirmware filters by status=Approved",
           "1. Send authenticated GraphQL query: listFirmware(status: 'Approved')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Approved'",
           "Only Approved firmware records are returned","REGRESSION","Active"),

          ("API-FW-005","listFirmware filters by status=Deprecated",
           "1. Send authenticated GraphQL query: listFirmware(status: 'Deprecated')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Deprecated'",
           "Only Deprecated firmware records are returned","REGRESSION","Active"),

          ("API-FW-006","listFirmware with limit=2 returns at most 2 items",
           "1. Send authenticated GraphQL query: listFirmware(limit: 2)\n2. Assert HTTP 200\n3. Parse items array\n4. Assert items.length ≤ 2",
           "Response contains at most 2 items when limit=2 is specified","REGRESSION","Active"),

          ("API-FW-007","listFirmware pagination: nextToken advances the cursor",
           "1. Send listFirmware(limit: 2) and capture nextToken from response\n2. Assert nextToken is defined and non-null\n3. Send listFirmware(limit: 2, nextToken: capturedToken)\n4. Assert HTTP 200 and response contains next page of items",
           "nextToken from page 1 correctly advances to page 2 with different items","REGRESSION","Active"),

          ("API-FW-008","getFirmware returns firmware by valid ID",
           "1. Seed a valid firmware ID from listFirmware in beforeAll\n2. Send authenticated GraphQL query: getFirmware(id: seededId)\n3. Assert HTTP 200\n4. Assert returned firmware.id === seededId",
           "Complete firmware record returned for the queried ID","SMOKE","Active"),

          ("API-FW-009","getFirmwareByModel returns firmware for a known model",
           "1. Seed a valid deviceModel from listFirmware in beforeAll\n2. Send authenticated GraphQL query: getFirmwareByModel(deviceModel: seededModel)\n3. Assert HTTP 200\n4. Parse items and assert length > 0",
           "Firmware records for the specified model are returned","REGRESSION","Active"),

          ("API-FW-010","getFirmwareWithRelations returns data for valid ID",
           "1. Seed a valid firmware ID from beforeAll\n2. Send authenticated GraphQL query: getFirmwareWithRelations(id: seededId)\n3. Assert HTTP 200\n4. Assert returned data is non-null and includes related entity fields",
           "Firmware record with related entities is returned for the queried ID","REGRESSION","Active"),

          ("API-FW-011","getFirmware with non-existent ID returns null or error",
           "1. Send authenticated GraphQL query: getFirmware(id: 'non-existent-fw-999')\n2. Assert HTTP 200\n3. Assert response.data.getFirmware is null OR response.errors is defined",
           "null data or descriptive error returned for unknown firmware ID","NEGATIVE","Active"),

          ("API-FW-012","getFirmwareByModel with unknown model returns empty items",
           "1. Send authenticated GraphQL query: getFirmwareByModel(deviceModel: 'UNKNOWN_MODEL_XYZ')\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
           "Empty items array returned for an unrecognised device model","NEGATIVE","Active"),

          ("API-FW-013","listFirmware with invalid status returns empty or error",
           "1. Send authenticated GraphQL query: listFirmware(status: 'INVALID_STATUS')\n2. Assert HTTP 200\n3. Assert items is empty OR response.errors is defined",
           "Invalid status returns empty items or a validation error","NEGATIVE","Active"),

          ("API-FW-014","getFirmwareWithRelations with non-existent ID returns null or error",
           "1. Send authenticated GraphQL query: getFirmwareWithRelations(id: 'non-existent-fw-999')\n2. Assert HTTP 200\n3. Assert data is null OR response.errors is defined",
           "null or error returned for unknown firmware ID with relations query","NEGATIVE","Active"),

          ("API-FW-015","Unauthenticated listFirmware request is rejected",
           "1. Build GraphQL request for listFirmware WITHOUT authentication token\n2. Send unauthenticated request to the GraphQL endpoint\n3. Assert HTTP 401 OR response contains authentication error",
           "Unauthenticated request is rejected with 401 or an auth error","NEGATIVE","Active"),
        ],
      },
      {
        "title": "8 · Compliance",
        "suite": "specs/api/compliance.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-COMP-001","listComplianceByStatus Approved returns only approved records",
           "1. Send authenticated GraphQL query: listComplianceByStatus(status: 'Approved')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Approved'",
           "Only Approved compliance records are returned","SMOKE","Active"),

          ("API-COMP-002","listComplianceByStatus Pending returns only pending records",
           "1. Send authenticated GraphQL query: listComplianceByStatus(status: 'Pending')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Pending'",
           "Only Pending compliance records are returned","SMOKE","Active"),

          ("API-COMP-003","listComplianceByStatus Deprecated returns only deprecated records",
           "1. Send authenticated GraphQL query: listComplianceByStatus(status: 'Deprecated')\n2. Assert HTTP 200\n3. Parse items and assert all have status === 'Deprecated'",
           "Only Deprecated compliance records are returned","REGRESSION","Active"),

          ("API-COMP-004","listComplianceByStatus response items have required fields",
           "1. Send authenticated GraphQL query: listComplianceByStatus with any valid status, limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.firmwareId, item.status, item.certifications are all present",
           "Each compliance item contains all required fields","SMOKE","Active"),

          ("API-COMP-005","getCompliance returns record by valid ID",
           "1. Seed a valid compliance ID from listComplianceByStatus in beforeAll\n2. Send authenticated GraphQL query: getCompliance(id: seededId)\n3. Assert HTTP 200\n4. Assert returned record.id === seededId",
           "Complete compliance record returned for the queried ID","SMOKE","Active"),

          ("API-COMP-006","getComplianceByCertification returns records for CE certification",
           "1. Send authenticated GraphQL query: getComplianceByCertification(certification: 'CE')\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
           "Compliance records with CE certification are returned","REGRESSION","Active"),

          ("API-COMP-007","getComplianceByCertification returns records for FCC certification",
           "1. Send authenticated GraphQL query: getComplianceByCertification(certification: 'FCC')\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
           "Compliance records with FCC certification are returned","REGRESSION","Active"),

          ("API-COMP-008","getComplianceByCertification returns records for UL certification",
           "1. Send authenticated GraphQL query: getComplianceByCertification(certification: 'UL')\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
           "Compliance records with UL certification are returned","REGRESSION","Active"),

          ("API-COMP-009","getComplianceByCertification returned items contain the requested cert",
           "1. Send authenticated GraphQL query: getComplianceByCertification(certification: 'CE')\n2. Assert HTTP 200\n3. Parse items\n4. Assert every item's certifications array includes 'CE'",
           "All returned items contain the requested certification in their certifications array","REGRESSION","Active"),

          ("API-COMP-010","getCompliance with non-existent ID returns null or error",
           "1. Send authenticated GraphQL query: getCompliance(id: 'non-existent-comp-999')\n2. Assert HTTP 200\n3. Assert response.data.getCompliance is null OR response.errors is defined",
           "null data or descriptive error returned for unknown compliance ID","NEGATIVE","Active"),

          ("API-COMP-011","listComplianceByStatus with invalid status returns empty or error",
           "1. Send authenticated GraphQL query: listComplianceByStatus(status: 'INVALID_STATUS')\n2. Assert HTTP 200\n3. Assert items is empty OR response.errors is defined",
           "Invalid status returns empty items or a validation error","NEGATIVE","Active"),

          ("API-COMP-012","getComplianceByCertification with unknown certification returns empty",
           "1. Send authenticated GraphQL query: getComplianceByCertification(certification: 'UNKNOWN_CERT')\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
           "Empty items array returned for an unrecognised certification","NEGATIVE","Active"),
        ],
      },
      {
        "title": "9 · Audit Logs",
        "suite": "specs/api/audit-logs.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-AUDIT-001","listAuditLogs returns records for last 24 hours",
           "1. Compute startDate = now - 24h, endDate = now (ISO format)\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Assert no errors\n5. Assert listAuditLogs result and items are defined",
           "Response is HTTP 200 with items defined for the 24-hour window","SMOKE","Active"),

          ("API-AUDIT-002","listAuditLogs returns records for last 7 days",
           "1. Compute startDate = now - 7 days, endDate = now\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Assert no errors and items are defined",
           "Response is HTTP 200 with items defined for the 7-day window","REGRESSION","Active"),

          ("API-AUDIT-003","listAuditLogs returns records for last 30 days",
           "1. Compute startDate = now - 30 days, endDate = now\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert length > 0",
           "Audit log records exist and are returned for the 30-day window","REGRESSION","Active"),

          ("API-AUDIT-004","listAuditLogs response items have required fields",
           "1. Send authenticated GraphQL query: listAuditLogs for last 30 days with limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id is truthy\n5. Assert item.action is truthy\n6. Assert at least one timestamp field exists (timestamp/createdAt/eventTime/date)",
           "Each audit log item contains id, action, and a timestamp field","SMOKE","Active"),

          ("API-AUDIT-005","listAuditLogs with limit=5 returns at most 5 items",
           "1. Send authenticated GraphQL query: listAuditLogs for last 30 days with limit=5\n2. Assert HTTP 200\n3. Parse items array\n4. Assert items.length ≤ 5",
           "Response contains at most 5 items when limit=5 is specified","REGRESSION","Active"),

          ("API-AUDIT-006","listAuditLogs timestamps are within requested range",
           "1. Compute startDate = now - 7 days, endDate = now\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Parse items\n4. For each item assert its timestamp >= startDate and <= endDate",
           "All returned audit log timestamps fall within the specified date range","REGRESSION","Active"),

          ("API-AUDIT-007","listAuditLogs totalCount is returned",
           "1. Send authenticated GraphQL query: listAuditLogs for last 30 days\n2. Assert HTTP 200\n3. Assert response.data.listAuditLogs.totalCount is a number ≥ 0",
           "totalCount field is present and is a non-negative integer","REGRESSION","Active"),

          ("API-AUDIT-008","getAuditLogsByUser returns logs for a known user",
           "1. Seed existingUserId from listAuditLogs in beforeAll\n2. Send authenticated GraphQL query: getAuditLogsByUser(userId: existingUserId)\n3. Assert HTTP 200\n4. Parse items and assert at least one item is returned",
           "Audit log entries for the specified user are returned","SMOKE","Active"),

          ("API-AUDIT-009","listAuditLogs returns empty for future date range",
           "1. Compute startDate = now + 1 year, endDate = now + 2 years\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert length === 0",
           "Empty items array returned for a fully future date range","NEGATIVE","Active"),

          ("API-AUDIT-010","listAuditLogs with reversed date range returns empty or error",
           "1. Set startDate = now, endDate = now - 1 day\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Assert items is empty OR response.errors is defined",
           "Reversed date range returns empty results or a validation error","NEGATIVE","Active"),

          ("API-AUDIT-011","getAuditLogsByUser with non-existent userId returns empty",
           "1. Send authenticated GraphQL query: getAuditLogsByUser(userId: 'unknown-user-999')\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
           "Empty items array returned for an unknown user ID","NEGATIVE","Active"),

          ("API-AUDIT-012","listAuditLogs with invalid ISO date format returns error",
           "1. Send authenticated GraphQL query: listAuditLogs(startDate: 'not-a-date', endDate: 'also-invalid')\n2. Assert response.errors is defined (GraphQL validation or parsing error)",
           "A GraphQL validation or parsing error is returned for malformed ISO dates","NEGATIVE","Active"),
        ],
      },
      {
        "title": "10 · Users & Customers",
        "suite": "specs/api/users.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-USER-001","getUserByEmail returns user profile for known email",
           "1. Seed a known registered email from the test environment\n2. Send authenticated GraphQL query: getUserByEmail(email: seededEmail)\n3. Assert HTTP 200\n4. Assert no errors\n5. Assert returned user object is defined",
           "Complete user profile returned for the queried email address","SMOKE","Active"),

          ("API-USER-002","getUserByEmail response has required fields",
           "1. Send authenticated GraphQL query: getUserByEmail(email: seededEmail)\n2. Assert HTTP 200\n3. Assert returned user.id, user.email, user.name, user.role are all present",
           "Returned user profile contains all required fields: id, email, name, role","SMOKE","Active"),

          ("API-USER-003","listUsersByRole Admin returns at least one admin user",
           "1. Send authenticated GraphQL query: listUsersByRole(role: 'Admin')\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
           "At least one Admin user is returned","SMOKE","Active"),

          ("API-USER-004","listUsersByRole returns a response for role=Admin",
           "1. Send authenticated GraphQL query: listUsersByRole(role: 'Admin')\n2. Assert HTTP 200\n3. Assert response.data.listUsersByRole is defined",
           "A valid response is returned for role=Admin","REGRESSION","Active"),

          ("API-USER-005","listUsersByRole returns a response for role=Technician",
           "1. Send authenticated GraphQL query: listUsersByRole(role: 'Technician')\n2. Assert HTTP 200\n3. Assert response.data.listUsersByRole is defined",
           "A valid response is returned for role=Technician","REGRESSION","Active"),

          ("API-USER-006","listUsersByRole returns a response for role=Manager",
           "1. Send authenticated GraphQL query: listUsersByRole(role: 'Manager')\n2. Assert HTTP 200\n3. Assert response.data.listUsersByRole is defined",
           "A valid response is returned for role=Manager","REGRESSION","Active"),

          ("API-USER-007","listUsersByRole items have required fields",
           "1. Send authenticated GraphQL query: listUsersByRole(role: 'Admin') with limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.email, item.name, item.role are all present",
           "Each user item contains all required fields: id, email, name, role","REGRESSION","Active"),

          ("API-USER-008","getCustomerWithRelations returns data for a known customer",
           "1. Seed a known customerId from the test environment\n2. Send authenticated GraphQL query: getCustomerWithRelations(customerId: seededId)\n3. Assert HTTP 200\n4. Assert returned customer data is non-null and includes related devices/locations",
           "Customer record with related entities is returned for the queried ID","SMOKE","Active"),

          ("API-USER-009","getUserByEmail with non-existent email returns null or error",
           "1. Send authenticated GraphQL query: getUserByEmail(email: 'nobody@nonexistent999.com')\n2. Assert HTTP 200\n3. Assert response.data.getUserByEmail is null OR response.errors is defined",
           "null data or descriptive error returned for an unregistered email","NEGATIVE","Active"),

          ("API-USER-010","listUsersByRole with unknown role returns 200",
           "1. Send authenticated GraphQL query: listUsersByRole(role: 'UnknownRole')\n2. Assert HTTP 200 (no server-side role validation enforced)\n3. Assert response does not throw an unhandled error",
           "HTTP 200 returned for unknown role (no server-side role enum validation)","NEGATIVE","Active"),

          ("API-USER-011","getCustomerWithRelations with non-existent customerId returns empty or error",
           "1. Send authenticated GraphQL query: getCustomerWithRelations(customerId: 'non-existent-cust-999')\n2. Assert HTTP 200\n3. Assert data is null/empty OR response.errors is defined",
           "null/empty data or error returned for unknown customer ID","NEGATIVE","Active"),

          ("API-USER-012","getUserByEmail with malformed email format returns null or error",
           "1. Send authenticated GraphQL query: getUserByEmail(email: 'not-an-email')\n2. Assert HTTP 200\n3. Assert response.data.getUserByEmail is null OR response.errors is defined",
           "null or validation error returned for a malformed email string","NEGATIVE","Active"),
        ],
      },
      {
        "title": "11 · Mutations",
        "suite": "specs/api/mutations.api.spec.ts  |  Playwright/TypeScript (GraphQL)",
        "rows": [
          ("API-MUT-001","Creates a new firmware record with Pending status",
           "1. Generate unique payload: name='Test-FW-<timestamp>', version='1.0.0', deviceModel='ModelX'\n2. Send authenticated GraphQL mutation: createFirmware(payload)\n3. Assert HTTP 200\n4. Assert returned firmware.id is truthy\n5. Assert returned firmware.status === 'Pending'",
           "New firmware record created with Pending status and a valid ID","SMOKE","Active"),

          ("API-MUT-002","createFirmware with minimal required fields succeeds",
           "1. Build payload with only name and version fields\n2. Send authenticated GraphQL mutation: createFirmware(minimalPayload)\n3. Assert HTTP 200 or graceful skip if resolver rejects missing optional fields\n4. Assert returned record has id and status='Pending'",
           "Firmware creation succeeds with only the minimum required fields","REGRESSION","Active"),

          ("API-MUT-003","createFirmware without required name returns GraphQL error",
           "1. Build payload omitting the 'name' field\n2. Send authenticated GraphQL mutation: createFirmware(payloadWithoutName)\n3. Assert response.errors is defined\n4. Assert error message references missing 'name' field",
           "GraphQL field-required error returned when name is omitted","NEGATIVE","Active"),

          ("API-MUT-004","createFirmware without required version returns GraphQL error",
           "1. Build payload omitting the 'version' field\n2. Send authenticated GraphQL mutation: createFirmware(payloadWithoutVersion)\n3. Assert response.errors is defined\n4. Assert error message references missing 'version' field",
           "GraphQL field-required error returned when version is omitted","NEGATIVE","Active"),

          ("API-MUT-005","Creates a new service order with Pending status",
           "1. Generate unique payload: title='Test-SO-<timestamp>', priority='Medium'\n2. Send authenticated GraphQL mutation: createServiceOrder(payload)\n3. Assert HTTP 200\n4. Assert returned order.id is truthy\n5. Assert returned order.status === 'Pending'",
           "New service order created with Pending status and a valid ID","SMOKE","Active"),

          ("API-MUT-006","createServiceOrder without required title returns GraphQL error",
           "1. Build payload omitting the 'title' field\n2. Send authenticated GraphQL mutation: createServiceOrder(payloadWithoutTitle)\n3. Assert response.errors is defined",
           "GraphQL field-required error returned when title is omitted","NEGATIVE","Active"),

          ("API-MUT-007","Creates a compliance submission in Pending status",
           "1. Seed a valid firmwareId from listFirmware\n2. Generate unique payload: firmwareId, firmwareVersion, deviceModel, certifications=['CE']\n3. Send authenticated GraphQL mutation: createCompliance(payload)\n4. Assert HTTP 200\n5. Assert returned compliance.id is truthy\n6. Assert returned compliance.status === 'Pending'",
           "New compliance record created with Pending status and a valid ID","SMOKE","Active"),

          ("API-MUT-008","createCompliance without required firmwareId returns error",
           "1. Build payload omitting the 'firmwareId' field\n2. Send authenticated GraphQL mutation: createCompliance(payloadWithoutFirmwareId)\n3. Assert response.errors is defined",
           "GraphQL field-required error returned when firmwareId is omitted","NEGATIVE","Active"),

          ("API-MUT-009","Approves a Pending firmware record",
           "1. Create a new firmware record (status=Pending) via createFirmware mutation\n2. Send authenticated mutation: updateEntityStatus(entityId, entityType:'firmware', newStatus:'Approved')\n3. Assert HTTP 200\n4. Assert returned entity status === 'Approved'",
           "Firmware record successfully transitions from Pending to Approved status","SMOKE","Active"),

          ("API-MUT-010","Deprecates a firmware record",
           "1. Create a new firmware record (status=Pending) via createFirmware mutation\n2. Send authenticated mutation: updateEntityStatus(entityId, entityType:'firmware', newStatus:'Deprecated')\n3. Assert HTTP 200\n4. Assert returned entity status === 'Deprecated'",
           "Firmware record successfully transitions to Deprecated status","REGRESSION","Active"),

          ("API-MUT-011","updateEntityStatus with non-existent ID returns 200 (upsert behaviour)",
           "1. Send authenticated mutation: updateEntityStatus(entityId:'non-existent-id-999', entityType:'firmware', newStatus:'Approved')\n2. Assert HTTP 200 (backend upserts on unknown ID)\n3. Assert no unhandled error in response",
           "HTTP 200 returned for unknown entity ID due to upsert behaviour","REGRESSION","Active"),

          ("API-MUT-012","updateEntityStatus with invalid newStatus returns 200",
           "1. Create a firmware record and capture its ID\n2. Send mutation: updateEntityStatus(entityId, entityType:'firmware', newStatus:'INVALID_STATUS')\n3. Assert HTTP 200 (no server-side enum validation)\n4. Assert no unhandled error thrown",
           "HTTP 200 returned for invalid newStatus value (no server-side enum enforcement)","NEGATIVE","Active"),

          ("API-MUT-013","updateEntityStatus without required entityType returns GraphQL error",
           "1. Build mutation payload omitting 'entityType'\n2. Send authenticated mutation: updateEntityStatus(entityId, newStatus:'Approved')\n3. Assert response.errors is defined",
           "GraphQL field-required error returned when entityType is omitted","NEGATIVE","Active"),

          ("API-MUT-014","Updates device coordinates with valid address and lat/lng",
           "1. Seed a valid deviceId from listDevices\n2. Send authenticated mutation: updateDeviceCoords(deviceId, address:'123 Test St', lat:37.7749, lng:-122.4194)\n3. Assert HTTP 200\n4. Assert returned device.location or coords fields are updated",
           "Device coordinates and address are successfully updated","SMOKE","Active"),

          ("API-MUT-015","updateDeviceCoords without required address returns GraphQL error",
           "1. Seed a valid deviceId\n2. Build mutation payload omitting the 'address' field\n3. Send authenticated mutation: updateDeviceCoords(deviceId, lat:37.7749, lng:-122.4194)\n4. Assert response.errors is defined",
           "GraphQL field-required error returned when address is omitted","NEGATIVE","Active"),

          ("API-MUT-016","updateDeviceCoords with out-of-range coordinates returns error",
           "1. Seed a valid deviceId\n2. Send authenticated mutation: updateDeviceCoords(deviceId, address:'Test', lat:999, lng:999)\n3. Assert response.errors is defined or HTTP 400 returned\n4. Assert error references invalid coordinate range",
           "Validation error returned when lat/lng values exceed valid geographic range","NEGATIVE","Active"),
        ],
      },
    ],
  },
]

# ── helpers ───────────────────────────────────────────────────────────────────
def write_section(ws, start_row, group):
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    c = ws.cell(row=r, column=1, value=group["title"])
    c.fill, c.font = mk_fill(CLR["title_bg"]), mk_font(bold=True, color=CLR["title_fg"], size=11)
    c.alignment, c.border = mk_align(wrap=False), mk_border()
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    c = ws.cell(row=r, column=1, value=group["suite"])
    c.fill, c.font = mk_fill("D9E1F2"), mk_font(color="1F3864", size=9)
    c.alignment, c.border = mk_align(wrap=False), mk_border()
    r += 1

    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=r, column=ci, value=col)
        c.fill, c.font = mk_fill(CLR["header_bg"]), mk_font(bold=True, color=CLR["header_fg"])
        c.alignment, c.border = mk_align(wrap=False, h="center"), mk_border()
    r += 1

    for idx, row in enumerate(group["rows"]):
        bg = CLR["row_white"] if idx % 2 == 0 else CLR["row_alt"]
        tc_id, scenario, steps, expected, typ, status = row
        vals = [tc_id, scenario, steps, expected, typ, status]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill, c.font = mk_fill(bg), mk_font(size=10)
            c.alignment = mk_align(wrap=True, h="left" if ci not in (1,5,6) else "center")
            c.border = mk_border()
        ws.cell(row=r, column=5).fill = TYPE_FILL.get(typ, mk_fill(bg))
        ws.cell(row=r, column=5).font = mk_font(bold=True, size=10)
        ws.cell(row=r, column=5).alignment = mk_align(wrap=False, h="center")
        ws.cell(row=r, column=6).fill = STATUS_FILL.get(status, mk_fill(bg))
        ws.cell(row=r, column=6).font = mk_font(bold=True, size=10)
        ws.cell(row=r, column=6).alignment = mk_align(wrap=False, h="center")
        r += 1
    return r + 1

# ── build sheets ──────────────────────────────────────────────────────────────
summary_data = []

for section in SECTIONS:
    ws = wb.create_sheet(title=section["sheet"])
    ws.freeze_panes = "A2"
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 5
    current_row = 2
    for group in section["groups"]:
        current_row = write_section(ws, current_row, group)
        smoke = sum(1 for r in group["rows"] if r[4] == "SMOKE")
        regr  = sum(1 for r in group["rows"] if r[4] == "REGRESSION")
        neg   = sum(1 for r in group["rows"] if r[4] == "NEGATIVE")
        summary_data.append({"sheet": section["sheet"], "module": group["title"],
                              "smoke": smoke, "regression": regr, "negative": neg,
                              "total": smoke + regr + neg})

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet(title="Summary", index=0)
ws_sum.freeze_panes = "A3"
ws_sum.merge_cells("A1:H1")
t = ws_sum["A1"]
t.value = "HLM-QA — Test Case Summary  |  Generated: 2026-03-25"
t.fill, t.font = mk_fill(CLR["header_bg"]), mk_font(bold=True, color="FFFFFF", size=13)
t.alignment, t.border = mk_align(wrap=False, h="center"), mk_border()
ws_sum.row_dimensions[1].height = 28

sum_cols   = ["#", "Sheet", "Module", "SMOKE", "REGRESSION", "NEGATIVE", "TOTAL", "Columns"]
sum_widths = [5, 14, 42, 12, 14, 12, 10, 55]
for ci, (col, w) in enumerate(zip(sum_cols, sum_widths), 1):
    c = ws_sum.cell(row=2, column=ci, value=col if col != "Columns" else "Columns: TC ID | Scenario | Test Steps | Expected Result | Type | Status")
    c.fill, c.font = mk_fill(CLR["title_bg"]), mk_font(bold=True, color="FFFFFF")
    c.alignment, c.border = mk_align(wrap=False, h="center"), mk_border()
    ws_sum.column_dimensions[get_column_letter(ci)].width = w

for idx, row in enumerate(summary_data, 1):
    bg = CLR["row_white"] if idx % 2 == 0 else CLR["row_alt"]
    vals = [idx, row["sheet"], row["module"], row["smoke"], row["regression"], row["negative"], row["total"], ""]
    for ci, val in enumerate(vals, 1):
        c = ws_sum.cell(row=idx + 2, column=ci, value=val)
        c.fill, c.font = mk_fill(bg), mk_font(size=10)
        c.alignment = mk_align(wrap=False, h="center" if ci != 3 else "left")
        c.border = mk_border()
    ws_sum.cell(row=idx+2, column=4).fill = TYPE_FILL["SMOKE"]
    ws_sum.cell(row=idx+2, column=5).fill = TYPE_FILL["REGRESSION"]
    ws_sum.cell(row=idx+2, column=6).fill = TYPE_FILL["NEGATIVE"]

total_row = len(summary_data) + 3
totals = ["", "", "TOTAL",
          sum(r["smoke"] for r in summary_data), sum(r["regression"] for r in summary_data),
          sum(r["negative"] for r in summary_data), sum(r["total"] for r in summary_data), ""]
for ci, val in enumerate(totals, 1):
    c = ws_sum.cell(row=total_row, column=ci, value=val)
    c.fill, c.font = mk_fill(CLR["header_bg"]), mk_font(bold=True, color="FFFFFF", size=11)
    c.alignment = mk_align(wrap=False, h="center" if ci != 3 else "right")
    c.border = mk_border()

leg_row = total_row + 2
ws_sum.cell(row=leg_row, column=1, value="Legend").font = mk_font(bold=True, size=10)
legends = [
    ("SMOKE",      CLR["smoke_bg"], "Core happy-path tests; must pass on every build"),
    ("REGRESSION", CLR["regr_bg"],  "Full regression coverage; run on release candidates"),
    ("NEGATIVE",   CLR["neg_bg"],   "Error / boundary / auth failure scenarios"),
    ("Active",     "C6EFCE",        "Test is implemented and included in a suite"),
    ("Pending",    "BDD7EE",        "Planned but not yet implemented"),
    ("Skipped",    "FFEB9C",        "Exists but excluded / marked skip"),
    ("Deprecated", "F4CCCC",        "Removed or no longer relevant"),
]
for i, (label, colour, meaning) in enumerate(legends, 1):
    c_lbl = ws_sum.cell(row=leg_row + i, column=1, value=label)
    c_lbl.fill, c_lbl.font = mk_fill(colour), mk_font(bold=True, size=10)
    c_lbl.border, c_lbl.alignment = mk_border(), mk_align(wrap=False, h="center")
    c_desc = ws_sum.cell(row=leg_row + i, column=2, value=meaning)
    c_desc.font = mk_font(size=10)
    ws_sum.merge_cells(start_row=leg_row+i, start_column=2, end_row=leg_row+i, end_column=8)
    c_desc.alignment = mk_align(wrap=False, h="left")

out = "/Users/ajaykumar.yadav/HLM-QA/TEST-CASES.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total test cases: {sum(r['total'] for r in summary_data)}")
