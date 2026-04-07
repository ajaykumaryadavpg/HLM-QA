#!/usr/bin/env python3
"""Regenerate TEST-CASES.xlsx from all 246 test cases in TEST-CASES.mdc"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
# DATA: All 246 test cases
# Format: (section, suite_file, test_class, tc_id, scenario, test_steps, expected_result, type_, status)
# ─────────────────────────────────────────────

TEST_CASES = [
    # ── Section 1: Login (5) ────────────────────────────────────────────────
    ("UI — Login", "inventory-login-suite.xml", "InventoryLoginTests",
     "UI-LOGIN-001", "Successful login with valid credentials",
     "1. Clear browser cookies and session/local storage\n2. Launch the HLM Platform login page\n3. Verify login page heading 'Sign in to your account' is visible\n4. Enter valid admin email and password, click Sign In\n5. Verify Dashboard h1 heading is visible after redirect\n6. Verify 'Inventory & Assets' nav link is present in sidebar\n7. Verify welcome message contains 'Welcome back'",
     "User is redirected to the Dashboard; h1 header, nav link, and welcome message are all visible",
     "SMOKE", "Active"),
    ("UI — Login", "inventory-login-suite.xml", "InventoryLoginTests",
     "UI-LOGIN-002", "Login fails with invalid credentials",
     "1. Clear session state and launch login page\n2. Enter invalid email and invalid password\n3. Click Sign In\n4. Verify red error banner (p.text-red-400) is visible\n5. Verify Sign In button is still present (no redirect occurred)",
     "Red error banner is displayed; user remains on the login page",
     "NEGATIVE", "Active"),
    ("UI — Login", "inventory-login-suite.xml", "InventoryLoginTests",
     "UI-LOGIN-003", "Login with empty email stays on login page",
     "1. Clear session state and launch login page\n2. Leave email field empty\n3. Enter valid password\n4. Click Sign In\n5. Verify Sign In button still visible (HTML5 blocked empty email)\n6. Verify login page heading is still present",
     "HTML5 validation blocks form submission; user remains on login page",
     "NEGATIVE", "Active"),
    ("UI — Login", "inventory-login-suite.xml", "InventoryLoginTests",
     "UI-LOGIN-004", "Login with empty password stays on login page",
     "1. Clear session state and launch login page\n2. Enter valid email address\n3. Leave password field empty\n4. Click Sign In\n5. Verify Sign In button still visible (no redirect)",
     "HTML5 validation blocks form submission; user remains on login page",
     "NEGATIVE", "Active"),
    ("UI — Login", "inventory-login-suite.xml", "InventoryLoginTests",
     "UI-LOGIN-005", "Login with both fields empty stays on login page",
     "1. Clear session state and launch login page\n2. Click Sign In without filling any credentials\n3. Verify Sign In button still visible\n4. Verify login page heading is still present",
     "HTML5 validation blocks empty form; user stays on login page",
     "NEGATIVE", "Active"),

    # ── Section 2: KPI Cards (18) ───────────────────────────────────────────
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-001", "Four KPI cards visible after login",
     "1. Re-navigate to Dashboard for a clean state\n2. Verify 'Welcome back' h2 heading is visible\n3. Verify Total Devices KPI card value element is visible\n4. Verify Active Deployments KPI card value element is visible\n5. Verify Pending Approvals KPI card value element is visible\n6. Verify Health Score KPI card value element is visible\n7. Verify no red error banner is present\n8. Verify no '—' loading placeholder remains",
     "All four KPI cards are visible, populated, and error-free", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-002", "Total Devices card has blue icon and numeric count",
     "1. Verify 'Total Devices' card label (div.text-sm) is visible\n2. Verify Total Devices numeric value (div.text-3xl) is visible\n3. Verify blue SVG Package icon (text-blue-* class) is present\n4. Verify no '—' loading placeholder remains",
     "Total Devices card shows label, blue icon, and a resolved numeric count", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-003", "Active Deployments card has green icon and in-progress count",
     "1. Verify 'Active Deployments' card label is visible\n2. Verify Active Deployments numeric value is visible\n3. Verify green SVG Download icon (text-green-* class) is present\n4. Verify no loading placeholder remains",
     "Active Deployments card shows label, green icon, and resolved count", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-004", "Pending Approvals card has orange icon and combined count",
     "1. Verify 'Pending Approvals' card label is visible\n2. Verify Pending Approvals numeric value (combined firmware + compliance) is visible\n3. Verify orange SVG Shield icon (text-orange-* class) is present\n4. Verify no loading placeholder remains",
     "Pending Approvals card shows label, orange icon, and aggregated count", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-005", "Health Score card has green icon and percentage format",
     "1. Verify 'Health Score' card label is visible\n2. Verify Health Score value contains '%' suffix (e.g., '77%')\n3. Verify green SVG Check-circle icon (text-green-* class) is present\n4. Verify no loading placeholder remains",
     "Health Score card shows label, green icon, and a value in N% format", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-006", "KPI cards show em-dash loading placeholder on page load",
     "1. Register route handler to delay all AppSync GraphQL responses by 2 s\n2. Re-navigate to Dashboard URL to trigger fresh page load\n3. Verify em-dash '—' (U+2014) is visible in at least one KPI card\n4. Unregister route handler (cleanup via finally block)",
     "At least one KPI card shows '—' placeholder while APIs are in-flight", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-007", "KPI cards show zero fallback when no data exists",
     "1. Navigate to Dashboard, wait for live data to confirm loaded state\n2. Register route to return empty arrays for all fetch/xhr requests\n3. Click Refresh Dashboard to re-trigger all KPI APIs\n4. Verify no error banner is shown\n5. Verify Total Devices card shows '0'\n6. Verify Active Deployments card shows '0'\n7. Verify Pending Approvals card shows '0'\n8. Verify Health Score card shows '0%'\n9. Unregister route handler (cleanup)",
     "All KPI cards display '0' or '0%' with no error banner when APIs return empty data", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-008", "KPI cards render in 4-column layout at desktop viewport",
     "1. Verify KPI grid container (div.grid) is visible at current desktop viewport\n2. Verify all four KPI card values are visible without horizontal scrolling\n3. Evaluate via JS whether grid-cols-4 Tailwind class is active on grid container\n4. Assert KPI grid container is present confirming 4-column layout",
     "All four cards visible in a single 4-column row at 1280px+ width", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-009", "KPI cards render in 2×2 grid at tablet viewport",
     "1. Resize browser viewport to 768×1024 (tablet)\n2. Re-navigate to Dashboard to trigger reflow\n3. Verify Total Devices card is visible at tablet width\n4. Verify Active Deployments, Pending Approvals, Health Score cards are visible\n5. Restore viewport to 1620×1080 (desktop)",
     "Cards reflow into a 2×2 grid; all four remain visible at tablet width", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-010", "KPI cards stack in single column at mobile viewport",
     "1. Resize browser viewport to 375×667 (mobile)\n2. Re-navigate to Dashboard to trigger reflow\n3. Verify Total Devices card is visible at mobile width\n4. Verify all four KPI cards are accessible without horizontal scrolling\n5. Restore viewport to 1620×1080 (desktop)",
     "All four cards stack vertically in a single column at 375px width", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-011", "Refresh Dashboard button re-fetches all KPI data",
     "1. Verify Dashboard is loaded with live data (Total Devices card visible)\n2. Click the Refresh Dashboard button\n3. Verify all four KPI cards are populated after refresh\n4. Verify no error banner is present post-refresh",
     "All KPI cards update with fresh data after clicking Refresh Dashboard", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-012", "API failure shows red error banner and KPI fallback values",
     "1. Register route to return HTTP 500 from GraphQL endpoint\n2. Navigate to Dashboard with API failing\n3. Verify red error banner container (div.bg-red-50) is visible\n4. Verify error banner message is visible\n5. Unregister route handler (cleanup)",
     "Red error banner appears with message; KPI cards show fallback values", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-013", "KPI cards render correctly in dark mode",
     "1. Activate dark mode on the Dashboard\n2. Verify all four KPI card value elements are visible\n3. Verify all four KPI card label elements are visible\n4. Verify all four KPI card icon elements are visible\n5. Verify no error banner is present in dark mode",
     "All KPI card icons, labels, and values render correctly with dark mode active", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-014", "Health Score value conforms to rounded percentage format",
     "1. Verify Health Score KPI card value element is visible\n2. Assert value text matches pattern ^\\d{1,3}%$ (e.g., '0%', '77%', '100%')\n3. Verify no loading placeholder '—' remains",
     "Health Score displays as a rounded integer percentage with no decimals", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-015", "KPI card icon colors match specification",
     "1. Verify Total Devices icon has a blue Tailwind class (text-blue-*)\n2. Verify Active Deployments icon has a green Tailwind class (text-green-*)\n3. Verify Pending Approvals icon has an orange Tailwind class (text-orange-*)\n4. Verify Health Score icon has a green Tailwind class (text-green-*)",
     "Each KPI card icon renders with the correct specification colour class", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-016", "KPI values are never negative or invalid artifacts",
     "1. Read text content of Total Devices value element\n2. Assert value is a non-negative integer string\n3. Read Active Deployments, Pending Approvals values and assert >= 0\n4. Read Health Score value and assert it matches N% with 0 <= N <= 100\n5. Verify no value is 'NaN', 'undefined', 'null', or empty",
     "All KPI values are valid non-negative numbers; no JavaScript artifacts present", "REGRESSION", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-017", "Unauthenticated access to dashboard redirects to login",
     "1. Clear all cookies and browser storage to remove auth session\n2. Navigate directly to the Dashboard URL\n3. Verify Sign In button is visible (redirected to login page)\n4. Verify login page heading is visible\n5. Verify Dashboard header is NOT present",
     "Unauthenticated navigation to dashboard redirects to login page", "SMOKE", "Active"),
    ("UI — Dashboard KPI Cards", "dashboard-kpi-suite.xml", "DashboardKpiTests",
     "UI-KPI-018", "Pending Approvals displays sum of firmware and compliance counts",
     "1. Intercept listFirmware and listCompliance GraphQL requests and capture responses\n2. Count records with status='Pending' from firmware API response\n3. Count records with status='Pending' from compliance API response\n4. Read Pending Approvals KPI card displayed value\n5. Assert displayed value = pending firmware count + pending compliance count",
     "Pending Approvals KPI equals the arithmetic sum of pending firmware + compliance records", "REGRESSION", "Active"),

    # ── Section 3: Dashboard API (31) ──────────────────────────────────────
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-001", "Parallel API calls on dashboard load",
     "1. Navigate to Dashboard and trigger fresh page load\n2. Verify Total Devices KPI value is visible (devices API responded)\n3. Verify Active Deployments KPI value is visible (service orders API responded)\n4. Verify Pending Approvals KPI value is visible (firmware + compliance APIs responded)\n5. Verify no loading placeholder '—' remains",
     "All six data-domain APIs respond and their values are reflected across KPI cards", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-002", "All device records are fetched",
     "1. Verify Total Devices KPI card value element is present in the DOM\n2. Verify no loading placeholder '—' remains — data has fully loaded",
     "Total Devices KPI card displays a resolved numeric count", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-003", "Offline devices are fetched separately",
     "1. Verify 'View Inventory' card is present in the Quick Actions section\n2. Verify orange offline-devices badge (span.bg-orange-500) is visible inside the 'View Inventory' card",
     "Offline device count badge is visible on the Quick Actions Inventory card", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-004", "In-progress service orders are fetched",
     "1. Verify Active Deployments KPI card value element is present\n2. Verify no loading placeholder '—' remains across all KPI cards",
     "Active Deployments KPI card shows a resolved count after APIs respond", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-005", "Scheduled service orders are fetched",
     "1. Verify 'Scheduled Service' card is present in Quick Actions section\n2. Verify orange scheduled-orders badge (span.bg-orange-500) is visible inside the 'Scheduled Service' card",
     "Scheduled orders count badge is visible on the Quick Actions Scheduled Service card", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-006", "Pending firmware records are fetched",
     "1. Verify Pending Approvals KPI card value element is visible\n2. Verify 'Check Compliance' quick action link is visible",
     "Pending Approvals KPI includes pending firmware count; compliance link is present", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-007", "Pending compliance records are fetched",
     "1. Verify Pending Approvals KPI card value element is visible (combined firmware + compliance)\n2. Verify no loading placeholder '—' remains in Pending Approvals card",
     "Pending Approvals KPI reflects both pending firmware and pending compliance counts", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-008", "KPI cards show loading placeholder on page load",
     "1. Register route to delay all GraphQL responses by 2 s\n2. Navigate directly to the Dashboard URL\n3. Verify '—' (em-dash, U+2014) is visible in at least one KPI card during fetch\n4. Unregister route handler (cleanup via finally block)",
     "Em-dash placeholder is visible in KPI cards while API responses are pending", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-009", "KPI cards transition from placeholder to live values",
     "1. Verify Total Devices KPI value is present (not placeholder)\n2. Verify Active Deployments KPI value is present\n3. Verify Pending Approvals KPI value is present\n4. Verify no loading placeholder '—' remains across any KPI card",
     "All KPI cards display live data values; no placeholder remains after load", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-010", "Error banner displayed on single API failure",
     "1. Register route to return HTTP 500 from GraphQL endpoint\n2. Navigate to Dashboard with API returning 500\n3. Verify red error banner container (div.bg-red-50) is visible\n4. Verify error message text is visible within the banner\n5. Unregister route handler (cleanup)",
     "Red error banner with message appears when any API returns a server error", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-011", "Error banner and placeholders when all APIs fail",
     "1. Register route to abort all GraphQL requests (network-level failure)\n2. Navigate to Dashboard with all APIs unavailable\n3. Verify red error banner is visible\n4. Verify KPI cards show '0' fallback (app resolved error state)\n5. Verify Dashboard header is still rendered (no page crash)\n6. Unregister route handler (cleanup)",
     "Error banner shown; KPI cards show fallback '0'; dashboard does not crash", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-012", "Error banner displays meaningful message",
     "1. Register route to return HTTP 500 with specific error message payload\n2. Navigate to Dashboard with error route active\n3. Verify red error banner container is visible\n4. Verify error banner message element is visible and non-empty\n5. Unregister route handler (cleanup)",
     "Error banner contains a visible, non-empty, API-specific error message", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-013", "Audit logs fetched for alerts panel",
     "1. Register route to intercept listAuditLogs GraphQL query and inject a mock audit log entry\n2. Re-navigate to Dashboard so mock response is applied\n3. Verify Recent Alerts panel container is visible\n4. Verify at least one audit log list item is rendered in the panel\n5. Unregister route handler (cleanup)",
     "Recent Alerts panel is visible and shows at least one audit log entry", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-014", "Audit logs 24-hour time window",
     "1. Register route to inject a mock audit log entry within last 24 hours\n2. Re-navigate to Dashboard so mock response is applied\n3. Verify Recent Alerts panel container is visible\n4. Verify alert items from within 24-hour window are displayed\n5. Verify 'No recent activity' message is NOT shown\n6. Unregister route handler (cleanup)",
     "Audit log entries from last 24 hours appear in the Alerts panel", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-015", "Alerts loading indicator displayed while fetching",
     "1. Register route to delay listAuditLogs GraphQL response by 2 s\n2. Re-navigate to Dashboard\n3. Verify loading indicator (spinner or skeleton) is visible in Alerts panel during fetch\n4. Unregister route handler (cleanup)",
     "A loading indicator is shown in the Alerts panel while audit log fetch is in-flight", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-016", "Alerts error message on API failure",
     "1. Register route to abort listAuditLogs GraphQL request\n2. Navigate to Dashboard with alerts API failing\n3. Verify Alerts panel container is visible\n4. Verify error message is displayed inside the Alerts panel\n5. Unregister route handler (cleanup)",
     "Alerts panel shows a descriptive error message when its API call fails", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-017", "No dashboard auto-refresh or polling",
     "1. Navigate to Dashboard and wait for full load\n2. Capture initial KPI values\n3. Wait 35 seconds without interaction\n4. Verify no additional API calls have been triggered\n5. Verify KPI values remain unchanged (no auto-refresh occurred)",
     "Dashboard data does not refresh automatically; values remain static until user action", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-018", "Device list is capped at 100 records",
     "1. Navigate to Dashboard and verify Total Devices KPI is visible\n2. Intercept the listDevices GraphQL request and inspect the limit variable\n3. Assert that limit <= 100 in the API request\n4. Verify Total Devices KPI value is <= 100",
     "The listDevices API call is sent with a limit of 100 or fewer records", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-019", "Dashboard loads with zero data",
     "1. Register route to return empty items arrays for all dashboard APIs\n2. Navigate to Dashboard\n3. Verify no red error banner appears\n4. Verify all four KPI cards show '0' or '0%'\n5. Verify dashboard header is still rendered\n6. Unregister route handler (cleanup)",
     "Dashboard renders cleanly with all-zero KPI values when APIs return empty datasets", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-020", "KPI and alerts fetch independently",
     "1. Register route to abort ONLY the listAuditLogs GraphQL request\n2. Navigate to Dashboard with alerts API failing\n3. Verify KPI cards are populated (Total Devices, Active Deployments, Pending Approvals, Health Score)\n4. Verify error message appears in Alerts panel only\n5. Unregister route handler (cleanup)",
     "KPI cards load successfully even when the Alerts panel API fails", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-021", "Refresh Dashboard button triggers complete data refetch",
     "1. Verify Dashboard is loaded with initial data\n2. Click the Refresh Dashboard button (circular arrow icon)\n3. Verify loading placeholders briefly appear in KPI cards\n4. Verify all KPI cards repopulate with fresh data\n5. Verify Alerts panel refreshes\n6. Verify no error banner is present post-refresh",
     "Clicking Refresh triggers re-fetch of all APIs; all panels update", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-022", "Refresh button visible with circular arrow icon",
     "1. Verify Dashboard is loaded\n2. Verify the Refresh Dashboard button element is visible on the page\n3. Verify the button contains a circular arrow SVG icon",
     "Refresh Dashboard button is visible with a circular arrow icon", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-023", "KPI cards show placeholders during refresh",
     "1. Navigate to Dashboard and wait for full load\n2. Register route to delay GraphQL responses by 2 s\n3. Click the Refresh Dashboard button\n4. Immediately verify KPI cards show '—' loading placeholders\n5. Unregister route handler (cleanup)",
     "KPI cards revert to '—' loading state while a manual refresh is in progress", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-024", "Alerts panel shows loading state during refresh",
     "1. Navigate to Dashboard and wait for full load\n2. Register route to delay listAuditLogs response by 2 s\n3. Click the Refresh Dashboard button\n4. Verify loading indicator appears in Alerts panel during fetch\n5. Unregister route handler (cleanup)",
     "Alerts panel shows a loading indicator while a manual refresh is fetching audit logs", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-025", "Error banners cleared when refresh starts",
     "1. Register route to return HTTP 500 from GraphQL endpoint\n2. Navigate to Dashboard so error banner appears\n3. Unregister error route and register success route\n4. Click the Refresh Dashboard button\n5. Verify the error banner is cleared immediately when refresh starts\n6. Verify KPI cards repopulate with data after successful refresh",
     "Any existing error banners are dismissed when Refresh Dashboard is clicked", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-026", "System Status panel displays four services",
     "1. Verify Dashboard is loaded\n2. Verify System Status panel container is visible\n3. Verify exactly four service rows are rendered inside the panel\n4. Verify each row has a service name label and a status indicator",
     "System Status panel renders with exactly four service entries", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-027", "System Status health percentages update from live KPI data",
     "1. Wait for Dashboard to fully load\n2. Read Total Devices and Health Score KPI values\n3. Read health percentage values displayed in the System Status panel\n4. Assert System Status health percentages correlate with live KPI data",
     "System Status health percentages are derived from and aligned with live KPI values", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-028", "System Status operational and degraded labels",
     "1. Wait for Dashboard to fully load\n2. Verify at least one service row with 100% health shows 'Operational' label\n3. Verify any service row below health threshold shows 'Degraded' label",
     "Service rows display 'Operational' or 'Degraded' labels based on health threshold", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-029", "Quick action badges hidden when count is zero",
     "1. Register route to return empty arrays for firmware and compliance APIs\n2. Click Refresh Dashboard\n3. Verify Quick Action Firmware badge is NOT visible\n4. Verify Quick Action Compliance badge is NOT visible\n5. Unregister route handler (cleanup)",
     "Quick Action count badges are hidden when the corresponding pending count is zero", "REGRESSION", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-030", "Quick action badges visible when count is greater than zero",
     "1. Verify Dashboard is loaded with live data\n2. Verify Pending Approvals KPI value is greater than zero\n3. Verify Quick Action badge for Firmware or Compliance is visible with a count",
     "Quick Action badges display numeric counts when pending items exist", "SMOKE", "Active"),
    ("UI — Dashboard API Data Layer", "dashboard-api-suite.xml", "DashboardApiTests",
     "UI-DAPI-031", "No recent activity message when no audit logs exist",
     "1. Register route to return empty items for listAuditLogs query\n2. Re-navigate to Dashboard so empty audit logs response is applied\n3. Verify Alerts panel container is visible\n4. Verify 'No recent activity' empty-state message is displayed\n5. Verify no audit log list items are rendered\n6. Unregister route handler (cleanup)",
     "Alerts panel shows 'No recent activity' empty-state when audit log API returns no items", "REGRESSION", "Active"),

    # ── Section 4: Geo Location (16) ───────────────────────────────────────
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-001", "Geo Location tab renders interactive MapLibre map",
     "1. Navigate to Dashboard -> Inventory module via sidebar\n2. Wait for Inventory page heading to appear\n3. Click the Geo Location tab\n4. Wait for MapLibre map canvas and device pins to render\n5. Verify div.maplibregl-map container is present\n6. Verify HTML5 canvas element is rendered inside map container\n7. Verify [aria-label='Map'] region is present\n8. Verify legacy static SVG map element is NOT present\n9. Verify at least one .maplibregl-marker pin is visible\n10. Verify MapLibre attribution link is visible",
     "Interactive MapLibre map renders; canvas and markers present; no legacy SVG map", "SMOKE", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-002", "Map legend displays correct status entries",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'Online' text entry is visible in the legend panel\n3. Verify 'Offline' text entry is visible in the legend panel\n4. Verify 'Maintenance' text entry is visible in the legend panel",
     "Map legend shows Online, Offline, and Maintenance entries with colour indicators", "SMOKE", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-003", "All twelve pins visible with correct color distribution",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'All' filter button is active by default\n3. Verify at least one map marker is visible\n4. Evaluate via JS: count background-color of all pin dot divs\n5. Assert total pin count = 12\n6. Assert 6 pins rgb(34,197,94) green, 3 pins rgb(239,68,68) red, 3 pins rgb(249,115,22) orange",
     "12 pins rendered; JS color evaluation confirms 6 green, 3 red, 3 orange distribution", "SMOKE", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-004", "Online filter shows six green pins",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click 'Online(6)' filter button\n3. Verify Online filter button is active/highlighted\n4. Wait for filtered pins to appear in DOM\n5. Evaluate via JS: count all map marker elements -> assert = 6\n6. Evaluate via JS: count green pins rgb(34,197,94) -> assert = 6",
     "Exactly 6 green pins visible after Online filter; all have correct green colour", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-005", "Offline filter shows three red pins",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click 'Offline(3)' filter button\n3. Wait for filtered pins to appear in DOM\n4. Evaluate via JS: count all map marker elements -> assert = 3\n5. Evaluate via JS: count red pins rgb(239,68,68) -> assert = 3",
     "Exactly 3 red pins visible after Offline filter; all have correct red colour", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-006", "Maintenance filter shows three orange pins",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click 'Maintenance(3)' filter button\n3. Wait for filtered pins to appear in DOM\n4. Evaluate via JS: count all map marker elements -> assert = 3\n5. Evaluate via JS: count orange pins rgb(249,115,22) -> assert = 3",
     "Exactly 3 orange pins visible after Maintenance filter; all have correct orange colour", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-007", "Stat pills display correct device counts",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify Total, Online, Offline, Maintenance stat pills are all visible\n3. Verify Online filter button label contains '6'\n4. Verify Offline filter button label contains '3'\n5. Verify Maintenance filter button label contains '3'",
     "All four stat pills visible; counts match Online(6), Offline(3), Maintenance(3)", "SMOKE", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-008", "All four filter buttons are present",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify 'All' filter button is visible and active by default\n3. Verify 'Online(N)' filter button is visible with count suffix\n4. Verify 'Offline(N)' filter button is visible with count suffix\n5. Verify 'Maintenance(N)' filter button is visible with count suffix",
     "All four filter buttons (All, Online(6), Offline(3), Maintenance(3)) are present", "SMOKE", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-009", "Filter sequence updates map markers correctly",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Assert initial pin count = 12 (All filter default)\n3. Click Online filter -> wait -> assert pin count = 6\n4. Click Offline filter -> wait -> assert pin count = 3\n5. Click Maintenance filter -> wait -> assert pin count = 3\n6. Click All filter -> wait -> assert pin count = 12",
     "Pin counts change correctly at each filter step: 12->6->3->3->12", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-010", "Maintenance pin opens detail card with correct fields",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click Maintenance filter to show only orange pins\n3. Click on the first Maintenance pin using browser.evaluate() (force click)\n4. Wait for device detail card to appear\n5. Verify detail card container is visible\n6. Verify device name, status, address, coordinates, and last-seen fields are present",
     "Clicking a Maintenance pin opens a detail card with all required device fields", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-011", "Device detail card closes on X button click",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Click on any map marker to open the device detail card\n3. Verify detail card is visible\n4. Click the X (close) button on the detail card\n5. Verify detail card container is no longer visible",
     "Device detail card is dismissed when the X button is clicked", "SMOKE", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-012", "Zoom in and zoom out controls are functional",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Evaluate via JS: get current map zoom level\n3. Click the Zoom In (+) control button\n4. Wait briefly for zoom animation\n5. Evaluate via JS: assert new zoom level > initial zoom level\n6. Click the Zoom Out (-) control button\n7. Evaluate via JS: assert zoom level decreased",
     "Zoom In increases map zoom level; Zoom Out decreases it", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-013", "Compass button is present and interactive",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Verify compass/bearing reset button element is visible\n3. Click the compass button\n4. Verify no error is thrown and map remains interactive",
     "Compass button is visible and responds to user clicks without error", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-014", "Hardware Inventory tab unaffected by PS20",
     "1. Navigate to Inventory module (via BeforeMethod)\n2. Click the Hardware Inventory tab\n3. Verify Hardware Inventory tab content is visible\n4. Verify no MapLibre or Geo Location elements leak into Hardware Inventory view\n5. Verify existing Hardware Inventory table/content renders correctly",
     "Hardware Inventory tab content is unchanged and unaffected by PS20 changes", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-015", "Firmware Status tab unaffected by PS20",
     "1. Navigate to Inventory module (via BeforeMethod)\n2. Click the Firmware Status tab\n3. Verify Firmware Status tab content is visible\n4. Verify no Geo Location UI elements are present in Firmware Status view\n5. Verify existing Firmware Status content renders correctly",
     "Firmware Status tab content is unchanged and unaffected by PS20 changes", "REGRESSION", "Active"),
    ("UI — Geo Location Map", "geo-location-suite.xml", "GeoLocationTests",
     "UI-GEO-016", "Stat pill counts are static across filter changes",
     "1. Navigate to Geo Location tab (via BeforeMethod)\n2. Read Online stat pill count (should be 6)\n3. Click Online filter -> assert Online stat pill still shows 6\n4. Click Offline filter -> assert Offline stat pill still shows 3\n5. Click Maintenance filter -> assert Maintenance stat pill still shows 3\n6. Click All filter and verify all counts unchanged",
     "Stat pill counts remain fixed at 6/3/3 regardless of which filter is active", "REGRESSION", "Active"),

    # ── Section 5: API — Devices (15) ──────────────────────────────────────
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-001", "listDevices returns paginated response with items",
     "1. Send authenticated GraphQL query: listDevices (no filter)\n2. Assert HTTP 200 status code\n3. Assert no errors in response\n4. Assert listDevices result object is defined\n5. Parse items array and assert length > 0",
     "Response is HTTP 200 with a populated items array", "SMOKE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-002", "listDevices with limit returns at most N items",
     "1. Send authenticated GraphQL query: listDevices(limit: 3)\n2. Assert HTTP 200 status code\n3. Parse items array\n4. Assert items.length <= 3",
     "Response contains at most 3 items when limit=3 is specified", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-003", "listDevices filters by status=Online",
     "1. Send authenticated GraphQL query: listDevices(status: \"Online\")\n2. Assert HTTP 200 status code\n3. Parse items array\n4. Assert every item has status === \"Online\"",
     "All returned items have status='Online'", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-004", "listDevices filters by status=Offline",
     "1. Send authenticated GraphQL query: listDevices(status: \"Offline\")\n2. Assert HTTP 200\n3. Parse items and assert every item has status === \"Offline\"",
     "All returned items have status='Offline'", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-005", "listDevices filters by status=Maintenance",
     "1. Send authenticated GraphQL query: listDevices(status: \"Maintenance\")\n2. Assert HTTP 200\n3. Parse items and assert every item has status === \"Maintenance\"",
     "All returned items have status='Maintenance'", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-006", "listDevices response items have required fields",
     "1. Send authenticated GraphQL query: listDevices(limit: 1)\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.deviceName, item.serialNumber, item.model, item.status, item.location are all present",
     "Each device item contains all required fields", "SMOKE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-007", "listDevices totalCount is a non-negative integer",
     "1. Send authenticated GraphQL query: listDevices\n2. Assert HTTP 200\n3. Read totalCount from response\n4. Assert totalCount is a number and totalCount >= 0",
     "totalCount field is present and is a non-negative integer", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-008", "getDevice returns device by valid ID",
     "1. Seed firstDeviceId in beforeAll by calling listDevices\n2. Send authenticated GraphQL query: getDevice(id: firstDeviceId)\n3. Assert HTTP 200\n4. Assert no errors in response\n5. Assert returned device.id === firstDeviceId",
     "getDevice returns the complete device record for the queried ID", "SMOKE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-009", "getDevicesByCustomer returns devices for a customer",
     "1. Seed firstCustomerId in beforeAll\n2. Send authenticated GraphQL query: getDevicesByCustomer(customerId: firstCustomerId)\n3. Assert HTTP 200\n4. Parse items array and assert length > 0",
     "All devices belonging to the specified customer are returned", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-010", "getDevicesByLocation returns devices for a location",
     "1. Seed firstLocation in beforeAll\n2. Send authenticated GraphQL query: getDevicesByLocation(location: firstLocation)\n3. Assert HTTP 200\n4. Parse items array and assert length > 0",
     "All devices at the specified location are returned", "REGRESSION", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-011", "getDevice with non-existent ID returns null or error",
     "1. Send authenticated GraphQL query: getDevice(id: \"non-existent-id-999\")\n2. Assert HTTP 200\n3. Assert response.data.getDevice is null OR response.errors is defined",
     "API returns null data or a descriptive error for an unknown device ID", "NEGATIVE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-012", "listDevices with invalid status returns empty items or error",
     "1. Send authenticated GraphQL query: listDevices(status: \"INVALID_STATUS\")\n2. Assert HTTP 200\n3. Assert items array is empty OR response contains a validation error",
     "Invalid status filter returns empty items or a GraphQL validation error", "NEGATIVE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-013", "getDevicesByCustomer with non-existent customer returns empty items",
     "1. Send authenticated GraphQL query: getDevicesByCustomer(customerId: \"non-existent-cust-999\")\n2. Assert HTTP 200\n3. Parse items array and assert length === 0",
     "Empty items array returned for an unknown customer ID", "NEGATIVE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-014", "getDevicesByLocation with unknown location returns empty items",
     "1. Send authenticated GraphQL query: getDevicesByLocation(location: \"UNKNOWN_LOCATION_XYZ\")\n2. Assert HTTP 200\n3. Parse items array and assert length === 0",
     "Empty items array returned for an unknown location", "NEGATIVE", "Active"),
    ("API — Devices", "N/A (TypeScript spec)", "device.api.spec.ts",
     "API-DEV-015", "Unauthenticated listDevices request returns 401 or auth error",
     "1. Build GraphQL request for listDevices WITHOUT authentication token\n2. Send unauthenticated request to the GraphQL endpoint\n3. Assert HTTP 401 OR response contains an auth error",
     "Unauthenticated request is rejected with 401 or an auth error", "NEGATIVE", "Active"),

    # ── Section 6: API — Service Orders (13) ──────────────────────────────
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-001", "listServiceOrdersByStatus returns Pending orders",
     "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: \"Pending\")\n2. Assert HTTP 200 and no errors\n3. Parse items array and assert length > 0\n4. Assert all items have status === \"Pending\"",
     "Only Pending service orders are returned", "SMOKE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-002", "listServiceOrdersByStatus returns InProgress orders",
     "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: \"In Progress\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"In Progress\"",
     "Only In Progress service orders are returned", "REGRESSION", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-003", "listServiceOrdersByStatus returns Completed orders",
     "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: \"Completed\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Completed\"",
     "Only Completed service orders are returned", "REGRESSION", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-004", "listServiceOrdersByStatus returns Scheduled orders",
     "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: \"Scheduled\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Scheduled\"",
     "Only Scheduled service orders are returned", "REGRESSION", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-005", "listServiceOrdersByStatus response items have required fields",
     "1. Send authenticated GraphQL query: listServiceOrdersByStatus with any valid status\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.title, item.status, item.technicianId, item.scheduledDate are all present",
     "Each service order item contains all required fields", "SMOKE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-006", "listServiceOrdersByDate returns results within last 30 days",
     "1. Compute startDate = now - 30 days, endDate = now (ISO format)\n2. Send authenticated GraphQL query: listServiceOrdersByDate(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert all scheduledDate values fall within the 30-day range",
     "All returned orders have scheduledDate within the last 30 days", "REGRESSION", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-007", "listServiceOrdersByDate with future range returns empty items",
     "1. Compute startDate = now + 1 year, endDate = now + 2 years\n2. Send authenticated GraphQL query: listServiceOrdersByDate(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert length === 0",
     "Empty items array returned for a fully future date range", "NEGATIVE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-008", "getServiceOrder returns order by valid ID",
     "1. Seed a valid service order ID from listServiceOrdersByStatus\n2. Send authenticated GraphQL query: getServiceOrder(id: seededId)\n3. Assert HTTP 200\n4. Assert returned order.id === seededId",
     "Complete service order record returned for the queried ID", "SMOKE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-009", "getServiceOrdersByTechnician returns orders for a technician",
     "1. Seed a valid technicianId from listServiceOrdersByStatus\n2. Send GraphQL query: getServiceOrdersByTechnician(technicianId: seededId)\n3. Assert HTTP 200\n4. Parse items and assert length > 0",
     "All orders for the specified technician are returned", "REGRESSION", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-010", "getServiceOrder with non-existent ID returns null or error",
     "1. Send authenticated GraphQL query: getServiceOrder(id: \"non-existent-so-999\")\n2. Assert HTTP 200\n3. Assert response.data.getServiceOrder is null OR response.errors is defined",
     "null data or descriptive error returned for unknown service order ID", "NEGATIVE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-011", "listServiceOrdersByStatus with invalid status returns empty or error",
     "1. Send authenticated GraphQL query: listServiceOrdersByStatus(status: \"FAKE_STATUS\")\n2. Assert HTTP 200\n3. Assert items array is empty OR response.errors is defined",
     "Invalid status returns empty items or a validation error", "NEGATIVE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-012", "getServiceOrdersByTechnician with unknown ID returns empty items",
     "1. Send GraphQL query: getServiceOrdersByTechnician(technicianId: \"unknown-tech-999\")\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
     "Empty items array returned for unknown technician ID", "NEGATIVE", "Active"),
    ("API — Service Orders", "N/A (TypeScript spec)", "service-orders.api.spec.ts",
     "API-SO-013", "listServiceOrdersByDate with endDate before startDate returns empty or error",
     "1. Set startDate = now, endDate = now - 1 day\n2. Send authenticated GraphQL query: listServiceOrdersByDate(startDate, endDate)\n3. Assert HTTP 200\n4. Assert items is empty OR response.errors is defined",
     "Reversed date range returns empty results or a validation error", "NEGATIVE", "Active"),

    # ── Section 7: API — Firmware (15) ─────────────────────────────────────
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-001", "listFirmware returns all firmware with expected fields",
     "1. Send authenticated GraphQL query: listFirmware (no filter)\n2. Assert HTTP 200\n3. Assert no errors in response\n4. Parse items and assert length > 0",
     "Response is HTTP 200 with a populated firmware items array", "SMOKE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-002", "listFirmware response items have required fields",
     "1. Send authenticated GraphQL query: listFirmware(limit: 1)\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.name, item.version, item.status, item.deviceModel are present and non-null",
     "Each firmware item contains all required fields", "SMOKE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-003", "listFirmware filters by status=Pending",
     "1. Send authenticated GraphQL query: listFirmware(status: \"Pending\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Pending\"",
     "Only Pending firmware records are returned", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-004", "listFirmware filters by status=Approved",
     "1. Send authenticated GraphQL query: listFirmware(status: \"Approved\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Approved\"",
     "Only Approved firmware records are returned", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-005", "listFirmware filters by status=Deprecated",
     "1. Send authenticated GraphQL query: listFirmware(status: \"Deprecated\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Deprecated\"",
     "Only Deprecated firmware records are returned", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-006", "listFirmware with limit=2 returns at most 2 items",
     "1. Send authenticated GraphQL query: listFirmware(limit: 2)\n2. Assert HTTP 200\n3. Parse items array\n4. Assert items.length <= 2",
     "Response contains at most 2 items when limit=2 is specified", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-007", "listFirmware pagination: nextToken advances the cursor",
     "1. Send listFirmware(limit: 2) and capture nextToken from response\n2. Assert nextToken is defined and non-null\n3. Send listFirmware(limit: 2, nextToken: capturedToken)\n4. Assert HTTP 200 and response contains the next page of items",
     "nextToken from page 1 correctly advances to page 2 with different items", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-008", "getFirmware returns firmware by valid ID",
     "1. Seed a valid firmware ID from listFirmware in beforeAll\n2. Send authenticated GraphQL query: getFirmware(id: seededId)\n3. Assert HTTP 200\n4. Assert returned firmware.id === seededId",
     "Complete firmware record returned for the queried ID", "SMOKE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-009", "getFirmwareByModel returns firmware for a known model",
     "1. Seed a valid deviceModel from listFirmware in beforeAll\n2. Send authenticated GraphQL query: getFirmwareByModel(deviceModel: seededModel)\n3. Assert HTTP 200\n4. Parse items and assert length > 0",
     "Firmware records for the specified model are returned", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-010", "getFirmwareWithRelations returns data for valid ID",
     "1. Seed a valid firmware ID from beforeAll\n2. Send authenticated GraphQL query: getFirmwareWithRelations(id: seededId)\n3. Assert HTTP 200\n4. Assert returned data is non-null and includes related entity fields",
     "Firmware record with related entities is returned for the queried ID", "REGRESSION", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-011", "getFirmware with non-existent ID returns null or error",
     "1. Send authenticated GraphQL query: getFirmware(id: \"non-existent-fw-999\")\n2. Assert HTTP 200\n3. Assert response.data.getFirmware is null OR response.errors is defined",
     "null data or descriptive error returned for unknown firmware ID", "NEGATIVE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-012", "getFirmwareByModel with unknown model returns empty items",
     "1. Send authenticated GraphQL query: getFirmwareByModel(deviceModel: \"UNKNOWN_MODEL_XYZ\")\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
     "Empty items array returned for an unrecognised device model", "NEGATIVE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-013", "listFirmware with invalid status returns empty or error",
     "1. Send authenticated GraphQL query: listFirmware(status: \"INVALID_STATUS\")\n2. Assert HTTP 200\n3. Assert items is empty OR response.errors is defined",
     "Invalid status returns empty items or a validation error", "NEGATIVE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-014", "getFirmwareWithRelations with non-existent ID returns null or error",
     "1. Send authenticated GraphQL query: getFirmwareWithRelations(id: \"non-existent-fw-999\")\n2. Assert HTTP 200\n3. Assert data is null OR response.errors is defined",
     "null or error returned for unknown firmware ID with relations query", "NEGATIVE", "Active"),
    ("API — Firmware", "N/A (TypeScript spec)", "firmware.api.spec.ts",
     "API-FW-015", "Unauthenticated listFirmware request is rejected",
     "1. Build GraphQL request for listFirmware WITHOUT authentication token\n2. Send unauthenticated request to the GraphQL endpoint\n3. Assert HTTP 401 OR response contains an authentication error",
     "Unauthenticated request is rejected with 401 or an auth error", "NEGATIVE", "Active"),

    # ── Section 8: API — Compliance (12) ───────────────────────────────────
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-001", "listComplianceByStatus Approved returns only approved records",
     "1. Send authenticated GraphQL query: listComplianceByStatus(status: \"Approved\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Approved\"",
     "Only Approved compliance records are returned", "SMOKE", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-002", "listComplianceByStatus Pending returns only pending records",
     "1. Send authenticated GraphQL query: listComplianceByStatus(status: \"Pending\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Pending\"",
     "Only Pending compliance records are returned", "SMOKE", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-003", "listComplianceByStatus Deprecated returns only deprecated records",
     "1. Send authenticated GraphQL query: listComplianceByStatus(status: \"Deprecated\")\n2. Assert HTTP 200\n3. Parse items and assert all have status === \"Deprecated\"",
     "Only Deprecated compliance records are returned", "REGRESSION", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-004", "listComplianceByStatus response items have required fields",
     "1. Send listComplianceByStatus with any valid status, limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.firmwareId, item.status, item.certifications are all present",
     "Each compliance item contains all required fields", "SMOKE", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-005", "getCompliance returns record by valid ID",
     "1. Seed a valid compliance ID from listComplianceByStatus in beforeAll\n2. Send authenticated GraphQL query: getCompliance(id: seededId)\n3. Assert HTTP 200\n4. Assert returned record.id === seededId",
     "Complete compliance record returned for the queried ID", "SMOKE", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-006", "getComplianceByCertification returns records for CE certification",
     "1. Send authenticated GraphQL query: getComplianceByCertification(certification: \"CE\")\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
     "Compliance records with CE certification are returned", "REGRESSION", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-007", "getComplianceByCertification returns records for FCC certification",
     "1. Send authenticated GraphQL query: getComplianceByCertification(certification: \"FCC\")\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
     "Compliance records with FCC certification are returned", "REGRESSION", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-008", "getComplianceByCertification returns records for UL certification",
     "1. Send authenticated GraphQL query: getComplianceByCertification(certification: \"UL\")\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
     "Compliance records with UL certification are returned", "REGRESSION", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-009", "getComplianceByCertification returned items contain the requested cert",
     "1. Send authenticated GraphQL query: getComplianceByCertification(certification: \"CE\")\n2. Assert HTTP 200\n3. Parse items\n4. Assert every item's certifications array includes \"CE\"",
     "All returned items contain the requested certification in their array", "REGRESSION", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-010", "getCompliance with non-existent ID returns null or error",
     "1. Send authenticated GraphQL query: getCompliance(id: \"non-existent-comp-999\")\n2. Assert HTTP 200\n3. Assert response.data.getCompliance is null OR response.errors is defined",
     "null data or descriptive error returned for unknown compliance ID", "NEGATIVE", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-011", "listComplianceByStatus with invalid status returns empty or error",
     "1. Send authenticated GraphQL query: listComplianceByStatus(status: \"INVALID_STATUS\")\n2. Assert HTTP 200\n3. Assert items is empty OR response.errors is defined",
     "Invalid status returns empty items or a validation error", "NEGATIVE", "Active"),
    ("API — Compliance", "N/A (TypeScript spec)", "compliance.api.spec.ts",
     "API-COMP-012", "getComplianceByCertification with unknown certification returns empty",
     "1. Send authenticated GraphQL query: getComplianceByCertification(certification: \"UNKNOWN_CERT\")\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
     "Empty items array returned for an unrecognised certification", "NEGATIVE", "Active"),

    # ── Section 9: API — Audit Logs (12) ───────────────────────────────────
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-001", "listAuditLogs returns records for last 24 hours",
     "1. Compute startDate = now - 24h, endDate = now (ISO format)\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Assert no errors\n5. Assert listAuditLogs result and items are defined",
     "Response is HTTP 200 with items defined for the 24-hour window", "SMOKE", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-002", "listAuditLogs returns records for last 7 days",
     "1. Compute startDate = now - 7 days, endDate = now\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Assert no errors and items are defined",
     "Response is HTTP 200 with items defined for the 7-day window", "REGRESSION", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-003", "listAuditLogs returns records for last 30 days",
     "1. Compute startDate = now - 30 days, endDate = now\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert length > 0",
     "Audit log records exist and are returned for the 30-day window", "REGRESSION", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-004", "listAuditLogs response items have required fields",
     "1. Send authenticated GraphQL query: listAuditLogs for last 30 days with limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id is truthy\n5. Assert item.action is truthy\n6. Assert at least one timestamp field exists (timestamp/createdAt/eventTime/date)",
     "Each audit log item contains id, action, and a timestamp field", "SMOKE", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-005", "listAuditLogs with limit=5 returns at most 5 items",
     "1. Send authenticated GraphQL query: listAuditLogs for last 30 days with limit=5\n2. Assert HTTP 200\n3. Parse items array\n4. Assert items.length <= 5",
     "Response contains at most 5 items when limit=5 is specified", "REGRESSION", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-006", "listAuditLogs timestamps are within requested range",
     "1. Compute startDate = now - 7 days, endDate = now\n2. Send listAuditLogs(startDate, endDate)\n3. Parse items\n4. For each item assert its timestamp >= startDate and <= endDate",
     "All returned audit log timestamps fall within the specified date range", "REGRESSION", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-007", "listAuditLogs totalCount is returned",
     "1. Send authenticated GraphQL query: listAuditLogs for last 30 days\n2. Assert HTTP 200\n3. Assert response.data.listAuditLogs.totalCount is a number >= 0",
     "totalCount field is present and is a non-negative integer", "REGRESSION", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-008", "getAuditLogsByUser returns logs for a known user",
     "1. Seed existingUserId from listAuditLogs in beforeAll\n2. Send authenticated GraphQL query: getAuditLogsByUser(userId: existingUserId)\n3. Assert HTTP 200\n4. Parse items and assert at least one item is returned",
     "Audit log entries for the specified user are returned", "SMOKE", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-009", "listAuditLogs returns empty for future date range",
     "1. Compute startDate = now + 1 year, endDate = now + 2 years\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Parse items and assert length === 0",
     "Empty items array returned for a fully future date range", "NEGATIVE", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-010", "listAuditLogs with reversed date range returns empty or error",
     "1. Set startDate = now, endDate = now - 1 day\n2. Send authenticated GraphQL query: listAuditLogs(startDate, endDate)\n3. Assert HTTP 200\n4. Assert items is empty OR response.errors is defined",
     "Reversed date range returns empty results or a validation error", "NEGATIVE", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-011", "getAuditLogsByUser with non-existent userId returns empty",
     "1. Send authenticated GraphQL query: getAuditLogsByUser(userId: \"unknown-user-999\")\n2. Assert HTTP 200\n3. Parse items and assert length === 0",
     "Empty items array returned for an unknown user ID", "NEGATIVE", "Active"),
    ("API — Audit Logs", "N/A (TypeScript spec)", "audit-logs.api.spec.ts",
     "API-AUDIT-012", "listAuditLogs with invalid ISO date format returns error",
     "1. Send authenticated GraphQL query: listAuditLogs(startDate: \"not-a-date\", endDate: \"also-invalid\")\n2. Assert response.errors is defined (GraphQL validation or parsing error)",
     "A GraphQL validation or parsing error is returned for malformed ISO dates", "NEGATIVE", "Active"),

    # ── Section 10: API — Users & Customers (12) ───────────────────────────
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-001", "getUserByEmail returns user profile for known email",
     "1. Seed a known registered email from the test environment\n2. Send authenticated GraphQL query: getUserByEmail(email: seededEmail)\n3. Assert HTTP 200\n4. Assert no errors\n5. Assert returned user object is defined",
     "Complete user profile returned for the queried email address", "SMOKE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-002", "getUserByEmail response has required fields",
     "1. Send authenticated GraphQL query: getUserByEmail(email: seededEmail)\n2. Assert HTTP 200\n3. Assert returned user.id, user.email, user.name, user.role are all present",
     "Returned user profile contains all required fields: id, email, name, role", "SMOKE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-003", "listUsersByRole Admin returns at least one admin user",
     "1. Send authenticated GraphQL query: listUsersByRole(role: \"Admin\")\n2. Assert HTTP 200\n3. Parse items and assert length > 0",
     "At least one Admin user is returned", "SMOKE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-004", "listUsersByRole returns a response for role=Admin",
     "1. Send authenticated GraphQL query: listUsersByRole(role: \"Admin\")\n2. Assert HTTP 200\n3. Assert response.data.listUsersByRole is defined",
     "A valid response is returned for role=Admin", "REGRESSION", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-005", "listUsersByRole returns a response for role=Technician",
     "1. Send authenticated GraphQL query: listUsersByRole(role: \"Technician\")\n2. Assert HTTP 200\n3. Assert response.data.listUsersByRole is defined",
     "A valid response is returned for role=Technician", "REGRESSION", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-006", "listUsersByRole returns a response for role=Manager",
     "1. Send authenticated GraphQL query: listUsersByRole(role: \"Manager\")\n2. Assert HTTP 200\n3. Assert response.data.listUsersByRole is defined",
     "A valid response is returned for role=Manager", "REGRESSION", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-007", "listUsersByRole items have required fields",
     "1. Send authenticated GraphQL query: listUsersByRole(role: \"Admin\") with limit=1\n2. Assert HTTP 200\n3. Parse first item\n4. Assert item.id, item.email, item.name, item.role are all present",
     "Each user item contains all required fields: id, email, name, role", "REGRESSION", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-008", "getCustomerWithRelations returns data for a known customer",
     "1. Seed a known customerId from the test environment\n2. Send authenticated GraphQL query: getCustomerWithRelations(customerId: seededId)\n3. Assert HTTP 200\n4. Assert returned customer data is non-null and includes related devices/locations",
     "Customer record with related entities is returned for the queried ID", "SMOKE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-009", "getUserByEmail with non-existent email returns null or error",
     "1. Send authenticated GraphQL query: getUserByEmail(email: \"nobody@nonexistent999.com\")\n2. Assert HTTP 200\n3. Assert response.data.getUserByEmail is null OR response.errors is defined",
     "null data or descriptive error returned for an unregistered email", "NEGATIVE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-010", "listUsersByRole with unknown role returns 200",
     "1. Send authenticated GraphQL query: listUsersByRole(role: \"UnknownRole\")\n2. Assert HTTP 200 (no server-side role validation enforced)\n3. Assert response does not throw an unhandled error",
     "HTTP 200 returned for unknown role (no server-side role enum validation)", "NEGATIVE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-011", "getCustomerWithRelations with non-existent customerId returns empty or error",
     "1. Send authenticated GraphQL query: getCustomerWithRelations(customerId: \"non-existent-cust-999\")\n2. Assert HTTP 200\n3. Assert data is null/empty OR response.errors is defined",
     "null/empty data or error returned for unknown customer ID", "NEGATIVE", "Active"),
    ("API — Users & Customers", "N/A (TypeScript spec)", "users.api.spec.ts",
     "API-USER-012", "getUserByEmail with malformed email format returns null or error",
     "1. Send authenticated GraphQL query: getUserByEmail(email: \"not-an-email\")\n2. Assert HTTP 200\n3. Assert response.data.getUserByEmail is null OR response.errors is defined",
     "null or validation error returned for a malformed email string", "NEGATIVE", "Active"),

    # ── Section 11: API — Mutations (16) ───────────────────────────────────
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-001", "Creates a new firmware record with Pending status",
     "1. Generate unique payload: name=\"Test-FW-<timestamp>\", version=\"1.0.0\", deviceModel=\"ModelX\"\n2. Send authenticated GraphQL mutation: createFirmware(payload)\n3. Assert HTTP 200\n4. Assert returned firmware.id is truthy\n5. Assert returned firmware.status === \"Pending\"",
     "New firmware record created with Pending status and a valid ID", "SMOKE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-002", "createFirmware with minimal required fields succeeds",
     "1. Build payload with only name and version fields\n2. Send authenticated GraphQL mutation: createFirmware(minimalPayload)\n3. Assert HTTP 200 or graceful skip if resolver rejects missing optional fields\n4. Assert returned record has id and status=\"Pending\"",
     "Firmware creation succeeds with only the minimum required fields", "REGRESSION", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-003", "createFirmware without required name returns GraphQL error",
     "1. Build payload omitting the name field\n2. Send authenticated GraphQL mutation: createFirmware(payloadWithoutName)\n3. Assert response.errors is defined\n4. Assert error message references missing name field",
     "GraphQL field-required error returned when name is omitted", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-004", "createFirmware without required version returns GraphQL error",
     "1. Build payload omitting the version field\n2. Send authenticated GraphQL mutation: createFirmware(payloadWithoutVersion)\n3. Assert response.errors is defined\n4. Assert error message references missing version field",
     "GraphQL field-required error returned when version is omitted", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-005", "Creates a new service order with Pending status",
     "1. Generate unique payload: title=\"Test-SO-<timestamp>\", priority=\"Medium\"\n2. Send authenticated GraphQL mutation: createServiceOrder(payload)\n3. Assert HTTP 200\n4. Assert returned order.id is truthy\n5. Assert returned order.status === \"Pending\"",
     "New service order created with Pending status and a valid ID", "SMOKE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-006", "createServiceOrder without required title returns GraphQL error",
     "1. Build payload omitting the title field\n2. Send authenticated GraphQL mutation: createServiceOrder(payloadWithoutTitle)\n3. Assert response.errors is defined",
     "GraphQL field-required error returned when title is omitted", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-007", "Creates a compliance submission in Pending status",
     "1. Seed a valid firmwareId from listFirmware\n2. Generate unique payload: firmwareId, firmwareVersion, deviceModel, certifications=[\"CE\"]\n3. Send authenticated GraphQL mutation: createCompliance(payload)\n4. Assert HTTP 200\n5. Assert returned compliance.id is truthy\n6. Assert returned compliance.status === \"Pending\"",
     "New compliance record created with Pending status and a valid ID", "SMOKE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-008", "createCompliance without required firmwareId returns error",
     "1. Build payload omitting the firmwareId field\n2. Send authenticated GraphQL mutation: createCompliance(payloadWithoutFirmwareId)\n3. Assert response.errors is defined",
     "GraphQL field-required error returned when firmwareId is omitted", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-009", "Approves a Pending firmware record",
     "1. Create a new firmware record (status=Pending) via createFirmware mutation\n2. Send authenticated mutation: updateEntityStatus(entityId, entityType:\"firmware\", newStatus:\"Approved\")\n3. Assert HTTP 200\n4. Assert returned entity status === \"Approved\"",
     "Firmware record successfully transitions from Pending to Approved", "SMOKE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-010", "Deprecates a firmware record",
     "1. Create a new firmware record (status=Pending) via createFirmware mutation\n2. Send authenticated mutation: updateEntityStatus(entityId, entityType:\"firmware\", newStatus:\"Deprecated\")\n3. Assert HTTP 200\n4. Assert returned entity status === \"Deprecated\"",
     "Firmware record successfully transitions to Deprecated status", "REGRESSION", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-011", "updateEntityStatus with non-existent ID returns 200 (upsert behaviour)",
     "1. Send authenticated mutation: updateEntityStatus(entityId:\"non-existent-id-999\", entityType:\"firmware\", newStatus:\"Approved\")\n2. Assert HTTP 200 (backend upserts on unknown ID)\n3. Assert no unhandled error in response",
     "HTTP 200 returned for unknown entity ID due to upsert behaviour", "REGRESSION", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-012", "updateEntityStatus with invalid newStatus returns 200",
     "1. Create a firmware record and capture its ID\n2. Send mutation: updateEntityStatus(entityId, entityType:\"firmware\", newStatus:\"INVALID_STATUS\")\n3. Assert HTTP 200 (no server-side enum validation)\n4. Assert no unhandled error thrown",
     "HTTP 200 returned for invalid newStatus value (no server-side enum enforcement)", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-013", "updateEntityStatus without required entityType returns GraphQL error",
     "1. Build mutation payload omitting entityType\n2. Send authenticated mutation: updateEntityStatus(entityId, newStatus:\"Approved\")\n3. Assert response.errors is defined",
     "GraphQL field-required error returned when entityType is omitted", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-014", "Updates device coordinates with valid address and lat/lng",
     "1. Seed a valid deviceId from listDevices\n2. Send authenticated mutation: updateDeviceCoords(deviceId, address:\"123 Test St\", lat:37.7749, lng:-122.4194)\n3. Assert HTTP 200\n4. Assert returned device location/coords fields are updated",
     "Device coordinates and address are successfully updated", "SMOKE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-015", "updateDeviceCoords without required address returns GraphQL error",
     "1. Seed a valid deviceId\n2. Build mutation payload omitting the address field\n3. Send authenticated mutation: updateDeviceCoords(deviceId, lat:37.7749, lng:-122.4194)\n4. Assert response.errors is defined",
     "GraphQL field-required error returned when address is omitted", "NEGATIVE", "Active"),
    ("API — Mutations", "N/A (TypeScript spec)", "mutations.api.spec.ts",
     "API-MUT-016", "updateDeviceCoords with out-of-range coordinates returns error",
     "1. Seed a valid deviceId\n2. Send authenticated mutation: updateDeviceCoords(deviceId, address:\"Test\", lat:999, lng:999)\n3. Assert response.errors is defined or HTTP 400 returned\n4. Assert error references invalid coordinate range",
     "Validation error returned when lat/lng values exceed valid geographic range", "NEGATIVE", "Active"),

    # ── Section 12: UI — Dashboard Recent Alerts (12) ──────────────────────
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-01", "Recent Alerts panel is visible after login",
     "1. Verify div.bg-card:has(h3:text-is('Recent Alerts')) is visible\n2. Verify panel contains text 'Recent Alerts'\n3. Verify no KPI loading placeholder '—' remains",
     "Alerts panel container rendered and visible; KPI cards have resolved", "SMOKE", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-02", "Panel displays audit-log entries from last 24 hours",
     "1. Register route to inject one approved audit-log entry (listAuditLogs)\n2. Re-navigate to Dashboard\n3. Verify div.bg-muted alert item row is visible\n4. Verify 'No recent activity' is NOT visible\n5. Unroute (finally)",
     "At least one alert item row rendered; empty-state absent", "SMOKE", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-03", "Alert item displays action text and timestamp",
     "1. Register route to inject entry with action 'Firmware update approved'\n2. Re-navigate to Dashboard\n3. Verify alert item row is visible\n4. Verify panel containsText 'Firmware update approved'\n5. Assert JS-evaluated item text is non-empty (timestamp present)\n6. Unroute (finally)",
     "Action text present in panel; alert item has non-empty text including timestamp", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-04", "Failed actions display orange warning icon",
     "1. Register route to inject failed-action entry (auditStatus='failed')\n2. Re-navigate to Dashboard\n3. Verify alert item row is visible\n4. Verify svg[class*='text-orange'] is visible inside div.bg-muted\n5. Unroute (finally)",
     "Orange SVG icon rendered in failed-action alert row", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-05", "Approved/approve actions display green check icon",
     "1. Register route to inject approved-action entry (auditStatus='approved')\n2. Re-navigate to Dashboard\n3. Verify alert item row is visible\n4. Verify svg[class*='text-green'] is visible inside div.bg-muted\n5. Unroute (finally)",
     "Green SVG icon rendered in approved-action alert row", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-06", "Other (info) actions display blue info icon",
     "1. Register route to inject info-action entry (auditStatus='info')\n2. Re-navigate to Dashboard\n3. Verify alert item row is visible\n4. Verify svg[class*='text-blue'] is visible inside div.bg-muted\n5. Unroute (finally)",
     "Blue SVG icon rendered in info-action alert row", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-07", "'View all' link visible and navigates to /analytics",
     "1. Verify a[href='/analytics'] inside Alerts panel is visible\n2. Click the 'View all' link\n3. Assert window.location.pathname equals '/analytics'",
     "'View all' link present; clicking routes browser to /analytics", "SMOKE", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-08", "Loading state — 'Loading alerts...' shown while API pending",
     "1. Register route to delay all AppSync responses by 2 s (background Thread)\n2. Re-navigate to Dashboard\n3. Verify 'Loading alerts' substring is visible in panel\n4. Unroute (finally)",
     "'Loading alerts...' text visible while audit-log API is in-flight", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-09", "Empty state — 'No recent activity' when no logs exist",
     "1. Register route to return empty listAuditLogs (items=[])\n2. Re-navigate to Dashboard\n3. Verify 'No recent activity' text is visible\n4. Verify no div.bg-muted alert item rows are present\n5. Verify panel container is still visible\n6. Unroute (finally)",
     "Empty-state message shown; no item rows; panel still rendered", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-10", "Error state — div.bg-red-50 shown when audit-log API returns 500",
     "1. Register route to return HTTP 500 only for listAuditLogs\n2. Re-navigate to Dashboard\n3. Verify div.bg-red-50 inside alerts card is visible\n4. Verify Total Devices KPI still shows a value (KPI unaffected)\n5. Verify global KPI error banner NOT visible\n6. Unroute (finally)",
     "Alerts card-level error element visible; KPI domain unaffected; no global error banner", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-11", "Alerts panel isolated from KPI API failures",
     "1. Confirm Dashboard loaded (Total Devices visible)\n2. Register route to return 500 for all fetch/xhr\n3. Click Refresh Dashboard\n4. Verify global KPI error banner appears\n5. Verify Alerts panel container still visible\n6. Unroute (finally)",
     "Global error banner shown; Alerts panel still rendered — separate error domains confirmed", "REGRESSION", "Active"),
    ("UI — Dashboard Recent Alerts", "dashboard-alerts-suite.xml", "DashboardAlertsTests",
     "TC-ALERTS-12", "Alerts panel updates correctly after Refresh Dashboard",
     "1. Verify Alerts panel visible before refresh\n2. Verify Refresh Dashboard button visible\n3. Click Refresh Dashboard\n4. Wait for KPI cards to repopulate\n5. Verify Alerts panel still visible post-refresh\n6. Verify no global error banner",
     "Panel remains rendered after refresh; KPI cards repopulate; no error banner", "SMOKE", "Active"),

    # ── Section 13: UI — Dashboard System Status (14) ──────────────────────
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-01", "System Status panel visible after login",
     "1. Verify div.bg-card:has(h3:text-is('System Status')) is visible\n2. Verify panel containsText 'System Status'",
     "Panel container rendered and heading text present", "SMOKE", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-02", "Panel displays exactly 4 service rows",
     "1. Verify SERVICE_ITEM is visible\n2. JS evaluate count of div.space-y-4 > div — assert equals 4\n3. Verify panel contains 'Deployment Service', 'Compliance Engine', 'Asset Database', 'Analytics Platform'",
     "Exactly 4 rows; all 4 service names present", "SMOKE", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-03", "Each service row has one status label (Operational or Degraded)",
     "1. Wait for KPI data to load\n2. JS evaluate count of span.text-green-600 + span.text-orange-600 — assert sum equals 4\n3. Assert at least one label visible",
     "Total green + orange label count equals 4; at least one visible", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-04", "Each service row displays a progress bar",
     "1. Wait for Dashboard load\n2. JS evaluate count of progress bar elements — assert >= 4\n3. Assert at least one bar has non-zero width",
     "4 progress bars present; at least one has non-zero computed width", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-05", "Analytics Platform always shows 'Operational'",
     "1. Wait for Dashboard load\n2. Verify ANALYTICS_PLATFORM_OPERATIONAL (span.text-green-600) visible\n3. Verify ANALYTICS_PLATFORM_DEGRADED (span.text-orange-600) NOT visible\n4. Verify text 'Operational' present in row",
     "Green label present; orange label absent; 'Operational' text confirmed", "SMOKE", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-06", "Deployment Service 'Operational' when in-progress orders exist",
     "1. Register route returning 1 in-progress service order\n2. Re-navigate to Dashboard\n3. Wait for Active Deployments KPI\n4. Verify DEPLOYMENT_SERVICE_OPERATIONAL visible\n5. Unroute (finally)",
     "span.text-green-600 visible in Deployment Service row", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-07", "Deployment Service 'Degraded' when no in-progress orders",
     "1. Register route returning empty in-progress orders\n2. Re-navigate to Dashboard\n3. Wait for Active Deployments KPI\n4. Verify DEPLOYMENT_SERVICE_DEGRADED visible\n5. Unroute (finally)",
     "span.text-orange-600 visible in Deployment Service row", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-08", "Compliance Engine 'Operational' when no pending compliance items",
     "1. Register route returning empty compliance items\n2. Re-navigate to Dashboard\n3. Wait for Pending Approvals KPI\n4. Verify COMPLIANCE_ENGINE_OPERATIONAL visible\n5. Unroute (finally)",
     "span.text-green-600 visible in Compliance Engine row", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-09", "Compliance Engine 'Degraded' when pending compliance items exist",
     "1. Register route returning 1 pending compliance item\n2. Re-navigate to Dashboard\n3. Wait for Pending Approvals KPI\n4. Verify COMPLIANCE_ENGINE_DEGRADED visible\n5. Unroute (finally)",
     "span.text-orange-600 visible in Compliance Engine row", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-10", "Asset Database reflects average device health score",
     "1. Register route with high-health devices (avg 96.5%)\n2. Re-navigate — verify ASSET_DATABASE_OPERATIONAL visible\n3. Switch to low-health devices (avg 50%) via Refresh Dashboard\n4. Verify ASSET_DATABASE_DEGRADED visible\n5. Unroute (finally)",
     "Green label for >=90% health; orange label for <90% health", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-11", "Panel visible when KPI error banner appears",
     "1. Confirm Dashboard loaded\n2. Register route returning 500 for all fetch/xhr\n3. Click Refresh Dashboard\n4. Verify global error banner visible\n5. Verify SystemStatus.CONTAINER still visible\n6. Unroute (finally)",
     "Error banner shown; System Status panel still rendered", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-12", "Panel updates correctly after Refresh Dashboard",
     "1. Verify panel visible before refresh\n2. Click Refresh Dashboard\n3. Wait for KPI cards to repopulate\n4. Verify panel still visible\n5. Assert at least one status label visible post-refresh",
     "Panel rendered after refresh; status labels re-populated from fresh KPI data", "SMOKE", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-13", "Service rows appear in correct order",
     "1. Wait for Dashboard load\n2. JS evaluate extract row text array — assert Row 1 contains 'Deployment Service', Row 4 contains 'Analytics Platform'\n3. Verify nth-child(1) containsText 'Deployment Service'\n4. Verify nth-child(4) containsText 'Analytics Platform'",
     "Rows in order: Deployment Service -> Compliance Engine -> Asset Database -> Analytics Platform", "REGRESSION", "Active"),
    ("UI — Dashboard System Status", "dashboard-system-status-suite.xml", "DashboardSystemStatusTests",
     "TC-STATUS-14", "Operational uses text-green-600, Degraded uses text-orange-600",
     "1. Verify OPERATIONAL_LABEL visible; JS assert className includes 'text-green-600'\n2. Register route with zero orders to force Deployment Service Degraded\n3. Re-navigate — verify DEPLOYMENT_SERVICE_DEGRADED visible\n4. Unroute (finally)",
     "Operational span className = text-green-600; Degraded span className = text-orange-600", "REGRESSION", "Active"),

    # ── Section 14: UI — Integration (4) ───────────────────────────────────
    ("UI — Dashboard Alerts & System Status Integration",
     "dashboard-alerts-system-status-integration-suite.xml",
     "DashboardAlertsSystemStatusIntegrationTests",
     "TC-INTEGRATION-01", "Both panels visible simultaneously on Dashboard",
     "1. Verify AlertsPanel.CONTAINER is visible\n2. Verify SystemStatus.CONTAINER is visible\n3. Verify KpiCard.KPI_GRID_CONTAINER is visible",
     "All three content sections coexist on Dashboard without layout breakage", "SMOKE", "Active"),
    ("UI — Dashboard Alerts & System Status Integration",
     "dashboard-alerts-system-status-integration-suite.xml",
     "DashboardAlertsSystemStatusIntegrationTests",
     "TC-INTEGRATION-02", "'View all' navigates to /analytics; back-navigation works",
     "1. Verify VIEW_ALL_LINK visible\n2. Click 'View all' link\n3. Assert URL path equals '/analytics'\n4. Launch Dashboard URL\n5. Verify Dashboard h1 header visible",
     "'View all' routes to /analytics; Dashboard loads correctly on return", "SMOKE", "Active"),
    ("UI — Dashboard Alerts & System Status Integration",
     "dashboard-alerts-system-status-integration-suite.xml",
     "DashboardAlertsSystemStatusIntegrationTests",
     "TC-INTEGRATION-03", "Both panels render correctly after Refresh Dashboard",
     "1. Verify Alerts panel, System Status panel, and Total Devices KPI loaded\n2. Click Refresh Dashboard\n3. Wait for KPI cards to repopulate\n4. Verify Alerts panel still visible\n5. Verify System Status panel still visible\n6. Verify no global error banner",
     "Both panels rendered and functional after refresh cycle", "SMOKE", "Active"),
    ("UI — Dashboard Alerts & System Status Integration",
     "dashboard-alerts-system-status-integration-suite.xml",
     "DashboardAlertsSystemStatusIntegrationTests",
     "TC-INTEGRATION-04", "Zero data state — both panels handle empty APIs gracefully",
     "1. Register route returning empty items for all GraphQL queries\n2. Re-navigate to Dashboard\n3. Verify Total Devices KPI shows '0'\n4. Verify 'No recent activity' shown in Alerts panel\n5. Verify System Status panel visible\n6. Verify no global error banner\n7. Unroute (finally)",
     "KPI shows '0'; Alerts shows empty state; System Status present; no error banner", "REGRESSION", "Active"),

    # ── Section 15: UI — Dashboard Quick Actions (25) ──────────────────────
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-QA-01", "Four Quick Action cards displayed on Dashboard",
     "1. Re-navigate to Dashboard for clean state\n2. Verify 'Welcome back' h2 heading visible\n3. Verify main a[href='/inventory'] (View Inventory card) visible\n4. Verify main a[href='/account-service'] (Schedule Service card) visible\n5. Verify a[href='/deployment'].relative.bg-card (Deploy Firmware card) visible\n6. Verify main a[href='/compliance'] (Check Compliance card) visible\n7. JS evaluate count of all 4 card selectors — assert equals 4",
     "All four Quick Action cards are visible in the main content area; exactly 4 cards are present", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-QA-02", "Each Quick Action card displays a label and SVG icon",
     "1. Verify 'View Inventory' label text visible inside card\n2. Verify SVG icon present inside View Inventory card\n3. Verify 'Schedule Service' label text visible\n4. Verify SVG icon present in Schedule Service card\n5. Verify 'Deploy Firmware' label text visible\n6. Verify SVG icon present in Deploy Firmware card\n7. Verify 'Check Compliance' label text visible\n8. Verify SVG icon present in Check Compliance card",
     "Each card renders its descriptive label and an SVG icon element", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-NAV-01", "'View Inventory' card navigates to /inventory",
     "1. Re-navigate to Dashboard\n2. Verify View Inventory card visible\n3. Click main a[href='/inventory']\n4. Assert browser.url() contains '/inventory'\n5. Verify Dashboard h1 is NOT visible (navigated away)",
     "URL contains '/inventory'; Inventory page renders; Dashboard header absent", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-NAV-02", "'Schedule Service' card navigates to /account-service",
     "1. Re-navigate to Dashboard\n2. Verify Schedule Service card visible\n3. Click main a[href='/account-service']\n4. Assert URL contains '/account-service'\n5. Verify Dashboard h1 NOT visible",
     "URL contains '/account-service'; Account & Service page renders", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-NAV-03", "'Deploy Firmware' card navigates to /deployment",
     "1. Re-navigate to Dashboard\n2. Verify Deploy Firmware card visible\n3. Click a[href='/deployment'].relative.bg-card\n4. Assert URL contains '/deployment'\n5. Verify Dashboard h1 NOT visible",
     "URL contains '/deployment'; Deployment page renders", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-NAV-04", "'Check Compliance' card navigates to /compliance",
     "1. Re-navigate to Dashboard\n2. Verify Check Compliance card visible\n3. Click main a[href='/compliance']\n4. Assert URL contains '/compliance'\n5. Verify Dashboard h1 NOT visible",
     "URL contains '/compliance'; Compliance page renders", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-01", "'View Inventory' badge shows offline device count",
     "1. Re-navigate to Dashboard and wait for badges to resolve\n2. Verify View Inventory card visible\n3. Check if main a[href='/inventory'] span.bg-orange-500 is visible\n4. If visible: read text, parse as int, assert > 0\n5. If absent: confirm offline count is 0 (badge correctly hidden)",
     "If offline devices exist, badge is visible with a positive integer count; if none exist, badge is absent", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-02", "'Schedule Service' badge shows scheduled orders count",
     "1. Re-navigate to Dashboard\n2. Verify Schedule Service card visible\n3. Check if main a[href='/account-service'] span.bg-orange-500 is visible\n4. If visible: parse text as int, assert > 0\n5. If absent: confirm scheduled orders count is 0",
     "If scheduled orders exist, badge visible with positive integer; if none, badge absent", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-03", "'Deploy Firmware' badge shows pending firmware count",
     "1. Re-navigate to Dashboard\n2. Verify Deploy Firmware card visible\n3. Check if a[href='/deployment'].relative.bg-card span.bg-orange-500 is visible\n4. If visible: parse text as int, assert > 0\n5. If absent: confirm pending firmware count is 0",
     "If pending firmware items exist, badge visible with positive integer; if none, badge absent", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-04", "'Check Compliance' badge shows pending compliance count",
     "1. Re-navigate to Dashboard\n2. Verify Check Compliance card visible\n3. Check if main a[href='/compliance'] span.bg-orange-500 is visible\n4. If visible: parse text as int, assert > 0\n5. If absent: confirm pending compliance count is 0",
     "If pending compliance records exist, badge visible with positive integer; if none, badge absent", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-05", "Badges hidden when all counts are zero (API mock)",
     "1. Register route intercepting all 4 badge GraphQL ops to return count=0\n2. Navigate to Dashboard with mocked responses\n3. Verify VIEW_INVENTORY_BADGE not visible\n4. Verify SCHEDULE_SERVICE_BADGE not visible\n5. Verify DEPLOY_FIRMWARE_BADGE not visible\n6. Verify CHECK_COMPLIANCE_BADGE not visible\n7. Unroute (finally)",
     "No orange badge spans appear on any card when all API counts are zero", "REGRESSION", "Disabled (requires zero-count env)"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-06", "Visible badge values are positive integers only",
     "1. Re-navigate to Dashboard, wait for badges to resolve\n2. Verify welcome heading visible (page loaded gate)\n3. JS evaluate all main span.bg-orange-500 elements\n4. For each: parse as int, assert not NaN, not <= 0, text matches integer string exactly",
     "All visible badge elements contain parseable positive integers; no NaN, decimals, negatives, or undefined strings", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-BADGE-07", "Sidebar badges isolated from Quick Action card badges",
     "1. Re-navigate to Dashboard, wait for sidebar and main content\n2. Verify sidebar nav Inventory link visible\n3. JS evaluate — assert no nav span.bg-orange-500 is inside main and vice versa\n4. Compare badge text for /inventory between sidebar and main card — assert match",
     "nav-scoped and main-scoped badge selectors do not overlap; badge counts are consistent between sidebar and Quick Action card", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-WELCOME-01", "Welcome message shows logged-in user email",
     "1. Re-navigate to Dashboard\n2. Verify h2:has-text('Welcome back') visible\n3. Verify welcome message containsText(adminUsername) — 'ajaykumar.yadav@3pillarglobal.com'",
     "Welcome h2 contains the admin email address after login", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-WELCOME-02", "Welcome message falls back to 'Admin' when no email (session mock)",
     "1. Strip email attribute from Cognito session in localStorage\n2. Reload Dashboard\n3. Verify 'Welcome back' h2 visible\n4. Verify h2 containsText('Admin') — fallback rendered",
     "h2 reads 'Welcome back, Admin' when user email is absent from the session", "REGRESSION", "Disabled (requires session mock)"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-WELCOME-03", "Welcome message format and immediate visibility",
     "1. Re-navigate to Dashboard\n2. Verify h2:has-text('Welcome back') visible (no loading delay)\n3. Verify h2 containsText('Welcome back,') — comma present\n4. JS read h2 textContent, assert matches regex `Welcome back,\\s+.+`",
     "Welcome message is immediately visible; format is 'Welcome back, {identifier}' with correct comma and space", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-REFRESH-01", "Refresh button visible with aria-label and SVG icon",
     "1. Re-navigate to Dashboard\n2. Verify button[aria-label='Refresh dashboard'] visible\n3. Verify button[aria-label='Refresh dashboard'] svg visible\n4. JS evaluate button.disabled — assert false",
     "Refresh button is visible in the welcome row, contains SVG icon, and is not disabled", "SMOKE", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-REFRESH-02", "Refresh button re-fetches all dashboard data",
     "1. Re-navigate to Dashboard; verify Total Devices KPI populated before refresh\n2. Click Refresh Dashboard\n3. Verify Total Devices, Active Deployments, Pending Approvals, Health Score KPI values all visible after refresh\n4. Verify no '—' loading placeholder remains\n5. Verify no error banner",
     "All KPI cards repopulate with valid data after refresh; no placeholders or error banners remain", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-REFRESH-03", "Refresh button shows animate-spin class during refresh",
     "1. Re-navigate to Dashboard; confirm Refresh button visible\n2. Click Refresh Dashboard\n3. Immediately JS evaluate button svg class — assert contains 'animate-spin'\n4. Wait for KPI card to repopulate (refresh complete gate)\n5. JS evaluate button svg class — assert does NOT contain 'animate-spin'",
     "SVG icon has 'animate-spin' class while refresh is active; class removed after completion", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-REFRESH-04", "Refresh button disabled during active refresh",
     "1. Re-navigate to Dashboard\n2. JS evaluate button.disabled before click — assert false\n3. Click Refresh Dashboard\n4. Immediately JS evaluate button.disabled — assert true\n5. Wait for KPI repopulation (refresh complete)\n6. JS evaluate button.disabled — assert false",
     "Button has disabled attribute during refresh; re-enabled after completion", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-REFRESH-05", "Badge counts valid after Refresh Dashboard",
     "1. Re-navigate to Dashboard; verify View Inventory card visible\n2. Click Refresh Dashboard; wait for KPI repopulation\n3. Verify all 4 Quick Action cards still visible\n4. JS evaluate all main span.bg-orange-500 — assert none contains U+2014 em-dash",
     "All Quick Action cards present after refresh; no badge shows a loading placeholder", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-ERR-01", "Badge values are never zero or negative",
     "1. Re-navigate to Dashboard; wait for badges to resolve\n2. Verify welcome heading visible (page loaded)\n3. JS evaluate all main span.bg-orange-500 — for each, parseInt and assert > 0",
     "All visible badge values parse as integers > 0; no zero, negative, or NaN values", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-ERR-02", "Quick Action cards accessible when GraphQL returns HTTP 500",
     "1. Register route **/graphql -> HTTP 500\n2. Navigate to Dashboard with all APIs failing\n3. Verify all 4 Quick Action cards still visible\n4. JS evaluate card text — assert no 'undefined', 'NaN', 'null' strings in badge positions\n5. Unroute (finally)",
     "All 4 cards rendered; no raw error strings in badge positions; cards not collapsed by API failure", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-ERR-03", "Dashboard survives hard page reload",
     "1. Re-navigate to Dashboard; verify welcome message and cards visible\n2. browser.reload() — hard reload\n3. Verify Dashboard h1 visible (session preserved, no login redirect)\n4. Verify welcome message containsText(adminUsername)\n5. Verify all 4 Quick Action cards visible\n6. Verify Refresh button visible and not disabled",
     "After reload: session preserved; welcome message, 4 cards, and Refresh button all re-render correctly", "REGRESSION", "Active"),
    ("UI — Dashboard Quick Actions", "dashboard-quick-actions-suite.xml", "DashboardQuickActionsTests",
     "TC-PS6-ERR-04", "Direct URL access renders all Quick Action features",
     "1. Launch.app(baseUrl) with existing authenticated session\n2. Verify Dashboard h1 visible (no redirect to login)\n3. Verify welcome message visible\n4. Verify all 4 Quick Action cards visible\n5. Verify Refresh button visible and not disabled",
     "Direct URL navigation renders welcome message, 4 cards, and Refresh button without re-login", "REGRESSION", "Active"),

    # ── Section 16: UI — Dashboard E2E Integration (27) ────────────────────
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-KPI-01", "All four KPI cards populate from parallel API calls on page load",
     "1. Re-navigate to Dashboard to trigger fresh page load\n2. Verify 'Welcome back' h2 heading visible\n3. Verify Total Devices, Active Deployments, Pending Approvals, Health Score KPI card values all visible\n4. Verify no '—' (U+2014) loading placeholder remains in any KPI card\n5. Verify no red error banner (div.bg-red-50) is present",
     "All 4 KPI cards populated from 6 parallel API calls; no placeholder; no error banner", "SMOKE", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-KPI-02", "KPI values display valid non-negative data — no JS artifacts",
     "1. Verify Total Devices, Active Deployments, Pending Approvals values visible\n2. Assert Health Score value containsText('%') — percentage format confirmed\n3. JS evaluate all div[class*='text-3xl'] textContent joined\n4. Assert result is non-empty and contains no 'NaN', 'undefined', or 'null'",
     "All KPI values are valid; Health Score has '%' suffix; no JavaScript error artifacts in any card", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-KPI-03", "KPI cards show em-dash loading placeholder during API fetch delay",
     "1. Register route to delay all AppSync GraphQL responses by 2 s via Thread.sleep\n2. Re-navigate to Dashboard to trigger fresh page load\n3. Assert KpiCard.LOADING_PLACEHOLDER ('—' U+2014) is visible immediately after navigation\n4. Unregister all route handlers (finally block)",
     "At least one KPI card displays '—' while six API calls are artificially delayed", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ALERTS-01", "Recent Alerts panel displays exactly 3 entries from 24-hour window",
     "1. Register route to inject 3-entry mock (approved/failed/info) for listAuditLogs\n2. Re-navigate to Dashboard with mock active\n3. Verify AlertsPanel.CONTAINER visible\n4. JS evaluate count of div.bg-muted rows — assert equals 3\n5. Verify AlertsPanel.EMPTY_STATE is NOT visible\n6. Unregister routes (finally)",
     "Exactly 3 div.bg-muted alert item rows rendered; 'No recent activity' absent", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ALERTS-02", "Recent Alerts panel shows empty state when no audit logs exist",
     "1. Register route to inject empty items[] for listAuditLogs\n2. Re-navigate to Dashboard\n3. Verify AlertsPanel.CONTAINER still visible\n4. Verify AlertsPanel.EMPTY_STATE ('No recent activity') is visible\n5. JS evaluate count of div.bg-muted rows — assert equals 0\n6. Unregister routes (finally)",
     "'No recent activity' text visible; zero div.bg-muted rows; panel container still present", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ALERTS-03", "'View all' link in Recent Alerts panel navigates to /analytics",
     "1. Verify AlertsPanel.VIEW_ALL_LINK (a[href='/analytics']) is visible\n2. Click 'View all' link via DashboardPageImpl.clickViewAllAlerts()\n3. JS evaluate window.location.pathname\n4. Assert pathname equals '/analytics'",
     "Clicking 'View all' navigates to the /analytics route", "SMOKE", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ICON-01", "Failed audit entry renders exclusive orange warning SVG icon",
     "1. Inject single auditStatus='failed' entry via listAuditLogs mock\n2. Re-navigate to Dashboard\n3. Verify AlertsPanel.FAILED_ACTION_ICON (svg[class*='text-orange']) visible\n4. Verify AlertsPanel.APPROVED_ACTION_ICON NOT visible\n5. Verify AlertsPanel.INFO_ACTION_ICON NOT visible\n6. Unregister routes (finally)",
     "Orange SVG icon exclusive to failed entry; no green or blue icon in same row", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ICON-02", "Approved audit entry renders exclusive green check SVG icon",
     "1. Inject single auditStatus='approved' entry via listAuditLogs mock\n2. Re-navigate to Dashboard\n3. Verify AlertsPanel.APPROVED_ACTION_ICON (svg[class*='text-green']) visible\n4. Verify AlertsPanel.FAILED_ACTION_ICON NOT visible\n5. Unregister routes (finally)",
     "Green SVG icon exclusive to approved entry; no orange icon in same row", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ICON-03", "Info audit entry renders exclusive blue info SVG icon",
     "1. Inject single auditStatus='info' entry via listAuditLogs mock\n2. Re-navigate to Dashboard\n3. Verify AlertsPanel.INFO_ACTION_ICON (svg[class*='text-blue']) visible\n4. Verify no orange or green icon in same row\n5. Unregister routes (finally)",
     "Blue SVG icon exclusive to info entry; no orange or green icon in same row", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ICON-04", "Each alert item row displays action text and a timestamp",
     "1. Inject one approved entry with action 'Firmware update approved'\n2. Re-navigate to Dashboard\n3. Verify AlertsPanel.ALERT_ITEM (div.bg-muted) visible\n4. Verify AlertsPanel.CONTAINER containsText('Firmware update approved')\n5. JS evaluate first div.bg-muted textContent — assert non-empty (includes time ref)\n6. Unregister routes (finally)",
     "Action text and non-empty timestamp both present in alert item row", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-BADGE-01", "All four Quick Action cards rendered with labels, icons, and badges",
     "1. Re-navigate to Dashboard\n2. Verify VIEW_INVENTORY_CARD_MAIN, SCHEDULE_SERVICE_CARD_MAIN, DEPLOY_FIRMWARE_CARD, CHECK_COMPLIANCE_CARD_MAIN all visible\n3. JS evaluate count of main div[class*='grid'] > a — assert 4\n4. Verify SVG icon visible in each card",
     "Exactly 4 Quick Action cards; each has label, icon, and badge span", "SMOKE", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-BADGE-02", "View Inventory card shows correct offline device badge count (3)",
     "1. Inject 3 offline devices for listDevices(Offline) mock\n2. Re-navigate to Dashboard\n3. Verify QuickActions.VIEW_INVENTORY_BADGE (span.bg-orange-500) visible\n4. JS evaluate badge textContent — assert '3'\n5. Unregister routes (finally)",
     "View Inventory badge displays '3' matching the 3 injected offline devices", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-BADGE-03", "Schedule Service card shows correct scheduled orders badge count (5)",
     "1. Inject 5 scheduled orders for listServiceOrdersByStatus(Scheduled) mock\n2. Re-navigate to Dashboard\n3. Verify QuickActions.SCHEDULE_SERVICE_BADGE visible\n4. JS evaluate badge textContent — assert '5'\n5. Unregister routes (finally)",
     "Schedule Service badge displays '5' matching the 5 injected scheduled orders", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-BADGE-04", "Deploy Firmware card shows correct pending firmware badge count (2)",
     "1. Inject 2 pending firmware records for listFirmware(Pending) mock\n2. Re-navigate to Dashboard\n3. Verify QuickActions.DEPLOY_FIRMWARE_BADGE visible\n4. JS evaluate badge textContent — assert '2'\n5. Unregister routes (finally)",
     "Deploy Firmware badge displays '2' matching the 2 injected pending firmware records", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-BADGE-05", "Check Compliance card shows correct pending compliance badge count (4)",
     "1. Inject 4 pending compliance records for listFirmwareCompliance(Pending) mock\n2. Re-navigate to Dashboard\n3. Verify QuickActions.CHECK_COMPLIANCE_BADGE visible\n4. JS evaluate badge textContent — assert '4'\n5. Unregister routes (finally)",
     "Check Compliance badge displays '4' matching the 4 injected pending compliance records", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-BADGE-06", "Quick Action cards navigate to correct module routes",
     "1. Re-navigate to Dashboard\n2. Click View Inventory -> JS evaluate pathname, assert '/inventory'\n3. Return to Dashboard; click Schedule Service -> assert '/account-service'\n4. Return; click Deploy Firmware -> assert '/deployment'\n5. Return; click Check Compliance -> assert '/compliance'",
     "All 4 Quick Action cards navigate to their correct routes", "SMOKE", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-STATUS-01", "Deployment Service shows Operational (green) when active orders exist",
     "1. Inject 1 In-Progress order for listServiceOrdersByStatus(In Progress)\n2. Re-navigate to Dashboard\n3. Verify SystemStatus.CONTAINER visible\n4. Verify SystemStatus.DEPLOYMENT_SERVICE_OPERATIONAL (span.text-green-600) visible\n5. Verify SystemStatus.DEPLOYMENT_SERVICE_DEGRADED NOT visible\n6. Unregister routes (finally)",
     "Deployment Service row shows green 'Operational' with active orders", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-STATUS-02", "Deployment Service shows Degraded (orange) when no active orders exist",
     "1. Inject empty orders for listServiceOrdersByStatus(In Progress)\n2. Re-navigate to Dashboard\n3. Verify SystemStatus.DEPLOYMENT_SERVICE_DEGRADED (span.text-orange-600) visible\n4. Verify SystemStatus.DEPLOYMENT_SERVICE_OPERATIONAL NOT visible\n5. Unregister routes (finally)",
     "Deployment Service row shows orange 'Degraded' with zero active orders", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-STATUS-03", "System Status panel renders exactly 4 service rows in fixed order",
     "1. JS evaluate count of div.space-y-4 > div in System Status panel — assert 4\n2. Verify panel containsText 'Deployment Service', 'Compliance Engine', 'Asset Database', 'Analytics Platform'\n3. Verify SystemStatus.ANALYTICS_PLATFORM_OPERATIONAL visible\n4. Verify SystemStatus.ANALYTICS_PLATFORM_DEGRADED NOT visible",
     "Exactly 4 service rows; all 4 names present; Analytics Platform always Operational", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-STATUS-04", "Compliance Engine shows Degraded when pending compliance items exist",
     "1. Inject 1 Pending compliance item for listFirmwareCompliance mock\n2. Re-navigate to Dashboard\n3. Verify SystemStatus.COMPLIANCE_ENGINE_DEGRADED (span.text-orange-600 in Compliance Engine row) visible\n4. Unregister routes (finally)",
     "Compliance Engine row shows orange 'Degraded' when pending compliance items are present", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-REFRESH-01", "Refresh Dashboard button is visible and accessible",
     "1. Verify DashboardPage.REFRESH_DASHBOARD_BUTTON (button[aria-label='Refresh dashboard']) visible\n2. JS evaluate button.disabled — assert false (button is enabled)",
     "Refresh button present in welcome header row and is enabled", "SMOKE", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-REFRESH-02", "Clicking Refresh re-triggers all API calls; KPI cards repopulate",
     "1. Verify no loading placeholder before click (stable state)\n2. Click Refresh Dashboard button\n3. Wait for KpiCard.LOADING_PLACEHOLDER to appear (refresh in progress)\n4. Wait for placeholder to disappear (refresh complete)\n5. Verify all 4 KPI card values visible\n6. Verify no error banner",
     "All KPI cards repopulated after manual refresh; no error banner", "SMOKE", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-REFRESH-03", "Refresh SVG icon applies animate-spin class during active refresh cycle",
     "1. Register route to delay all AppSync responses by 2 s\n2. Click Refresh Dashboard\n3. Immediately JS evaluate button[aria-label='Refresh dashboard'] svg class — assert contains 'animate-spin'\n4. Wait for placeholder to disappear (refresh complete)\n5. JS evaluate SVG class again — assert does NOT contain 'animate-spin'\n6. Unregister routes (finally)",
     "SVG has 'animate-spin' during refresh; class removed on completion", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-REFRESH-04", "Dashboard data updates after refresh when backend data changed",
     "1. Register route returning 1 device for listDevices (non-offline) — simulates changed data\n2. Click Refresh Dashboard\n3. Wait for placeholder to appear then disappear (refresh complete)\n4. JS evaluate Total Devices div[class*='text-3xl'] textContent — assert '1'\n5. Unregister routes (finally)",
     "Total Devices KPI displays '1' after refresh reflecting the updated mock backend data", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ERR-01", "Red error banner displayed when all KPI API calls fail with HTTP 500",
     "1. Register route returning HTTP 500 for all AppSync requests\n2. Re-navigate to Dashboard\n3. Verify ErrorBanner.CONTAINER (div.p-6.space-y-6 > div.bg-red-50) visible\n4. Verify ErrorBanner.MESSAGE containsText('network error')\n5. Unregister routes (finally)",
     "Red error banner with network error message visible when all KPI APIs return 500", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ERR-02", "KPI cards fall back to '0' on full API failure — no crash or blank screen",
     "1. Register route returning HTTP 500 for all AppSync requests\n2. Re-navigate to Dashboard\n3. Wait for KpiCard.ZERO_FALLBACK ('0') to become visible\n4. Verify no '—' loading placeholder remains\n5. Verify Total Devices and Health Score card containers present\n6. JS evaluate !!document.querySelector('h1') — assert true (no blank screen)\n7. Unregister routes (finally)",
     "KPI cards show '0' gracefully; no crash overlay; Dashboard layout intact", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ERR-03", "Audit-log API failure does not trigger global KPI error banner",
     "1. Register route returning HTTP 500 ONLY for listAuditLogs; all other requests pass through\n2. Re-navigate to Dashboard\n3. Verify ErrorBanner.CONTAINER is NOT visible (global banner absent)\n4. Verify KpiCard.TOTAL_DEVICES_VALUE visible (KPI domain unaffected)\n5. Assert alerts panel shows in-panel error OR empty state\n6. Unregister routes (finally)",
     "No global error banner; KPI cards populated; alerts panel handles its own failure in isolation", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ERR-04", "Error banner disappears after successful manual refresh following failure",
     "1. Register route returning HTTP 500 for all AppSync requests\n2. Re-navigate to Dashboard — confirm ErrorBanner.CONTAINER visible\n3. Remove all error routes via browser.unrouteAll()\n4. Click Refresh Dashboard\n5. Wait for KPI placeholder to disappear (refresh complete)\n6. Verify ErrorBanner.CONTAINER is NOT visible",
     "Error banner dismissed after a successful manual refresh; error state is recoverable", "REGRESSION", "Active"),
    ("UI — Dashboard E2E Integration (PS-7)", "dashboard-e2e-integration-suite.xml", "DashboardE2EIntegrationTests",
     "TC-PS7-ERR-05", "Unauthenticated access to Dashboard redirects to login page",
     "1. Clear all cookies via browser.context().clearCookies()\n2. Clear localStorage and sessionStorage via JS evaluate\n3. Navigate directly to base URL\n4. Verify LoginPage.EMAIL_FIELD visible (redirected to login)\n5. Verify DashboardPage.DASHBOARD_HEADER is NOT visible",
     "Unauthenticated navigation redirects to login; Dashboard is not exposed", "SMOKE", "Active"),
]

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=12)
SECTION_FONT = Font(bold=True, size=11)

COLOR_SMOKE      = "C6EFCE"  # light green
COLOR_REGRESSION = "BDD7EE"  # light blue
COLOR_NEGATIVE   = "FCE4D6"  # light orange
COLOR_HEADER     = "1F4E79"  # dark blue
COLOR_SECTION_BG = "D6E4F0"  # section header bg

FILL_SMOKE      = PatternFill("solid", fgColor=COLOR_SMOKE)
FILL_REGRESSION = PatternFill("solid", fgColor=COLOR_REGRESSION)
FILL_NEGATIVE   = PatternFill("solid", fgColor=COLOR_NEGATIVE)
FILL_HEADER     = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_SECTION    = PatternFill("solid", fgColor=COLOR_SECTION_BG)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_header_row(ws, row, values, fill=FILL_HEADER, font=HEADER_FONT):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, type_val):
    if type_val == "SMOKE":
        cell.fill = FILL_SMOKE
    elif type_val == "REGRESSION":
        cell.fill = FILL_REGRESSION
    elif type_val == "NEGATIVE":
        cell.fill = FILL_NEGATIVE
    cell.border = THIN_BORDER
    cell.alignment = WRAP


# ─────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Summary ──────────────────────────────────────────────────────
ws_summary = wb.active
ws_summary.title = "Summary"

ws_summary.merge_cells("A1:F1")
title_cell = ws_summary["A1"]
title_cell.value = "HLM-QA — Test Case Summary  |  Generated: 2026-04-02  |  Total: 246 Tests"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = CENTER
title_cell.fill = PatternFill("solid", fgColor="1F4E79")
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
ws_summary.row_dimensions[1].height = 30

ws_summary.append([])  # blank row

set_header_row(ws_summary, 3, ["Module / Section", "Suite File", "Test Class", "SMOKE", "REGRESSION", "NEGATIVE", "Total"])

# Count per section
from collections import defaultdict
section_counts = defaultdict(lambda: {"SMOKE": 0, "REGRESSION": 0, "NEGATIVE": 0})
for tc in TEST_CASES:
    section, _, _, _, _, _, _, type_, _ = tc
    if type_ in section_counts[section]:
        section_counts[section][type_] += 1
    else:
        section_counts[section][type_] = 1

# Get unique sections in order
seen = []
for tc in TEST_CASES:
    if tc[0] not in seen:
        seen.append(tc[0])

suite_map = {}
class_map = {}
for tc in TEST_CASES:
    if tc[0] not in suite_map:
        suite_map[tc[0]] = tc[1]
        class_map[tc[0]] = tc[2]

totals = {"SMOKE": 0, "REGRESSION": 0, "NEGATIVE": 0}
row_num = 4
for section in seen:
    s = section_counts[section]["SMOKE"]
    r = section_counts[section]["REGRESSION"]
    n = section_counts[section]["NEGATIVE"]
    t = s + r + n
    totals["SMOKE"] += s
    totals["REGRESSION"] += r
    totals["NEGATIVE"] += n
    ws_summary.append([section, suite_map[section], class_map[section], s, r, n, t])
    for col in range(1, 8):
        cell = ws_summary.cell(row=row_num, column=col)
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        if col >= 4:
            cell.alignment = CENTER
    row_num += 1

# Total row
total_row = row_num
ws_summary.append(["TOTAL", "", "", totals["SMOKE"], totals["REGRESSION"], totals["NEGATIVE"],
                   totals["SMOKE"] + totals["REGRESSION"] + totals["NEGATIVE"]])
for col in range(1, 8):
    cell = ws_summary.cell(row=total_row, column=col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FFD700")
    cell.border = THIN_BORDER
    cell.alignment = CENTER

# Column widths for summary
ws_summary.column_dimensions["A"].width = 55
ws_summary.column_dimensions["B"].width = 42
ws_summary.column_dimensions["C"].width = 40
for col in ["D", "E", "F", "G"]:
    ws_summary.column_dimensions[col].width = 12

# Legend
ws_summary.append([])
ws_summary.append(["Type Legend"])
ws_summary.cell(row=total_row + 2, column=1).font = Font(bold=True, size=11)
legend = [
    ("SMOKE", "Core happy-path tests; must pass on every build", COLOR_SMOKE),
    ("REGRESSION", "Full regression coverage; run on release candidates", COLOR_REGRESSION),
    ("NEGATIVE", "Error / boundary / auth failure scenarios", COLOR_NEGATIVE),
]
for label, desc, color in legend:
    ws_summary.append([label, desc])
    r = ws_summary.max_row
    ws_summary.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
    ws_summary.cell(row=r, column=1).font = Font(bold=True)
    ws_summary.cell(row=r, column=1).border = THIN_BORDER
    ws_summary.cell(row=r, column=2).border = THIN_BORDER

ws_summary.freeze_panes = "A4"


# ── Sheet 2: All Test Cases ───────────────────────────────────────────────
ws_all = wb.create_sheet("All Test Cases")

ws_all.merge_cells("A1:I1")
ws_all["A1"].value = "HLM-QA — All Test Cases (246 Total)  |  Generated: 2026-04-02"
ws_all["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws_all["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws_all["A1"].alignment = CENTER
ws_all.row_dimensions[1].height = 28

headers = ["Section", "Suite File", "Test Class", "TC ID", "Scenario", "Test Steps", "Expected Result", "Type", "Status"]
set_header_row(ws_all, 2, headers)

current_section = None
data_row = 3
for tc in TEST_CASES:
    section, suite, cls, tc_id, scenario, steps, expected, type_, status = tc

    # Section separator row
    if section != current_section:
        current_section = section
        ws_all.append([""] * 9)
        ws_all.merge_cells(f"A{ws_all.max_row}:I{ws_all.max_row}")
        sec_cell = ws_all.cell(row=ws_all.max_row, column=1)
        sec_cell.value = f"  {section}"
        sec_cell.font = Font(bold=True, size=11)
        sec_cell.fill = FILL_SECTION
        sec_cell.border = THIN_BORDER
        sec_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_all.row_dimensions[ws_all.max_row].height = 20
        data_row = ws_all.max_row + 1

    ws_all.append([section, suite, cls, tc_id, scenario, steps, expected, type_, status])
    r = ws_all.max_row
    for col in range(1, 10):
        style_data_cell(ws_all.cell(row=r, column=col), type_)
    ws_all.row_dimensions[r].height = 80

ws_all.freeze_panes = "A3"

# Column widths
col_widths = [38, 38, 38, 18, 45, 70, 55, 12, 12]
for i, w in enumerate(col_widths, 1):
    ws_all.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 3: UI Tests ─────────────────────────────────────────────────────
ws_ui = wb.create_sheet("UI Tests")
ws_ui.merge_cells("A1:I1")
ws_ui["A1"].value = "HLM-QA — UI Test Cases"
ws_ui["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws_ui["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws_ui["A1"].alignment = CENTER
ws_ui.row_dimensions[1].height = 28

set_header_row(ws_ui, 2, headers)

ui_sections = [s for s in seen if s.startswith("UI")]
current_section = None
for tc in TEST_CASES:
    section, suite, cls, tc_id, scenario, steps, expected, type_, status = tc
    if not section.startswith("UI"):
        continue
    if section != current_section:
        current_section = section
        ws_ui.append([""] * 9)
        ws_ui.merge_cells(f"A{ws_ui.max_row}:I{ws_ui.max_row}")
        sec_cell = ws_ui.cell(row=ws_ui.max_row, column=1)
        sec_cell.value = f"  {section}"
        sec_cell.font = Font(bold=True, size=11)
        sec_cell.fill = FILL_SECTION
        sec_cell.border = THIN_BORDER
        sec_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_ui.row_dimensions[ws_ui.max_row].height = 20

    ws_ui.append([section, suite, cls, tc_id, scenario, steps, expected, type_, status])
    r = ws_ui.max_row
    for col in range(1, 10):
        style_data_cell(ws_ui.cell(row=r, column=col), type_)
    ws_ui.row_dimensions[r].height = 80

ws_ui.freeze_panes = "A3"
for i, w in enumerate(col_widths, 1):
    ws_ui.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 4: API Tests ────────────────────────────────────────────────────
ws_api = wb.create_sheet("API Tests")
ws_api.merge_cells("A1:I1")
ws_api["A1"].value = "HLM-QA — API Test Cases"
ws_api["A1"].font = Font(bold=True, size=13, color="FFFFFF")
ws_api["A1"].fill = PatternFill("solid", fgColor="1F4E79")
ws_api["A1"].alignment = CENTER
ws_api.row_dimensions[1].height = 28

set_header_row(ws_api, 2, headers)

current_section = None
for tc in TEST_CASES:
    section, suite, cls, tc_id, scenario, steps, expected, type_, status = tc
    if not section.startswith("API"):
        continue
    if section != current_section:
        current_section = section
        ws_api.append([""] * 9)
        ws_api.merge_cells(f"A{ws_api.max_row}:I{ws_api.max_row}")
        sec_cell = ws_api.cell(row=ws_api.max_row, column=1)
        sec_cell.value = f"  {section}"
        sec_cell.font = Font(bold=True, size=11)
        sec_cell.fill = FILL_SECTION
        sec_cell.border = THIN_BORDER
        sec_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_api.row_dimensions[ws_api.max_row].height = 20

    ws_api.append([section, suite, cls, tc_id, scenario, steps, expected, type_, status])
    r = ws_api.max_row
    for col in range(1, 10):
        style_data_cell(ws_api.cell(row=r, column=col), type_)
    ws_api.row_dimensions[r].height = 80

ws_api.freeze_panes = "A3"
for i, w in enumerate(col_widths, 1):
    ws_api.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────
output_path = "/Users/ajaykumar.yadav/HLM-QA/TEST-CASES.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Total test cases written: {len(TEST_CASES)}")
print(f"Sheets: {wb.sheetnames}")
