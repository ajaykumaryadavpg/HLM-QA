"""
generate_api_catalogue_xls.py
Generates HLM-API-Catalogue.xlsx for Dev & QA manual API testing.
Run: python3 generate_api_catalogue_xls.py
Requires: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constants ─────────────────────────────────────────────────────────────────
GRAPHQL_URL  = "https://rw3sl7mwgzcvbg7zuyqj7ppjv4.appsync-api.us-east-2.amazonaws.com/graphql"
COGNITO_IDP  = "https://cognito-idp.us-east-2.amazonaws.com/"
COGNITO_ID   = "https://cognito-identity.us-east-2.amazonaws.com/"
AMPLIFY_URL  = "https://main.dddsig2mih3hw.amplifyapp.com"
MAP_BASE     = "https://maps.geo.us-east-2.amazonaws.com"
MAP_NAME     = "HlmMapOpenData-a3b4d250-131a-11f1-90d8-0aa113e32f65"
AUTH_VALUE   = "Authorization: <Cognito idToken>   (raw JWT — NO 'Bearer' prefix)"
AUTH_FULL    = (
    "Header: Authorization: <Cognito idToken>  (raw JWT, NO 'Bearer' prefix)\n"
    "Header: Content-Type: application/json\n"
    "Token type: idToken  (NOT accessToken)\n"
    "Unauthenticated → 401 UnauthorizedException"
)

# ── Colour palette ────────────────────────────────────────────────────────────
CLR = {
    "header_bg":  "1F3864", "header_fg":  "FFFFFF",
    "section_bg": "2E75B6", "section_fg": "FFFFFF",
    "query_hdr":  "C5D9F1",
    "mut_hdr":    "FFE699",
    "rest_hdr":   "C6EFCE",
    "alt_row":    "F7F7F7",
    "white":      "FFFFFF",
    "border":     "BFBFBF",
    "active":     "C6EFCE",
    "pending":    "FFEB9C",
    "known_bug":  "FCE4D6",
    "method_post":"DDEEFF",
    "method_get": "E2EFDA",
    "url_bg":     "EBF3FB",
    "auth_bg":    "FFF2CC",
}

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(h):
    return PatternFill("solid", fgColor=h)

def font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def border():
    s = Side(style="thin", color=CLR["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def align(wrap=True, h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_cell(cell, bg=None, bold=False, fg="000000", sz=10, h="left", wrap=True):
    if bg:
        cell.fill = fill(bg)
    cell.font      = font(bold=bold, color=fg, size=sz)
    cell.alignment = align(wrap=wrap, h=h)
    cell.border    = border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title(ws, n_cols, text, bg=None):
    bg = bg or CLR["header_bg"]
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = text
    style_cell(c, bg=bg, bold=True, fg=CLR["header_fg"], sz=12, h="center")
    ws.row_dimensions[1].height = 28

def write_col_headers(ws, row, headers, bg):
    for col_i, name in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_i, value=name)
        style_cell(c, bg=bg, bold=True)
    ws.row_dimensions[row].height = 20


# ══════════════════════════════════════════════════════════════════════════════
# Workbook
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Infrastructure & Auth
# ══════════════════════════════════════════════════════════════════════════════
ws_infra = wb.create_sheet("Infrastructure & Auth")
set_col_widths(ws_infra, [22, 16, 18, 70, 38])

write_title(ws_infra, 5, "HLM Platform — Infrastructure & Authentication Reference")

write_col_headers(ws_infra, 2, ["Layer", "Method", "Service", "URL / Value", "Auth / Notes"], CLR["section_bg"])

infra_rows = [
    ("GraphQL API",         "POST",  "AWS AppSync",              GRAPHQL_URL,
     AUTH_FULL),
    ("Auth — User Pool",    "POST",  "AWS Cognito IDP",          COGNITO_IDP,
     "Pool ID: us-east-2_Q36YWNsEC\nClient ID: 3aseu1rf3q3tae7u4dgplllii8"),
    ("Auth — Identity",     "POST",  "AWS Cognito Identity",     COGNITO_ID,
     "Returns temporary AWS credentials for map tile signing"),
    ("Map Tiles",           "GET",   "AWS Location (MapLibre)",  f"{MAP_BASE}/maps/v0/maps/{MAP_NAME}/",
     "Signed with AWS Sig v4 via Amplify session credentials"),
    ("App Hosting",         "GET",   "AWS Amplify",              AMPLIFY_URL,
     "Frontend SPA"),
]

for r_i, (layer, method, svc, url, notes) in enumerate(infra_rows, 3):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]
    method_bg = CLR["method_post"] if method == "POST" else CLR["method_get"]
    vals = [layer, method, svc, url, notes]
    for col_i, val in enumerate(vals, 1):
        c = ws_infra.cell(row=r_i, column=col_i, value=val)
        if col_i == 2:
            style_cell(c, bg=method_bg, bold=True, h="center")
        elif col_i == 4:
            style_cell(c, bg=CLR["url_bg"])
        elif col_i == 5:
            style_cell(c, bg=CLR["auth_bg"])
        else:
            style_cell(c, bg=bg)
    ws_infra.row_dimensions[r_i].height = 60

# operationName constraint note
note_row = len(infra_rows) + 4
ws_infra.merge_cells(f"A{note_row}:E{note_row}")
c = ws_infra.cell(row=note_row, column=1,
    value="⚠  AppSync operationName Constraint:  The 'operationName' field in the POST body MUST exactly match "
          "the named operation in the query string.  Suffixes or aliased names → 400 BadRequestException.")
style_cell(c, bg=CLR["known_bug"], sz=9)
ws_infra.row_dimensions[note_row].height = 30

ws_infra.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# Helper: build full GraphQL request body string for display
# ══════════════════════════════════════════════════════════════════════════════
def gql_body(operation_name, gql_str, sample_vars_str):
    """Return a ready-to-paste JSON POST body string."""
    # Collapse the GQL to one line for the JSON field value representation
    one_line = " ".join(gql_str.split())
    return (
        f'{{\n'
        f'  "operationName": "{operation_name}",\n'
        f'  "query": "{one_line}",\n'
        f'  "variables": {sample_vars_str}\n'
        f'}}'
    )


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — GraphQL Queries
# ══════════════════════════════════════════════════════════════════════════════
#  Columns: #  | Operation | Category | Method | API URL | Auth | Purpose/Used On
#           | Arguments | Returns | Request Body (Full JSON) | Notes | Status
Q_COLS   = ["#", "Operation", "Category", "Method", "API URL",
            "Auth", "Purpose / Used On", "Arguments", "Returns",
            "Request Body (Full JSON — paste into Postman / Insomnia)",
            "Notes / Known Quirks", "Test Status"]
Q_WIDTHS = [5, 26, 16, 10, 60, 36, 42, 36, 22, 72, 42, 13]

# (id, operation, category, purpose, arguments, returns, gql_doc, first_sample_vars, extra_vars_note, notes)
QUERIES_RAW = [
    # ── Devices ─────────────────────────────────────────────────────────────
    ("Q-01", "listDevices", "Device",
     "List all devices, optionally filtered by status.\nUsed on: Dashboard (KPI cards), Hardware Inventory tab, Reporting & Analytics",
     "status?: String\nlimit?: Int\nnextToken?: String",
     "PaginatedResponse\n(items · nextToken · totalCount)",
     'query listDevices($status: String, $limit: Int, $nextToken: String) { listDevices(status: $status, limit: $limit, nextToken: $nextToken) { items nextToken totalCount } }',
     '{ "status": "Online" }',
     'Try also:\n{ "status": "Offline" }\n{ "status": "Maintenance" }\n{} ← no filter → all',
     "Status values: Online | Offline | Maintenance\nOmit status to return all records."),

    ("Q-02", "getDevice", "Device",
     "Fetch a single device record by its unique ID.\nUsed on: Hardware Inventory (device detail view)",
     "id: String!",
     "DeviceType",
     'query getDevice($id: String!) { getDevice(id: $id) { id deviceName serialNumber model status location firmwareVersion } }',
     '{ "id": "a11bc9b9-ba1e-4e1f-8523-45cd003a6845" }',
     "",
     "Returns null for non-existent ID"),

    ("Q-03", "getDevicesByCustomer", "Device",
     "Retrieve all devices belonging to a specific customer account.\nUsed on: Account & Service (customer device listing)",
     "customerId: String!",
     "PaginatedResponse",
     'query getDevicesByCustomer($customerId: String!) { getDevicesByCustomer(customerId: $customerId) { items totalCount } }',
     '{ "customerId": "c001" }',
     'Try: c001 – c005',
     "Demo customer IDs: c001 – c005"),

    ("Q-04", "getDevicesByLocation", "Device",
     "Retrieve all devices at a named location (building / site).\nUsed on: Inventory → Geo Location tab (map-pin grouping)",
     "location: String!",
     "PaginatedResponse",
     'query getDevicesByLocation($location: String!) { getDevicesByLocation(location: $location) { items totalCount } }',
     '{ "location": "New York Office" }',
     "",
     "location must exactly match the string stored in the device record"),

    # ── Firmware ─────────────────────────────────────────────────────────────
    ("Q-05", "listFirmware", "Firmware",
     "List all firmware entries, optionally filtered by status.\nUsed on: Firmware Status tab, Deployment → Firmware Versions, Dashboard (pending approvals KPI)",
     "status?: String\nlimit?: Int\nnextToken?: String",
     "PaginatedResponse\n(items · nextToken · totalCount)",
     'query listFirmware($status: String, $limit: Int, $nextToken: String) { listFirmware(status: $status, limit: $limit, nextToken: $nextToken) { items nextToken totalCount } }',
     '{ "status": "Pending" }',
     'Try also:\n{ "status": "Approved" }\n{ "status": "Rejected" }\n{ "status": "Deprecated" } ← ⚠ filter ignored',
     "⚠ KNOWN BUG: Deprecated filter not applied — returns ALL records"),

    ("Q-06", "getFirmware", "Firmware",
     "Fetch a single firmware record by ID.\nUsed on: Deployment → Firmware Versions detail panel, Firmware Compliance detail view",
     "id: String!",
     "FirmwareType",
     'query getFirmware($id: String!) { getFirmware(id: $id) { id name version deviceModel status releaseDate fileSize checksum } }',
     '{ "id": "94688ba5-81b0-4c73-9f2a-5923f0b4b256" }',
     "",
     "fileSize is Int (bytes) — not Float"),

    ("Q-07", "getFirmwareByModel", "Firmware",
     "Retrieve all firmware versions applicable to a given device model.\nUsed on: Inventory → Firmware Status tab (model-scoped filtering)",
     "deviceModel: String!",
     "PaginatedResponse",
     'query getFirmwareByModel($deviceModel: String!) { getFirmwareByModel(deviceModel: $deviceModel) { items totalCount } }',
     '{ "deviceModel": "XR-5000" }',
     "",
     "deviceModel must exactly match the model string stored in firmware records"),

    ("Q-08", "getFirmwareWithRelations", "Firmware",
     "Fetch a firmware record together with related compliance records and audit log entries.\nUsed on: Deployment → Firmware Versions (expanded details panel)",
     "id: String!",
     "PaginatedResponse\n(firmware + nested compliance + audit)",
     'query getFirmwareWithRelations($id: String!) { getFirmwareWithRelations(id: $id) { items totalCount } }',
     '{ "id": "94688ba5-81b0-4c73-9f2a-5923f0b4b256" }',
     "",
     "Returns nested related records inside the items AWSJSON scalar — parse client-side"),

    # ── Service Orders ────────────────────────────────────────────────────────
    ("Q-09", "listServiceOrdersByStatus", "Service Order",
     "List service orders filtered by workflow status.\nUsed on: Account & Service page, Dashboard (KPI — pending actions)",
     "status: String!",
     "PaginatedResponse",
     'query listServiceOrdersByStatus($status: String!) { listServiceOrdersByStatus(status: $status) { items totalCount } }',
     '{ "status": "Pending" }',
     'Try also:\n{ "status": "In Progress" }\n{ "status": "Completed" }\n{ "status": "Cancelled" }',
     "Status values: Pending | In Progress | Completed | Cancelled"),

    ("Q-10", "listServiceOrdersByDate", "Service Order",
     "List service orders within a date range.\nUsed on: Account & Service (date-range filter), Reporting & Analytics (deployment trend chart)",
     "startDate: String! (ISO 8601)\nendDate: String! (ISO 8601)",
     "PaginatedResponse",
     'query listServiceOrdersByDate($startDate: String!, $endDate: String!) { listServiceOrdersByDate(startDate: $startDate, endDate: $endDate) { items totalCount } }',
     '{ "startDate": "2025-01-01", "endDate": "2025-12-31" }',
     'Edge cases:\n{ "startDate": "2030-01-01", "endDate": "2030-12-31" } ← future → []\n{ "startDate": "2025-12-31", "endDate": "2025-01-01" } ← reversed → []',
     "Returns empty array for reversed ranges (end < start) or future date ranges"),

    ("Q-11", "getServiceOrder", "Service Order",
     "Fetch a single service order by ID.\nUsed on: Account & Service (order detail view)",
     "id: String!",
     "ServiceOrderType",
     'query getServiceOrder($id: String!) { getServiceOrder(id: $id) { id title description status priority technicianId scheduledDate } }',
     '{ "id": "acf1df53-cf21-476d-9410-251efbf04b37" }',
     "",
     'Sample: "Core Switch Firmware Upgrade" — status: Completed'),

    ("Q-12", "getServiceOrdersByTechnician", "Service Order",
     "Retrieve all service orders assigned to a specific technician.\nUsed on: Account & Service (technician filter)",
     "technicianId: String!",
     "PaginatedResponse",
     'query getServiceOrdersByTechnician($technicianId: String!) { getServiceOrdersByTechnician(technicianId: $technicianId) { items totalCount } }',
     '{ "technicianId": "u003" }',
     'Try: u001 / u002 / u003',
     "technicianId maps to UserType.id (u001 = Alice, u002 = Bob, u003 = Carol)"),

    # ── Compliance ────────────────────────────────────────────────────────────
    ("Q-13", "listComplianceByStatus", "Compliance",
     "List compliance records filtered by certification status.\nUsed on: Firmware Compliance page, Dashboard (quick actions badge), Reporting & Analytics",
     "status: String!",
     "PaginatedResponse",
     'query listComplianceByStatus($status: String!) { listComplianceByStatus(status: $status) { items totalCount } }',
     '{ "status": "Approved" }',
     'Try also:\n{ "status": "Pending" }\n{ "status": "Deprecated" }',
     "Status values: Approved | Pending | Deprecated"),

    ("Q-14", "getCompliance", "Compliance",
     "Fetch a single compliance record by ID.\nUsed on: Firmware Compliance (record detail)",
     "id: String!",
     "ComplianceType",
     'query getCompliance($id: String!) { getCompliance(id: $id) { id firmwareId firmwareVersion deviceModel status certifications vulnerabilities } }',
     '{ "id": "85a9549f-..." }',
     "",
     "certifications is AWSJSON array — parse client-side"),

    ("Q-15", "getComplianceByCertification", "Compliance",
     "Filter compliance records by a specific certification standard.\nUsed on: Firmware Compliance (certification filter)",
     "certification: String!",
     "PaginatedResponse",
     'query getComplianceByCertification($certification: String!) { getComplianceByCertification(certification: $certification) { items totalCount } }',
     '{ "certification": "HIPAA" }',
     'Try also:\n{ "certification": "ISO 27001" }\n{ "certification": "FCC" }\n{ "certification": "WiFi Alliance" }\n{ "certification": "CE" }',
     "Known certifications: HIPAA | ISO 27001 | FCC | WiFi Alliance | CE"),

    # ── Audit Logs ────────────────────────────────────────────────────────────
    ("Q-16", "listAuditLogs", "Audit Log",
     "Retrieve audit log entries within a time window.\nDashboard uses 24-hour window. Deployment Audit Log: last 50 entries.",
     "startDate: String! (ISO 8601)\nendDate: String! (ISO 8601)\nlimit?: Int",
     "PaginatedResponse\n(items · nextToken · totalCount)",
     'query listAuditLogs($startDate: String!, $endDate: String!, $limit: Int) { listAuditLogs(startDate: $startDate, endDate: $endDate, limit: $limit) { items nextToken totalCount } }',
     '{ "startDate": "2026-03-26T00:00:00Z", "endDate": "2026-03-27T00:00:00Z" }',
     'Last 50 entries:\n{ "startDate": "2025-01-01", "endDate": "2026-12-31", "limit": 50 }\nFuture range → []',
     "Returns empty array for future or reversed date ranges"),

    ("Q-17", "getAuditLogsByUser", "Audit Log",
     "Retrieve the audit trail for a specific user by their user ID.\nUsed on: Reporting & Analytics (filtered audit log view)",
     "userId: String!",
     "PaginatedResponse",
     'query getAuditLogsByUser($userId: String!) { getAuditLogsByUser(userId: $userId) { items totalCount } }',
     '{ "userId": "u001" }',
     'Try: u001 / u002 / u003',
     "userId maps to UserType.id (u001 = Alice, u002 = Bob, u003 = Carol)"),

    # ── Users & Customers ─────────────────────────────────────────────────────
    ("Q-18", "getUserByEmail", "User",
     "Look up a user profile by email address.\nUsed on: App bootstrap / session init (all pages) — resolves logged-in user's display name and role",
     "email: String!",
     "UserType",
     'query getUserByEmail($email: String!) { getUserByEmail(email: $email) { id email role firstName lastName } }',
     '{ "email": "alice@acmecorp.com" }',
     'Try also:\n{ "email": "bob@acmecorp.com" }\n{ "email": "carol@acmecorp.com" }',
     "Returns null if email exists in Cognito but not in the application DB"),

    ("Q-19", "listUsersByRole", "User",
     "Retrieve users by role for assignment dropdowns.\nUsed on: Account & Service (technician assignment dropdown)",
     "role: String!",
     "PaginatedResponse",
     'query listUsersByRole($role: String!) { listUsersByRole(role: $role) { items totalCount } }',
     '{ "role": "Technician" }',
     'Try also:\n{ "role": "Admin" }\n{ "role": "Manager" }',
     "⚠ KNOWN BUG: role filter NOT applied server-side — returns ALL users regardless of value"),

    ("Q-20", "getCustomerWithRelations", "Customer",
     "Fetch a customer record along with all related devices and service orders.\nUsed on: Account & Service (customer detail panel)",
     "customerId: String!",
     "PaginatedResponse\n(customer + nested devices + service orders)",
     'query getCustomerWithRelations($customerId: String!) { getCustomerWithRelations(customerId: $customerId) { items totalCount } }',
     '{ "customerId": "c001" }',
     'Try: c001 – c005',
     "Nested related records returned inside items AWSJSON scalar — parse client-side"),
]

ws_q = wb.create_sheet("GraphQL Queries")
set_col_widths(ws_q, Q_WIDTHS)
write_title(ws_q, len(Q_COLS),
    f"HLM Platform — GraphQL Queries  (20 operations)  |  All requests: POST {GRAPHQL_URL}")
write_col_headers(ws_q, 2, Q_COLS, CLR["query_hdr"])

for r_i, (qid, op, cat, purpose, args, ret, gql_doc, sample_vars, extra_vars, notes) in \
        enumerate(QUERIES_RAW, 3):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]

    # Build full request body
    vars_for_body = sample_vars
    if extra_vars:
        vars_for_body = sample_vars  # use first sample as the primary body
    full_body = gql_body(op, gql_doc, vars_for_body)
    if extra_vars:
        full_body += f"\n\n// Additional variable examples:\n{extra_vars}"

    row_vals = [qid, op, cat, "POST", GRAPHQL_URL, AUTH_VALUE,
                purpose, args, ret, full_body, notes, "Active"]

    for col_i, val in enumerate(row_vals, 1):
        c = ws_q.cell(row=r_i, column=col_i, value=val)
        if col_i == 4:   # Method
            style_cell(c, bg=CLR["method_post"], bold=True, h="center")
        elif col_i == 5: # URL
            style_cell(c, bg=CLR["url_bg"], sz=9)
        elif col_i == 6: # Auth
            style_cell(c, bg=CLR["auth_bg"], sz=9)
        elif col_i == 10: # Request body
            style_cell(c, bg=CLR["white"], sz=9)
        elif col_i == 11 and "⚠" in str(val): # Known bug note
            style_cell(c, bg=CLR["known_bug"])
        elif col_i == 12: # Status
            style_cell(c, bg=CLR["active"], bold=True, h="center")
        else:
            style_cell(c, bg=bg)
    ws_q.row_dimensions[r_i].height = 100

ws_q.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — GraphQL Mutations
# ══════════════════════════════════════════════════════════════════════════════
M_COLS   = ["#", "Operation", "Category", "Method", "API URL",
            "Auth", "Purpose / Used On", "Arguments", "Returns",
            "Request Body (Full JSON — paste into Postman / Insomnia)",
            "Notes / Known Quirks", "Test Status"]
M_WIDTHS = [5, 26, 16, 10, 60, 36, 42, 42, 22, 72, 42, 13]

MUTATIONS_RAW = [
    ("M-01", "createDevice", "Device",
     "Register a new device in the system (hardware onboarding).\nUsed on: Inventory → Hardware Inventory tab (Add Device form)",
     "deviceName: String!\nserialNumber: String!\nmodel: String!\n(+ location, customerId, firmwareVersion, lat, lng depending on resolver)",
     "DeviceType",
     'mutation createDevice($deviceName: String!, $serialNumber: String!, $model: String!) { createDevice(deviceName: $deviceName, serialNumber: $serialNumber, model: $model) { id deviceName serialNumber status } }',
     '{ "deviceName": "Test-Device-01", "serialNumber": "SN-TEST-001", "model": "XR-5000" }',
     "",
     "New device defaults to Online status. Resolver may require additional fields."),

    ("M-02", "updateDeviceCoords", "Device",
     "Update GPS coordinates of a device (geocoding from address).\nUsed on: Inventory → Geo Location tab (pin position update)",
     "address: String!\nlat: Float!\nlng: Float!",
     "GeocodedDeviceType\n(id · address · lat · lng)",
     'mutation updateDeviceCoords($address: String!, $lat: Float!, $lng: Float!) { updateDeviceCoords(address: $address, lat: $lat, lng: $lng) { id address lat lng } }',
     '{ "address": "350 Fifth Avenue, New York, NY 10118", "lat": 40.7484, "lng": -73.9967 }',
     "",
     "Called when a device address is set or changed so the map pin moves."),

    ("M-03", "createFirmware", "Firmware",
     "Upload and register a new firmware package. Creates a Pending entry.\nUsed on: Deployment → 'Upload Firmware' button",
     "name: String!\nversion: String!\ndeviceModel: String!\nfileSize?: Int (bytes)\nreleaseDate?: String\nchecksum?: String",
     "FirmwareType",
     'mutation createFirmware($name: String!, $version: String!, $deviceModel: String!, $releaseDate: String, $fileSize: Int, $checksum: String) { createFirmware(name: $name, version: $version, deviceModel: $deviceModel, releaseDate: $releaseDate, fileSize: $fileSize, checksum: $checksum) { id name version deviceModel status } }',
     '{ "name": "XR-5000 Security Patch v2.1", "version": "2.1.0", "deviceModel": "XR-5000", "releaseDate": "2026-03-27", "fileSize": 2048000, "checksum": "sha256:abc123" }',
     "",
     "fileSize is Int (bytes) — not Float.\nNew firmware defaults to Pending.\nResolver may also require: fileName, s3Key, s3Bucket, uploadedBy"),

    ("M-04", "approveFirmware", "Firmware",
     "Approve a firmware package for deployment.\nTransition: Pending → Approved\nUsed on: Deployment → 'Approve' button",
     "id: String!",
     "FirmwareType",
     'mutation approveFirmware($id: String!) { approveFirmware(id: $id) { id name status } }',
     '{ "id": "94688ba5-81b0-4c73-9f2a-5923f0b4b256" }',
     "",
     "Check that firmware is in Pending status before calling. Only valid transition is Pending → Approved."),

    ("M-05", "createServiceOrder", "Service Order",
     "Create a new service / maintenance ticket.\nUsed on: Account & Service page (New Service Order form)",
     "title: String!\ndescription?: String\npriority?: String\nscheduledDate?: String",
     "ServiceOrderType",
     'mutation createServiceOrder($title: String!, $description: String, $priority: String, $scheduledDate: String) { createServiceOrder(title: $title, description: $description, priority: $priority, scheduledDate: $scheduledDate) { id title status priority } }',
     '{ "title": "Router Firmware Upgrade", "description": "Apply latest security patch", "priority": "High", "scheduledDate": "2026-04-01" }',
     "",
     "New orders default to Pending status.\nResolver may also require: serviceType, location, customerId, createdBy, technicianName"),

    ("M-06", "createCompliance", "Compliance",
     "Submit a firmware version for compliance certification review.\nUsed on: Firmware Compliance → 'Submit for Review' button",
     "firmwareId: String!\nfirmwareVersion: String!\ndeviceModel: String!\ncertifications?: [String]\nvulnerabilities?: Int\ndescription?: String",
     "ComplianceType",
     'mutation createCompliance($firmwareId: String!, $firmwareVersion: String!, $deviceModel: String!, $certifications: [String], $vulnerabilities: Int, $description: String) { createCompliance(firmwareId: $firmwareId, firmwareVersion: $firmwareVersion, deviceModel: $deviceModel, certifications: $certifications, vulnerabilities: $vulnerabilities, description: $description) { id firmwareId status certifications } }',
     '{ "firmwareId": "94688ba5-81b0-4c73-9f2a-5923f0b4b256", "firmwareVersion": "2.1.0", "deviceModel": "XR-5000", "certifications": ["HIPAA", "FCC"], "vulnerabilities": 0, "description": "Security patch compliance review" }',
     "",
     "New submissions default to Pending status."),

    ("M-07", "updateEntityStatus", "Cross-Entity",
     "Generic status-transition mutation across multiple entity types.\nHandles: firmware lifecycle (Pending→Approved/Deprecated/Rejected) and compliance transitions.\nUsed on: Deployment 'Approve'/'Deprecate' buttons, Compliance 'Approve'/'Review' buttons",
     'entityType: String! ("firmware" | "compliance")\nid: String!\nnewStatus: String!',
     "AWSJSON\n(raw JSON of updated record)",
     'mutation updateEntityStatus($entityType: String!, $id: String!, $newStatus: String!) { updateEntityStatus(entityType: $entityType, id: $id, newStatus: $newStatus) }',
     '{ "entityType": "firmware", "id": "94688ba5-81b0-4c73-9f2a-5923f0b4b256", "newStatus": "Approved" }',
     'Other examples:\n{ "entityType": "firmware", "id": "98910328-...", "newStatus": "Deprecated" }\n{ "entityType": "compliance", "id": "85a9549f-...", "newStatus": "Approved" }',
     "⚠ KNOWN BUG: No server-side validation of newStatus — any string accepted without error.\n⚠ KNOWN BUG: Non-existent id succeeds (upsert behaviour) instead of 404."),
]

ws_m = wb.create_sheet("GraphQL Mutations")
set_col_widths(ws_m, M_WIDTHS)
write_title(ws_m, len(M_COLS),
    f"HLM Platform — GraphQL Mutations  (7 operations)  |  All requests: POST {GRAPHQL_URL}")
write_col_headers(ws_m, 2, M_COLS, CLR["mut_hdr"])

for r_i, (mid, op, cat, purpose, args, ret, gql_doc, sample_vars, extra_vars, notes) in \
        enumerate(MUTATIONS_RAW, 3):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]

    full_body = gql_body(op, gql_doc, sample_vars)
    if extra_vars:
        full_body += f"\n\n// Additional examples:\n{extra_vars}"

    row_vals = [mid, op, cat, "POST", GRAPHQL_URL, AUTH_VALUE,
                purpose, args, ret, full_body, notes, "Active"]

    for col_i, val in enumerate(row_vals, 1):
        c = ws_m.cell(row=r_i, column=col_i, value=val)
        if col_i == 4:
            style_cell(c, bg=CLR["method_post"], bold=True, h="center")
        elif col_i == 5:
            style_cell(c, bg=CLR["url_bg"], sz=9)
        elif col_i == 6:
            style_cell(c, bg=CLR["auth_bg"], sz=9)
        elif col_i == 10:
            style_cell(c, bg=CLR["white"], sz=9)
        elif col_i == 11 and "⚠" in str(val):
            style_cell(c, bg=CLR["known_bug"])
        elif col_i == 12:
            style_cell(c, bg=CLR["active"], bold=True, h="center")
        else:
            style_cell(c, bg=bg)
    ws_m.row_dimensions[r_i].height = 110

ws_m.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — REST Map Tiles
# ══════════════════════════════════════════════════════════════════════════════
R_COLS   = ["#", "Method", "API URL (Full Path)", "Auth", "Request Body", "Purpose", "Test Status"]
R_WIDTHS = [5, 9, 90, 42, 20, 40, 13]

REST_ROWS = [
    ("R-01", "GET",
     f"{MAP_BASE}/maps/v0/maps/{MAP_NAME}/style-descriptor",
     "AWS Signature v4 via Amplify session credentials\n(added automatically by MapLibre + Amplify SDK)",
     "None — GET request, no body",
     "Fetch map style JSON (colours, layers, fonts). Called once on map init.",
     "Active"),
    ("R-02", "GET",
     f"{MAP_BASE}/maps/v0/maps/{MAP_NAME}/sprites/sprites@2x.png",
     "AWS Signature v4 via Amplify session credentials",
     "None — GET request, no body",
     "Icon sprite sheet (PNG) for map symbols.",
     "Active"),
    ("R-03", "GET",
     f"{MAP_BASE}/maps/v0/maps/{MAP_NAME}/sprites/sprites@2x.json",
     "AWS Signature v4 via Amplify session credentials",
     "None — GET request, no body",
     "Icon sprite coordinate metadata (JSON). Used alongside sprite PNG.",
     "Active"),
    ("R-04", "GET",
     f"{MAP_BASE}/maps/v0/maps/{MAP_NAME}/tiles/{{z}}/{{x}}/{{y}}\n(Example: .../tiles/7/35/48)",
     "AWS Signature v4 via Amplify session credentials",
     "None — GET request, no body",
     "Vector map tile at z/x/y coordinates. Loaded repeatedly on pan and zoom — replace z/x/y with actual tile coordinates.",
     "Active"),
]

ws_rest = wb.create_sheet("REST - Map Tiles")
set_col_widths(ws_rest, R_WIDTHS)
write_title(ws_rest, len(R_COLS),
    f"HLM Platform — AWS Location Service REST Endpoints (MapLibre Map Tiles)  |  Map: {MAP_NAME}")
write_col_headers(ws_rest, 2, R_COLS, CLR["rest_hdr"])

for r_i, row_data in enumerate(REST_ROWS, 3):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]
    for col_i, val in enumerate(row_data, 1):
        c = ws_rest.cell(row=r_i, column=col_i, value=val)
        if col_i == 2:
            style_cell(c, bg=CLR["method_get"], bold=True, h="center")
        elif col_i == 3:
            style_cell(c, bg=CLR["url_bg"], sz=9)
        elif col_i == 4:
            style_cell(c, bg=CLR["auth_bg"], sz=9)
        elif col_i == 7:
            style_cell(c, bg=CLR["active"], bold=True, h="center")
        else:
            style_cell(c, bg=bg)
    ws_rest.row_dimensions[r_i].height = 55

ws_rest.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Page → API Mapping
# ══════════════════════════════════════════════════════════════════════════════
ws_map = wb.create_sheet("Page-API Mapping")
set_col_widths(ws_map, [34, 10, 62, 42, 20])

write_title(ws_map, 5, "HLM Platform — Page → API Call Mapping")
write_col_headers(ws_map, 2,
    ["Page / Route", "Method", "Operations Called on Load", "API URL", "Notes"],
    CLR["section_bg"])

page_map = [
    ("Dashboard  (/)",
     "POST", "getUserByEmail\nlistDevices\nlistFirmware\nlistComplianceByStatus(Approved)\nlistAuditLogs  [24-hour window]",
     GRAPHQL_URL, "5 operations on load"),
    ("Inventory — Hardware Inventory",
     "POST", "listDevices",
     GRAPHQL_URL, "Status filter optional"),
    ("Inventory — Firmware Status",
     "POST", "listFirmware\ngetFirmwareByModel",
     GRAPHQL_URL, "Model filter triggers getFirmwareByModel"),
    ("Inventory — Geo Location",
     "POST / GET", "listDevices  (POST → AppSync)\nAWS Location Service map tiles  (GET → see REST sheet)",
     GRAPHQL_URL, "listDevices populates pins; tiles loaded by MapLibre SDK"),
    ("Account & Service  (/account-service)",
     "POST", "listServiceOrdersByStatus\nlistUsersByRole",
     GRAPHQL_URL, "⚠ listUsersByRole filter not applied server-side"),
    ("Deployment — Firmware Versions",
     "POST", "listFirmware",
     GRAPHQL_URL, ""),
    ("Deployment — Audit Log",
     "POST", "listAuditLogs  [limit: 50]",
     GRAPHQL_URL, ""),
    ("Firmware Compliance  (/compliance)",
     "POST", "listComplianceByStatus\nlistFirmware",
     GRAPHQL_URL, ""),
    ("Reporting & Analytics  (/analytics)",
     "POST", "listDevices\nlistFirmware\nlistComplianceByStatus\nlistAuditLogs\nlistServiceOrdersByDate",
     GRAPHQL_URL, "5 operations on load"),
]

for r_i, (page, method, ops, url, notes) in enumerate(page_map, 3):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]
    for col_i, val in enumerate([page, method, ops, url, notes], 1):
        c = ws_map.cell(row=r_i, column=col_i, value=val)
        if col_i == 2:
            style_cell(c, bg=CLR["method_post"], bold=True, h="center")
        elif col_i == 4:
            style_cell(c, bg=CLR["url_bg"], sz=9)
        elif col_i == 5 and "⚠" in str(val):
            style_cell(c, bg=CLR["known_bug"])
        else:
            style_cell(c, bg=bg)
    ws_map.row_dimensions[r_i].height = 70

ws_map.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Sample Data & Entity Types
# ══════════════════════════════════════════════════════════════════════════════
ws_data = wb.create_sheet("Sample Data & Types")
set_col_widths(ws_data, [22, 40, 44])

write_title(ws_data, 3,
    "HLM Platform — Live Sample Data & Entity Type Field Reference  (Discovered 2026-03-25)")
write_col_headers(ws_data, 2, ["Entity Type", "ID", "Name / Description"], CLR["section_bg"])

sample_rows = [
    ("Device",        "a11bc9b9-ba1e-4e1f-8523-45cd003a6845", "UPS-POWER-05"),
    ("Device",        "5c8da3fb-…",                           "Switch-CORE-01"),
    ("Firmware",      "94688ba5-81b0-4c73-9f2a-5923f0b4b256", "XR-5000 Security Patch  [status: Pending]"),
    ("Firmware",      "98910328-e3cb-4d03-8690-5212c30ccece", "Dell SRV-9000 Firmware  [status: Rejected]"),
    ("Compliance",    "85a9549f-…",                           "HIPAA-certified  [status: Approved]"),
    ("Service Order", "acf1df53-cf21-476d-9410-251efbf04b37", "Core Switch Firmware Upgrade  [status: Completed]"),
    ("User",          "u001",                                  "Alice Turner — alice@acmecorp.com  [role: Admin]"),
    ("User",          "u002",                                  "Bob Nguyen — bob@acmecorp.com  [role: Manager]"),
    ("User",          "u003",                                  "Carol Smith — carol@acmecorp.com  [role: Technician]"),
    ("Customer",      "c001 – c005",                          "Seeded demo customers"),
]

for r_i, (entity, id_, desc) in enumerate(sample_rows, 3):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]
    for col_i, val in enumerate([entity, id_, desc], 1):
        c = ws_data.cell(row=r_i, column=col_i, value=val)
        style_cell(c, bg=bg)
    ws_data.row_dimensions[r_i].height = 22

gap = len(sample_rows) + 4
ws_data.merge_cells(f"A{gap}:C{gap}")
c = ws_data.cell(row=gap, column=1, value="Entity Type — Confirmed Field Reference (via GraphQL introspection)")
style_cell(c, bg=CLR["section_bg"], bold=True, fg=CLR["section_fg"])

write_col_headers(ws_data, gap + 1, ["Type", "Confirmed Fields", "Notes"], CLR["section_bg"])

entity_types = [
    ("DeviceType",         "id · deviceName · serialNumber · model · status · location · lat · lng · customerId · firmwareVersion", ""),
    ("FirmwareType",       "id · name · version · deviceModel · status · releaseDate · fileSize · checksum", "fileSize is Int (bytes)"),
    ("ServiceOrderType",   "id · title · description · status · technicianId · scheduledDate · priority", ""),
    ("ComplianceType",     "id · firmwareId · firmwareVersion · deviceModel · status · certifications · vulnerabilities", "certifications is AWSJSON array — parse client-side"),
    ("AuditLogType",       "id · userId · userEmail · action · auditStatus · resourceType · resourceId", ""),
    ("UserType",           "id · email · role · firstName · lastName", ""),
    ("GeocodedDeviceType", "id · address · lat · lng", "Returned by updateDeviceCoords mutation"),
    ("PaginatedResponse",  "items · nextToken · totalCount",
     "items is AWSJSON scalar — cannot sub-select fields in GQL; parse in client.\nField is totalCount NOT total."),
]

for r_i, (typ, fields, notes) in enumerate(entity_types, gap + 2):
    bg = CLR["white"] if r_i % 2 == 1 else CLR["alt_row"]
    for col_i, val in enumerate([typ, fields, notes], 1):
        c = ws_data.cell(row=r_i, column=col_i, value=val)
        style_cell(c, bg=bg)
    ws_data.row_dimensions[r_i].height = 30

ws_data.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Known Backend Bugs
# ══════════════════════════════════════════════════════════════════════════════
ws_bugs = wb.create_sheet("Known Backend Bugs")
set_col_widths(ws_bugs, [28, 50, 44])

write_title(ws_bugs, 3, "HLM Platform — Known Backend Limitations / Bugs")
write_col_headers(ws_bugs, 2, ["Operation", "Limitation / Bug", "Impact & Suggested Test"], CLR["known_bug"])

bugs = [
    ('listFirmware(status: "Deprecated")',
     'Filter not applied server-side — returns ALL firmware records regardless of status value',
     'Impact: cannot isolate deprecated firmware via this query.\n'
     'Test: pass "Deprecated" and verify totalCount equals ALL firmware count.\n'
     'Workaround: filter client-side after fetching all records.'),
    ('listUsersByRole(role: any)',
     'role parameter is completely ignored server-side — all users returned regardless of value',
     'Impact: technician dropdown always shows all users.\n'
     'Test: pass role="NonExistentRole" and verify users are still returned.\n'
     'Workaround: filter client-side by role field in returned items.'),
    ('updateEntityStatus — newStatus validation',
     'No server-side enum validation on newStatus — any arbitrary string is accepted without error',
     'Impact: typos/junk strings are silently persisted.\n'
     'Test: pass newStatus="GARBAGE_VALUE" and verify 200 response with no error.\n'
     'Risk: data corruption if callers pass invalid status values.'),
    ('updateEntityStatus — id existence check',
     'No existence check on id — non-existent IDs succeed with upsert behaviour instead of returning 404',
     'Impact: calling with a random UUID creates or overwrites a record unexpectedly.\n'
     'Test: pass a random UUID and verify the mutation succeeds.\n'
     'Risk: phantom records can be created.'),
]

for r_i, (op, bug, impact) in enumerate(bugs, 3):
    for col_i, val in enumerate([op, bug, impact], 1):
        c = ws_bugs.cell(row=r_i, column=col_i, value=val)
        style_cell(c, bg=CLR["known_bug"])
    ws_bugs.row_dimensions[r_i].height = 70

ws_bugs.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
OUTPUT = "/Users/ajaykumar.yadav/HLM-QA/HLM-API-Catalogue.xlsx"
wb.save(OUTPUT)
print(f"✓  Saved: {OUTPUT}")
print(f"   Sheets : {', '.join(wb.sheetnames)}")
print(f"   Queries: {len(QUERIES_RAW)}  |  Mutations: {len(MUTATIONS_RAW)}  |  REST: {len(REST_ROWS)}")
